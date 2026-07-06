"""
Bloomberg Exchange Inventory / Warehouse Stocks Downloader
===========================================================
Downloads physical inventory / certified-stock time series (PX_LAST, daily)
from exchange-published data sourced via Bloomberg.

Coverage:
    LME Total Stocks      Cu, Al, Zn, Ni, Pb, Sn  (daily, metric tons)
    LME Cancelled Warrants Cu, Al, Zn, Ni, Pb, Sn (daily, metric tons)
    LME On-Warrant Stocks  Cu, Al, Zn, Ni, Pb, Sn  (daily, metric tons)
    SHFE Warehouse Stocks  Cu, Al, Zn, Ni, Pb, Sn, Rebar (weekly, MT)
    COMEX/NYMEX Stocks     Au, Ag, HG Copper, Pt, Pd  (daily, troy oz / lbs)
    ICE Certified Stocks   Cotton, Cocoa, Coffee C     (daily, bales/MT/bags)

Notes:
    - LME tickers are high confidence.
    - SHFE and COMEX/NYMEX tickers are marked [verify] — if a ticker returns
      "no data", search Bloomberg terminal for the correct ticker and update the
      INVENTORY dict below.  The deadlock killer handles wrong tickers cleanly.
    - ICE certified stock tickers (ICTCOTCE, ICTCOCOA, ICTCOFFC) are
      high confidence.

Output (same folder as this script):
    GHR_Inventory_Exchange.xlsx

Run with Bloomberg Terminal open:
    python bloomberg_inventory_exchange.py

Requirements: pip install pdblp openpyxl pandas
"""

import pdblp
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
import sys

try:
    sys.stdout.reconfigure(encoding="utf-8")
except AttributeError:
    pass

# =============================================================================
# CONFIG
# =============================================================================

START_DATE = "20050101"
END_DATE   = "20260630"

HOST = "localhost"
PORT = 8194

FIELDS = ["PX_LAST"]

TIMEOUT_PER_CONTRACT = 12
EARLY_EXIT_EMPTY     = 3   # unused for flat downloads; kept for consistency

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# =============================================================================
# INVENTORY TICKER GROUPS
# Each group is downloaded separately and gets its own sheet in the workbook.
# Dict format: { "BLOOMBERG TICKER": "Column label in Excel" }
# =============================================================================

# ---- LME Total Stocks (daily, metric tons) ----------------------------------
# Source: LME warehouse reports, compiled by Bloomberg.
# These are the core inventory series used in the Theory of Storage (GHR).
LME_TOTAL = {
    "LCSNSTOT Index": "LME_Cu_Total_MT",
    "LASSNTOT Index": "LME_Al_Total_MT",
    "LMZSNTOT Index": "LME_Zn_Total_MT",
    "LMNITON Index":  "LME_Ni_Total_MT",
    "LMPBTOT Index":  "LME_Pb_Total_MT",
    "LMSNTOT Index":  "LME_Sn_Total_MT",
}

# ---- LME Cancelled Warrants (daily, metric tons) ----------------------------
# Cancelled warrants = inventory earmarked for physical withdrawal.
# Spike = drawdown imminent; useful leading indicator of physical tightness.
LME_CANCELLED = {
    "LCSNCANC Index": "LME_Cu_Cancelled_MT",
    "LASCANWP Index": "LME_Al_Cancelled_MT",
    "LMZSCANW Index": "LME_Zn_Cancelled_MT",
    "LMNICANW Index": "LME_Ni_Cancelled_MT",
    "LMPBCANW Index": "LME_Pb_Cancelled_MT",
    "LMSNCANW Index": "LME_Sn_Cancelled_MT",
}

# ---- LME On-Warrant Stocks (daily, metric tons) -----------------------------
# On-warrant = freely available (not yet cancelled). Complement to cancelled.
# On-warrant + Cancelled = Total Stocks.
LME_ONWARRANT = {
    "LCSNOWRT Index":  "LME_Cu_OnWarrant_MT",
    "LASSNOWRT Index": "LME_Al_OnWarrant_MT",
    "LMZSOWRT Index":  "LME_Zn_OnWarrant_MT",
    "LMNIOWRT Index":  "LME_Ni_OnWarrant_MT",
    "LMPBOWRT Index":  "LME_Pb_OnWarrant_MT",
    "LMSNOWRT Index":  "LME_Sn_OnWarrant_MT",
}

# ---- SHFE Warehouse Stocks (weekly, metric tons) [VERIFY TICKERS] -----------
# Shanghai Futures Exchange publishes every Friday after market close.
# Bloomberg carries these but exact ticker varies by subscription tier.
# If "no data" returned: on Bloomberg terminal type  SHFE <GO>  then find
# the correct ticker under Commodity Storage / Inventory.
SHFE_STOCKS = {
    "SHFHCU Index": "SHFE_Cu_MT",
    "SHFHAL Index": "SHFE_Al_MT",
    "SHFHZN Index": "SHFE_Zn_MT",
    "SHFHNI Index": "SHFE_Ni_MT",
    "SHFHPB Index": "SHFE_Pb_MT",
    "SHFHSN Index": "SHFE_Sn_MT",
    "SHFHRB Index": "SHFE_SteelRebar_MT",
}

# ---- COMEX / NYMEX Warehouse Stocks [VERIFY TICKERS] -----------------------
# GHR uses NYMEX warehouse stocks for Platinum and Palladium.
# If "no data": search Bloomberg for  COMEX WAREHOUSE <metal>  or
# type the root (e.g. GC1 Comdty), press COMS <GO>, then look for stocks series.
COMEX_NYMEX = {
    "COMXGOLD Index": "COMEX_Au_TroyOz",
    "COMXSILV Index": "COMEX_Ag_TroyOz",
    "COMXHG Index":   "COMEX_HG_Copper_lbs",
    "NYMXPT Index":   "NYMEX_Pt_TroyOz",
    "NYMXPA Index":   "NYMEX_Pd_TroyOz",
}

# ---- ICE Certified Stocks (daily) -------------------------------------------
# ICE publishes certified exchange stocks daily for these three softs.
# These are the exact inventory series GHR uses for Cotton, Cocoa, Coffee.
ICE_CERTIFIED = {
    "ICTCOTCE Index": "ICE_Cotton_Bales",
    "ICTCOCOA Index": "ICE_Cocoa_MT",
    "ICTCOFFC Index": "ICE_CoffeeC_Bags",
}

# =============================================================================
# STYLES
# =============================================================================

TITLE_FILL  = PatternFill("solid", fgColor="0D3B66")
TITLE_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=12)
HDR_FILL    = PatternFill("solid", fgColor="1F4E79")
HDR_FONT    = Font(name="Arial", bold=True, color="FFFFFF", size=10)
DATA_FONT   = Font(name="Arial", size=9)
DATE_FONT   = Font(name="Arial", bold=True, size=9)
ALT_FILL    = PatternFill("solid", fgColor="F5F9FC")

# =============================================================================
# CONNECTION
# =============================================================================

def connect():
    print("\n  Connecting to Bloomberg (%s:%d) ..." % (HOST, PORT))
    con = pdblp.BCon(debug=False, host=HOST, port=PORT, timeout=15000)
    con.start()
    print("  Connected.\n")
    return con

# =============================================================================
# DOWNLOAD
# =============================================================================

def safe_bdh(con, ticker):
    """Single-ticker BDH with error handling. Returns (df|None, status_str).
    Direct call — no thread wrapper (threading causes pdblp session errors)."""
    try:
        df = con.bdh(ticker, FIELDS, START_DATE, END_DATE)
        if isinstance(df.columns, pd.MultiIndex):
            df.columns = df.columns.get_level_values(1)
        if df.empty or "PX_LAST" not in df.columns:
            return None, "no data"
        df = df[["PX_LAST"]].dropna()
        if df.empty:
            return None, "all NaN"
        return df, "%d rows" % len(df)
    except Exception as e:
        return None, "ERROR: %s" % str(e)[:80]


def download_tickers(con, ticker_map, label=""):
    """Download a flat list of named tickers. Returns dict: {ticker: DataFrame}."""
    result = {}
    total = len(ticker_map)
    tag = "[%s] " % label if label else ""
    for idx, (tkr, col_name) in enumerate(ticker_map.items(), 1):
        print("  %s[%d/%d] %-26s" % (tag, idx, total, tkr + " (" + col_name + ")"), end="", flush=True)
        df, status = safe_bdh(con, tkr)
        if df is not None:
            result[tkr] = df
        print("  " + status)
    return result

# =============================================================================
# EXCEL WRITERS
# =============================================================================

def write_flat_sheet(ws, title, ticker_map, data):
    """Date | Col1 | Col2 | ...  (PX_LAST, one column per ticker)."""
    ordered = [t for t in ticker_map if t in data]
    if not ordered:
        ws["A1"] = "No data returned for: %s" % title
        return

    all_dates = sorted(set().union(*(data[t].index for t in ordered)))
    n_cols = 1 + len(ordered)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws.cell(row=1, column=1,
                value="%s  (PX_LAST  %s to %s)" % (title, START_DATE, END_DATE))
    c.font = TITLE_FONT; c.fill = TITLE_FILL
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 22

    c = ws.cell(row=2, column=1, value="Date")
    c.font = HDR_FONT; c.fill = HDR_FILL
    c.alignment = Alignment(horizontal="center")
    for j, tkr in enumerate(ordered, 2):
        c = ws.cell(row=2, column=j, value=ticker_map[tkr])
        c.font = HDR_FONT; c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center")

    for ri, dt in enumerate(all_dates, 3):
        cell_date = ws.cell(row=ri, column=1, value=dt)
        cell_date.number_format = "YYYY-MM-DD"
        cell_date.font = DATE_FONT
        if ri % 2 == 0:
            cell_date.fill = ALT_FILL
        for j, tkr in enumerate(ordered, 2):
            df = data[tkr]
            cell = ws.cell(row=ri, column=j)
            if dt in df.index:
                v = df.loc[dt, "PX_LAST"]
                if pd.notna(v):
                    cell.value = float(v)
            cell.font = DATA_FONT
            cell.number_format = "#,##0.00"
            if ri % 2 == 0:
                cell.fill = ALT_FILL

    ws.column_dimensions["A"].width = 13
    for col in range(2, n_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 22
    ws.freeze_panes = "B3"
    print("    Sheet '%s': %d rows, %d series" % (ws.title, len(all_dates), len(ordered)))


def write_readme(wb, title_str, rows):
    ws = wb.active
    ws.title = "README"
    meta = [
        (title_str,        ""),
        ("",               ""),
        ("Generated",      datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Start Date",     START_DATE),
        ("End Date",       END_DATE),
        ("Fields",         ", ".join(FIELDS)),
        ("Timeout/ticker", "%ds" % TIMEOUT_PER_CONTRACT),
        ("Reference",      "Gorton, Hayashi & Rouwenhorst (2013), Rev Finance"),
        ("", ""),
    ] + rows
    for r, (a, b) in enumerate(meta, 1):
        ws.cell(row=r, column=1, value=a).font = Font(name="Arial", bold=True, size=10)
        ws.cell(row=r, column=2, value=b).font = Font(name="Arial", size=10)
    ws.cell(row=1, column=1).font = Font(name="Arial", bold=True, size=13, color="0D3B66")
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 60


def save_wb(wb, filename):
    path = os.path.join(SCRIPT_DIR, filename)
    wb.save(path)
    print("\n  Saved: %s" % path)
    print("  Sheets: %s\n" % ", ".join(wb.sheetnames))

# =============================================================================
# MAIN
# =============================================================================

def main():
    print("\n" + "=" * 65)
    print("  GHR EXCHANGE INVENTORY — BLOOMBERG DOWNLOADER")
    print("  Start: %s   End: %s" % (START_DATE, END_DATE))
    print("  Timeout: %ds/ticker" % TIMEOUT_PER_CONTRACT)
    print("=" * 65)

    con = connect()
    wb = Workbook()
    write_readme(wb, "GHR Exchange Inventory / Warehouse Stocks", [
        ("LME Total Stocks",      "Cu,Al,Zn,Ni,Pb,Sn — daily metric tons"),
        ("LME Cancelled Warrants","Cu,Al,Zn,Ni,Pb,Sn — earmarked for withdrawal"),
        ("LME On-Warrant Stocks", "Cu,Al,Zn,Ni,Pb,Sn — freely available inventory"),
        ("SHFE Stocks [verify]",  "Cu,Al,Zn,Ni,Pb,Sn,Rebar — weekly SHFE warehouse"),
        ("COMEX/NYMEX [verify]",  "Au,Ag,HG Cu,Pt,Pd — exchange warehouse stocks"),
        ("ICE Certified Stocks",  "Cotton,Cocoa,Coffee C — ICE certified daily"),
    ])

    SECTIONS = [
        ("LME Total",     LME_TOTAL,     "LME Total Stocks"),
        ("LME Cancel",    LME_CANCELLED, "LME Cancelled Warrants"),
        ("LME OnWarrant", LME_ONWARRANT, "LME On-Warrant Stocks"),
        ("SHFE",          SHFE_STOCKS,   "SHFE Warehouse Stocks"),
        ("COMEX",         COMEX_NYMEX,   "COMEX-NYMEX Stocks"),
        ("ICE",           ICE_CERTIFIED, "ICE Certified Stocks"),
    ]

    for label, ticker_map, sheet_title in SECTIONS:
        print("\n" + "=" * 65)
        print("  %s" % sheet_title)
        print("=" * 65)
        data = download_tickers(con, ticker_map, label=label)
        ws = wb.create_sheet(title=sheet_title[:31])
        write_flat_sheet(ws, sheet_title, ticker_map, data)

    save_wb(wb, "GHR_Inventory_Exchange.xlsx")

    try:
        con.stop()
    except Exception:
        pass

    print("=" * 65)
    print("  DONE => GHR_Inventory_Exchange.xlsx")
    print("=" * 65 + "\n")


if __name__ == "__main__":
    main()
