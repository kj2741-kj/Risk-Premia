"""
Bloomberg Govt / Agency Inventory Data Downloader
==================================================
Downloads government and agency physical inventory / stock series (PX_LAST)
sourced via Bloomberg Index tickers.

Coverage:
    EIA Energy Stocks     US crude oil, distillate, gasoline, propane,
                          nat gas storage, coal stocks — published weekly
                          by US Energy Information Administration.
    USDA Grain Stocks     US wheat, corn, soybean stocks — published quarterly
                          by USDA in the Grain Stocks Report.
    USDA Other            US cotton, peanut stocks (if available via Bloomberg).
    Macro Cross-Reference DXY, US 10Y yield, BCOM, GSCI — for normalised
                          inventory regression controls (GHR Appendix).

Notes:
    - EIA data is published weekly (Wednesdays for petroleum, Thursdays for
      nat gas). Bloomberg carries these as Index tickers with daily frequency
      (value held constant between releases).
    - USDA data is quarterly (March, June, September, December). Bloomberg
      carries these as point-in-time macro releases.
    - Tickers marked [verify] should be confirmed on Bloomberg terminal if
      "no data" is returned.

Output (same folder as this script):
    GHR_Inventory_Govt.xlsx

Run with Bloomberg Terminal open:
    python bloomberg_inventory_govt.py

Requirements: pip install pdblp openpyxl pandas
"""

import pdblp
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
import sys

try:
    sys.stdout.reconfigure(encoding="utf-8")
except AttributeError:
    pass

# =============================================================================
# CONFIG
# =============================================================================

START_DATE = "20050101"
END_DATE   = "20260630"

HOST = "localhost"
PORT = 8194

FIELDS = ["PX_LAST"]

TIMEOUT_PER_CONTRACT = 12

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# =============================================================================
# INVENTORY TICKER GROUPS
# =============================================================================

# ---- EIA Energy Stocks via Bloomberg (weekly, US) ---------------------------
# Source: US Energy Information Administration via Bloomberg Index tickers.
# These are the exact inventory series GHR uses for all 6 energy commodities.
# GHR energies: Crude Oil (CL), Heating Oil (HO), Unleaded Gas (XB),
#               Propane (PN), Natural Gas (NG), Coal (QL).
EIA_ENERGY = {
    "DOESCRUS Index": "EIA_Crude_Oil_excl_SPR_000bbl",   # Weekly Petroleum Status Report
    "DOESDS1U Index": "EIA_Distillate_FuelOil_000bbl",    # Heating oil proxy
    "DOESG1U Index":  "EIA_Motor_Gasoline_000bbl",         # Unleaded gas proxy
    "DOESPP1U Index": "EIA_Propane_Propylene_000bbl",
    "DOENUSGW Index": "EIA_NatGas_Underground_Bcf",        # Weekly Natural Gas Storage
    "DOECSTTT Index": "EIA_Coal_Stocks_Total_000ST",       # Monthly coal stocks [verify]
}

# ---- USDA Grain Stocks via Bloomberg (quarterly release) --------------------
# Source: USDA Grain Stocks Report (released ~March 30, June 30, Sept 30, Dec 1).
# These are the inventory series GHR uses for all 6 grain commodities.
# Bloomberg carries these as macro data — value updated on release date.
USDA_GRAINS = {
    "USSTWT Index":    "USDA_US_Wheat_Stocks_mnBu",
    "USSTCORN Index":  "USDA_US_Corn_Stocks_mnBu",
    "USSTSBNS Index":  "USDA_US_Soybean_Stocks_mnBu",
    "USSTSOBO Index":  "USDA_US_SoybeanOil_Stocks_mnlbs",  # [verify]
    "USSTOATS Index":  "USDA_US_Oats_Stocks_mnBu",          # [verify]
}

# ---- USDA Other Agricultural Stocks via Bloomberg [verify] ------------------
# Cotton: USDA Cotton Ginnings / EWG data. OJ: cold storage.
# These are less standardized on Bloomberg; verify tickers on terminal.
USDA_OTHER = {
    "USSTCTN Index":   "USDA_US_Cotton_Stocks_000bales",    # [verify]
    "USSTSUGR Index":  "USDA_US_Sugar_Stocks",              # [verify]
}

# ---- Macro Cross-Reference --------------------------------------------------
# Used as controls in GHR normalised-inventory regressions.
# Also useful for risk-premia decomposition (dollar, rates, risk appetite).
MACRO_CONTROLS = {
    "DXY Curncy":     "DXY_USD_Index",
    "USGG10YR Index": "US_10Y_Treasury_Yield_pct",
    "USGG2YR Index":  "US_2Y_Treasury_Yield_pct",
    "BCOM Index":     "BCOM_Bloomberg_Commodity_Index",
    "GSCI Index":     "SP_GSCI_Commodity_Index",
    "VIX Index":      "CBOE_VIX",
    "CRY Index":      "CRB_Reuters_Commodity_Index",
    "SPGSCI Index":   "SP_GSCI_Total_Return",               # [verify ticker]
}

# =============================================================================
# STYLES
# =============================================================================

TITLE_FILL  = PatternFill("solid", fgColor="0D3B66")
TITLE_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=12)
HDR_FILL    = PatternFill("solid", fgColor="1F4E79")
HDR_FONT    = Font(name="Arial", bold=True, color="FFFFFF", size=10)
DATA_FONT   = Font(name="Arial", size=9)
DATE_FONT   = Font(name="Arial", bold=True, size=9)
ALT_FILL    = PatternFill("solid", fgColor="F5F9FC")

# =============================================================================
# CONNECTION
# =============================================================================

def connect():
    print("\n  Connecting to Bloomberg (%s:%d) ..." % (HOST, PORT))
    con = pdblp.BCon(debug=False, host=HOST, port=PORT, timeout=15000)
    con.start()
    print("  Connected.\n")
    return con

# =============================================================================
# DOWNLOAD
# =============================================================================

def safe_bdh(con, ticker):
    """Single-ticker BDH with error handling. Returns (df|None, status_str).
    Direct call — no thread wrapper (threading causes pdblp session errors)."""
    try:
        df = con.bdh(ticker, FIELDS, START_DATE, END_DATE)
        if isinstance(df.columns, pd.MultiIndex):
            df.columns = df.columns.get_level_values(1)
        if df.empty or "PX_LAST" not in df.columns:
            return None, "no data"
        df = df[["PX_LAST"]].dropna()
        if df.empty:
            return None, "all NaN"
        return df, "%d rows" % len(df)
    except Exception as e:
        return None, "ERROR: %s" % str(e)[:80]


def download_tickers(con, ticker_map, label=""):
    """Download a flat list of named tickers. Returns dict: {ticker: DataFrame}."""
    result = {}
    total = len(ticker_map)
    tag = "[%s] " % label if label else ""
    for idx, (tkr, col_name) in enumerate(ticker_map.items(), 1):
        print("  %s[%d/%d] %-26s" % (tag, idx, total, tkr + " (" + col_name + ")"), end="", flush=True)
        df, status = safe_bdh(con, tkr)
        if df is not None:
            result[tkr] = df
        print("  " + status)
    return result

# =============================================================================
# EXCEL WRITERS
# =============================================================================

def write_flat_sheet(ws, title, ticker_map, data):
    """Date | Col1 | Col2 | ...  (PX_LAST, one column per ticker)."""
    ordered = [t for t in ticker_map if t in data]
    if not ordered:
        ws["A1"] = "No data returned for: %s" % title
        return

    all_dates = sorted(set().union(*(data[t].index for t in ordered)))
    n_cols = 1 + len(ordered)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws.cell(row=1, column=1,
                value="%s  (PX_LAST  %s to %s)" % (title, START_DATE, END_DATE))
    c.font = TITLE_FONT; c.fill = TITLE_FILL
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 22

    c = ws.cell(row=2, column=1, value="Date")
    c.font = HDR_FONT; c.fill = HDR_FILL
    c.alignment = Alignment(horizontal="center")
    for j, tkr in enumerate(ordered, 2):
        c = ws.cell(row=2, column=j, value=ticker_map[tkr])
        c.font = HDR_FONT; c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center")

    for ri, dt in enumerate(all_dates, 3):
        cell_date = ws.cell(row=ri, column=1, value=dt)
        cell_date.number_format = "YYYY-MM-DD"
        cell_date.font = DATE_FONT
        if ri % 2 == 0:
            cell_date.fill = ALT_FILL
        for j, tkr in enumerate(ordered, 2):
            df = data[tkr]
            cell = ws.cell(row=ri, column=j)
            if dt in df.index:
                v = df.loc[dt, "PX_LAST"]
                if pd.notna(v):
                    cell.value = float(v)
            cell.font = DATA_FONT
            cell.number_format = "#,##0.000"
            if ri % 2 == 0:
                cell.fill = ALT_FILL

    ws.column_dimensions["A"].width = 13
    for col in range(2, n_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 28
    ws.freeze_panes = "B3"
    print("    Sheet '%s': %d rows, %d series" % (ws.title, len(all_dates), len(ordered)))


def write_readme(wb, title_str, rows):
    ws = wb.active
    ws.title = "README"
    meta = [
        (title_str,        ""),
        ("",               ""),
        ("Generated",      datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Start Date",     START_DATE),
        ("End Date",       END_DATE),
        ("Fields",         ", ".join(FIELDS)),
        ("Timeout/ticker", "%ds" % TIMEOUT_PER_CONTRACT),
        ("Reference",      "Gorton, Hayashi & Rouwenhorst (2013), Rev Finance — Appendix B1"),
        ("", ""),
    ] + rows
    for r, (a, b) in enumerate(meta, 1):
        ws.cell(row=r, column=1, value=a).font = Font(name="Arial", bold=True, size=10)
        ws.cell(row=r, column=2, value=b).font = Font(name="Arial", size=10)
    ws.cell(row=1, column=1).font = Font(name="Arial", bold=True, size=13, color="0D3B66")
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 65


def save_wb(wb, filename):
    path = os.path.join(SCRIPT_DIR, filename)
    wb.save(path)
    print("\n  Saved: %s" % path)
    print("  Sheets: %s\n" % ", ".join(wb.sheetnames))

# =============================================================================
# MAIN
# =============================================================================

def main():
    print("\n" + "=" * 65)
    print("  GHR GOVT INVENTORY — BLOOMBERG DOWNLOADER")
    print("  Start: %s   End: %s" % (START_DATE, END_DATE))
    print("  Timeout: %ds/ticker" % TIMEOUT_PER_CONTRACT)
    print("=" * 65)
    print("\n  NOTE: EIA data is weekly; USDA data is quarterly.")
    print("  Tickers marked [verify] in the dict: update if 'no data' returned.\n")

    con = connect()
    wb = Workbook()
    write_readme(wb, "GHR Govt / Agency Inventory Data — via Bloomberg", [
        ("EIA Energy Stocks",   "DOE weekly: crude, distillate, gasoline, propane, nat gas, coal"),
        ("USDA Grain Stocks",   "USDA quarterly: wheat, corn, soybeans, soybean oil, oats"),
        ("USDA Other",          "USDA cotton, sugar stocks [verify tickers on terminal]"),
        ("Macro Controls",      "DXY, US 10Y, VIX, BCOM, GSCI — regression controls"),
        ("",                    ""),
        ("If empty [verify]",   "For missing tickers: search Bloomberg SRCH <GO> or type"),
        ("",                    "e.g. 'EIA crude oil inventories' in Bloomberg Search"),
        ("Alternative (EIA)",   "EIA API available free at https://api.eia.gov/v2/"),
        ("Alternative (USDA)",  "USDA NASS API: https://quickstats.nass.usda.gov/api"),
    ])

    SECTIONS = [
        ("EIA",   EIA_ENERGY,    "EIA Energy Stocks (DOE)"),
        ("USDA",  USDA_GRAINS,   "USDA Grain Stocks"),
        ("USDA2", USDA_OTHER,    "USDA Other Agri Stocks"),
        ("MACRO", MACRO_CONTROLS,"Macro Cross-Reference"),
    ]

    for label, ticker_map, sheet_title in SECTIONS:
        print("\n" + "=" * 65)
        print("  %s" % sheet_title)
        print("=" * 65)
        data = download_tickers(con, ticker_map, label=label)
        ws = wb.create_sheet(title=sheet_title[:31])
        write_flat_sheet(ws, sheet_title, ticker_map, data)

    save_wb(wb, "GHR_Inventory_Govt.xlsx")

    try:
        con.stop()
    except Exception:
        pass

    print("=" * 65)
    print("  DONE => GHR_Inventory_Govt.xlsx")
    print("=" * 65 + "\n")


if __name__ == "__main__":
    main()
