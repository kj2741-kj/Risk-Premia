"""
Bloomberg GHR Replication Data Downloader
==========================================
Downloads all data needed to replicate Gorton, Hayashi & Rouwenhorst (2013)
"The Fundamentals of Commodity Futures Returns" — extended to June 30, 2026.

SECTION 1  Futures prices (F1-Fn, PX_LAST) for all 31 GHR commodities
           Minimum F1 + F2 required for basis and excess-return calculations
           Groups: LME Metals, Precious, Grains, Softs, Meats, Energies

SECTION 2  Physical inventory / warehouse stocks (PX_LAST, daily or weekly)
           — LME warehouse stocks + cancelled warrants (Cu, Al, Zn, Ni, Pb, Sn)
           — SHFE warehouse stocks (Cu, Al, Zn, Ni, Pb, Sn, Rebar)
           — COMEX/NYMEX warehouse stocks (Au, Ag, HG copper, Pt, Pd)
           — ICE certified stocks (Cotton, Cocoa, Coffee C)
           — EIA energy stocks via Bloomberg (Crude, Distillate, Gasoline,
             Propane, Nat Gas, Coal)
           — USDA grain stocks via Bloomberg (Wheat, Corn, Soybeans)

Deadlock protection: 12s per-ticker timeout + early-exit after 3 consecutive empties.

Output files (saved to same folder as this script):
  GHR_Futures_Metals.xlsx
  GHR_Futures_Grains.xlsx
  GHR_Futures_Softs.xlsx
  GHR_Futures_Meats.xlsx
  GHR_Futures_Energies.xlsx
  GHR_Inventory_Exchange.xlsx   (LME, SHFE, COMEX/NYMEX, ICE certified)
  GHR_Inventory_Govt.xlsx       (EIA energy, USDA grain stocks via Bloomberg)

Ticker notes:
  - SHFE tickers marked [verify] — confirm on Bloomberg terminal if empty
  - COMEX/NYMEX warehouse tickers marked [verify] — confirm if empty
  - EIA/USDA Bloomberg Index tickers: widely used but check if your subscription includes them

Run with Bloomberg Terminal open:
    python bloomberg_ghr_replication.py

Requirements:
    pip install pdblp openpyxl pandas
"""

import pdblp
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from concurrent.futures import ThreadPoolExecutor, TimeoutError as FuturesTimeoutError
from collections import defaultdict
from datetime import datetime
import os
import sys

try:
    sys.stdout.reconfigure(encoding="utf-8")
except AttributeError:
    pass

# =============================================================================
# CONFIG
# =============================================================================

START_DATE = "20050101"
END_DATE   = "20260630"       # June 30, 2026

HOST = "localhost"
PORT = 8194

FIELDS = ["PX_LAST"]

TIMEOUT_PER_TICKER = 12       # seconds hard cap per Bloomberg call
EARLY_EXIT_EMPTY   = 3        # stop futures strip after N consecutive empties

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# =============================================================================
# SECTION 1 — FUTURES STRIPS
# Ticker format: {ROOT}{n} Comdty  e.g. W1 Comdty, W2 Comdty ... W12 Comdty
# max_contracts: how many generic contracts to attempt; deadlock-killer stops
#                early if the back months are illiquid/unlisted.
# For GHR you need F1 and F2 at minimum. More contracts give the full curve.
# =============================================================================

GHR_FUTURES = {

    # ---- LME Base Metals (F1-F27) -------------------------------------------
    # Used in GHR: Copper, Zinc, Lead, Nickel, Aluminium, Tin
    "LP": {"name": "Copper LME",    "max_contracts": 27, "group": "Metals"},
    "LA": {"name": "Aluminium LME", "max_contracts": 27, "group": "Metals"},
    "LX": {"name": "Zinc LME",      "max_contracts": 27, "group": "Metals"},
    "LN": {"name": "Nickel LME",    "max_contracts": 27, "group": "Metals"},
    "LL": {"name": "Lead LME",      "max_contracts": 27, "group": "Metals"},
    "LT": {"name": "Tin LME",       "max_contracts": 15, "group": "Metals"},

    # ---- Precious Metals (COMEX / NYMEX) ------------------------------------
    # Used in GHR: Platinum (PL), Palladium (PA)
    # Gold (GC) and Silver (SI) excluded by GHR (essentially financial)
    # but included here as useful references
    "PL": {"name": "Platinum NYMEX",   "max_contracts": 13, "group": "Metals"},
    "PA": {"name": "Palladium NYMEX",  "max_contracts": 12, "group": "Metals"},
    "GC": {"name": "Gold COMEX",       "max_contracts": 20, "group": "Metals"},
    "SI": {"name": "Silver COMEX",     "max_contracts": 17, "group": "Metals"},
    "HG": {"name": "Copper CME (HG)",  "max_contracts": 27, "group": "Metals"},

    # ---- CBOT Grains --------------------------------------------------------
    # All 6 used in GHR
    "W":  {"name": "Wheat CBOT",        "max_contracts": 12, "group": "Grains"},
    "C":  {"name": "Corn CBOT",         "max_contracts": 12, "group": "Grains"},
    "S":  {"name": "Soybeans CBOT",     "max_contracts": 12, "group": "Grains"},
    "BO": {"name": "Soybean Oil CBOT",  "max_contracts": 12, "group": "Grains"},
    "SM": {"name": "Soybean Meal CBOT", "max_contracts": 12, "group": "Grains"},
    "O":  {"name": "Oats CBOT",         "max_contracts": 12, "group": "Grains"},

    # ---- ICE / CME Softs ----------------------------------------------------
    # All 5 used in GHR: Cotton, Cocoa, Coffee, Orange Juice, Lumber
    "CT": {"name": "Cotton ICE",       "max_contracts": 12, "group": "Softs"},
    "CC": {"name": "Cocoa ICE",        "max_contracts": 12, "group": "Softs"},
    "KC": {"name": "Coffee C ICE",     "max_contracts": 12, "group": "Softs"},
    "JO": {"name": "Orange Juice ICE", "max_contracts": 12, "group": "Softs"},
    "LB": {"name": "Lumber CME",       "max_contracts": 12, "group": "Softs"},

    # ---- CME Meats ----------------------------------------------------------
    # GHR uses: Live Cattle, Lean Hogs, Feeder Cattle, Pork Bellies (PB
    # delisted 2011), Milk, Butter
    # Note: Pork Bellies (PB) contract was delisted July 2011; include for
    # historical coverage but OOS data will be absent.
    "LC": {"name": "Live Cattle CME",   "max_contracts": 12, "group": "Meats"},
    "LH": {"name": "Lean Hogs CME",     "max_contracts": 12, "group": "Meats"},
    "FC": {"name": "Feeder Cattle CME", "max_contracts": 12, "group": "Meats"},
    "DA": {"name": "Milk Class III CME","max_contracts": 12, "group": "Meats"},
    "CB": {"name": "Butter CME",        "max_contracts": 12, "group": "Meats"},
    "PB": {"name": "Pork Bellies CME",  "max_contracts":  6, "group": "Meats"},

    # ---- NYMEX / ICE Energies -----------------------------------------------
    # GHR uses: Heating Oil, Crude Oil, Unleaded Gas (now RBOB=RB),
    #           Propane, Natural Gas, Coal
    # Brent (CO) not in GHR but included for cross-reference
    "CL": {"name": "WTI Crude NYMEX",         "max_contracts": 27, "group": "Energies"},
    "HO": {"name": "Heating Oil NYMEX",        "max_contracts": 27, "group": "Energies"},
    "RB": {"name": "RBOB Gasoline NYMEX",      "max_contracts": 27, "group": "Energies"},
    "PN": {"name": "Propane NYMEX",            "max_contracts": 12, "group": "Energies"},
    "NG": {"name": "Nat Gas Henry Hub NYMEX",  "max_contracts": 27, "group": "Energies"},
    "QL": {"name": "Coal NYMEX",               "max_contracts": 12, "group": "Energies"},
    "CO": {"name": "Brent Crude ICE",          "max_contracts": 27, "group": "Energies"},
    "QS": {"name": "Singapore Gasoil ICE",     "max_contracts": 27, "group": "Energies"},
    "GO": {"name": "ICE Gasoil London",        "max_contracts": 27, "group": "Energies"},
}

# =============================================================================
# SECTION 2 — INVENTORY / WAREHOUSE STOCK TICKERS (flat BDH download)
# All fetched as daily PX_LAST time series via con.bdh()
# Tickers marked [verify] should be confirmed on Bloomberg terminal;
# the deadlock killer handles no-data cases cleanly.
# =============================================================================

INVENTORY_TICKERS = {

    # ---- LME WAREHOUSE STOCKS (daily, metric tons) --------------------------
    # Source: LME published warehouse inventory, compiled by Bloomberg.
    # High confidence — these are the same tickers used in the paper's
    # copper inventory series (LCSNSTOT) plus the equivalent for each metal.
    "LCSNSTOT Index":  "LME Copper Total Stocks (MT)",
    "LASSNTOT Index":  "LME Aluminium Total Stocks (MT)",
    "LMZSNTOT Index":  "LME Zinc Total Stocks (MT)",
    "LMNITON Index":   "LME Nickel Total Stocks (MT)",
    "LMPBTOT Index":   "LME Lead Total Stocks (MT)",
    "LMSNTOT Index":   "LME Tin Total Stocks (MT)",

    # ---- LME CANCELLED WARRANTS (daily, metric tons) ------------------------
    # Cancelled warrants = inventory earmarked for withdrawal; useful leading
    # indicator of physical tightness (spike = drawdown imminent).
    "LCSNCANC Index":  "LME Copper Cancelled Warrants (MT)",
    "LASCANWP Index":  "LME Aluminium Cancelled Warrants (MT)",
    "LMZSCANW Index":  "LME Zinc Cancelled Warrants (MT)",
    "LMNICANW Index":  "LME Nickel Cancelled Warrants (MT)",
    "LMPBCANW Index":  "LME Lead Cancelled Warrants (MT)",
    "LMSNCANW Index":  "LME Tin Cancelled Warrants (MT)",

    # ---- LME ON-WARRANT STOCKS (daily, metric tons) -------------------------
    # On-warrant = freely available (not cancelled). Complement to cancelled.
    "LCSNOWRT Index":  "LME Copper On-Warrant Stocks (MT)",
    "LASSNOWRT Index": "LME Aluminium On-Warrant Stocks (MT)",
    "LMZSOWRT Index":  "LME Zinc On-Warrant Stocks (MT)",
    "LMNIOWRT Index":  "LME Nickel On-Warrant Stocks (MT)",
    "LMPBOWRT Index":  "LME Lead On-Warrant Stocks (MT)",
    "LMSNOWRT Index":  "LME Tin On-Warrant Stocks (MT)",

    # ---- SHFE WAREHOUSE STOCKS (weekly, metric tons) [VERIFY TICKERS] -------
    # Shanghai Futures Exchange publishes every Friday after market close.
    # Bloomberg carries these but ticker format varies by subscription.
    # If empty, search Bloomberg for "SHFE" + metal name to find exact ticker.
    "SHFHCU Index":    "SHFE Copper Stocks (MT) [verify ticker]",
    "SHFHAL Index":    "SHFE Aluminium Stocks (MT) [verify ticker]",
    "SHFHZN Index":    "SHFE Zinc Stocks (MT) [verify ticker]",
    "SHFHNI Index":    "SHFE Nickel Stocks (MT) [verify ticker]",
    "SHFHPB Index":    "SHFE Lead Stocks (MT) [verify ticker]",
    "SHFHSN Index":    "SHFE Tin Stocks (MT) [verify ticker]",
    "SHFHRB Index":    "SHFE Steel Rebar Stocks (MT) [verify ticker]",

    # ---- COMEX / NYMEX WAREHOUSE STOCKS [VERIFY TICKERS] -------------------
    # COMEX warehouse stocks for Gold, Silver, Copper (HG).
    # NYMEX warehouse stocks for Platinum, Palladium.
    # These are the exact inventory series GHR uses for Pt and Pd.
    # If Bloomberg ticker incorrect, search: COMEX <metal> inventory warehouse
    "COMXGOLD Index":  "COMEX Gold Warehouse Stocks (troy oz) [verify]",
    "COMXSILV Index":  "COMEX Silver Warehouse Stocks (troy oz) [verify]",
    "COMXHG Index":    "COMEX Copper HG Warehouse Stocks (lbs) [verify]",
    "NYMXPT Index":    "NYMEX Platinum Warehouse Stocks (troy oz) [verify]",
    "NYMXPA Index":    "NYMEX Palladium Warehouse Stocks (troy oz) [verify]",

    # ---- ICE CERTIFIED STOCKS (daily) ---------------------------------------
    # ICE publishes certified exchange stocks daily for cotton, cocoa, coffee.
    # These are the exact inventory series GHR uses for those softs.
    "ICTCOTCE Index":  "ICE Certified Cotton Stocks (bales)",
    "ICTCOCOA Index":  "ICE Certified Cocoa Stocks (MT)",
    "ICTCOFFC Index":  "ICE Certified Coffee C Stocks (bags)",

    # ---- EIA ENERGY STOCKS VIA BLOOMBERG (weekly, US) -----------------------
    # US Energy Information Administration data carried by Bloomberg.
    # These are the exact inventory series GHR uses for all 6 energy commodities.
    # Published weekly (Wednesday for most, Thursday for Nat Gas storage).
    "DOESCRUS Index":  "EIA US Crude Oil Stocks excl SPR (000 bbls)",
    "DOESDS1U Index":  "EIA US Distillate Fuel Oil Stocks (000 bbls)",   # Heating oil
    "DOESG1U Index":   "EIA US Motor Gasoline Stocks (000 bbls)",         # Unleaded/RBOB
    "DOESPP1U Index":  "EIA US Propane-Propylene Stocks (000 bbls)",
    "DOENUSGW Index":  "EIA US Nat Gas Underground Storage Working (Bcf)",
    "DOECSTTT Index":  "EIA US Coal Stocks Total (000 short tons) [verify]",

    # ---- USDA GRAIN STOCKS VIA BLOOMBERG (quarterly release) ----------------
    # USDA Grain Stocks Report (March, June, September, December).
    # Bloomberg carries these as macro data points updated on release date.
    "USSTWT Index":    "USDA US Wheat Total Stocks (million bushels)",
    "USSTCORN Index":  "USDA US Corn Total Stocks (million bushels)",
    "USSTSBNS Index":  "USDA US Soybean Total Stocks (million bushels)",

    # ---- CROSS-REFERENCE MACRO (for normalised-inventory regressions) -------
    "DXY Curncy":      "US Dollar Index (DXY)",
    "USGG10YR Index":  "US 10Y Treasury Yield",
    "CL1 Comdty":      "WTI Crude F1 (cross-ref for energy regressions)",
    "GC1 Comdty":      "Gold Spot F1 (cross-ref)",
    "BCOM Index":      "Bloomberg Commodity Index",
    "GSCI Index":      "S&P GSCI Commodity Index",
}

# =============================================================================
# STYLES
# =============================================================================

TITLE_FILL  = PatternFill("solid", fgColor="0D3B66")
TITLE_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=12)
HDR_FILL    = PatternFill("solid", fgColor="1F4E79")
HDR_FONT    = Font(name="Arial", bold=True, color="FFFFFF", size=10)
GRP_FILLS   = {
    "Metals":   PatternFill("solid", fgColor="1A3A5C"),
    "Grains":   PatternFill("solid", fgColor="2E5902"),
    "Softs":    PatternFill("solid", fgColor="7B3F00"),
    "Meats":    PatternFill("solid", fgColor="6B1A1A"),
    "Energies": PatternFill("solid", fgColor="1A3A1A"),
}
DATA_FONT   = Font(name="Arial", size=9)
DATE_FONT   = Font(name="Arial", bold=True, size=9)
ALT_FILL    = PatternFill("solid", fgColor="F5F9FC")

# =============================================================================
# CONNECTION
# =============================================================================

def connect():
    print("\n  Connecting to Bloomberg (%s:%d) ..." % (HOST, PORT))
    con = pdblp.BCon(debug=False, host=HOST, port=PORT, timeout=15000)
    con.start()
    print("  Connected.\n")
    return con

# =============================================================================
# DEADLOCK-SAFE DOWNLOAD
# =============================================================================

def safe_bdh(con, ticker):
    """Single-ticker BDH with hard timeout. Returns (DataFrame|None, status)."""
    def _call():
        return con.bdh(ticker, FIELDS, START_DATE, END_DATE)

    with ThreadPoolExecutor(max_workers=1) as ex:
        fut = ex.submit(_call)
        try:
            df = fut.result(timeout=TIMEOUT_PER_TICKER)
            if isinstance(df.columns, pd.MultiIndex):
                df.columns = df.columns.get_level_values(1)
            if df.empty or "PX_LAST" not in df.columns:
                return None, "no data"
            df = df[["PX_LAST"]].dropna()
            return (None, "all NaN") if df.empty else (df, "%d rows" % len(df))
        except FuturesTimeoutError:
            return None, "TIMEOUT (%ds)" % TIMEOUT_PER_TICKER
        except Exception as e:
            return None, "ERROR: %s" % str(e)[:80]


def download_strip(con, root, max_contracts, name):
    """Download F1-Fn with early-exit after EARLY_EXIT_EMPTY consecutive failures."""
    print("  %-28s root=%-4s F1-F%d" % (name, root, max_contracts))
    result = {}
    consecutive_empty = 0
    for i in range(1, max_contracts + 1):
        tkr = "%s%d Comdty" % (root, i)
        df, status = safe_bdh(con, tkr)
        if df is not None:
            result[tkr] = df
            consecutive_empty = 0
            print("    %-22s %s" % (tkr, status))
        else:
            consecutive_empty += 1
            print("    %-22s %s" % (tkr, status))
            if consecutive_empty >= EARLY_EXIT_EMPTY:
                remaining = max_contracts - i
                if remaining > 0:
                    print("    => %d empties at F%d; skipping F%d-F%d"
                          % (EARLY_EXIT_EMPTY, i, i + 1, max_contracts))
                break
    print("  => %d / %d contracts\n" % (len(result), max_contracts))
    return result


def download_flat(con, ticker_map):
    """Download arbitrary named tickers. Returns {ticker: DataFrame}."""
    result = {}
    total = len(ticker_map)
    for idx, (tkr, desc) in enumerate(ticker_map.items(), 1):
        short = desc[:50]
        print("  [%d/%d] %-28s %s" % (idx, total, tkr, short), end="", flush=True)
        df, status = safe_bdh(con, tkr)
        if df is not None:
            result[tkr] = df
        print("  => %s" % status)
    return result

# =============================================================================
# EXCEL WRITERS
# =============================================================================

def write_strip_sheet(ws, root, name, group, strip_data):
    """One sheet: Date | F1 | F2 | ... | Fn (PX_LAST)."""
    contracts = sorted(
        strip_data.keys(),
        key=lambda t: int("".join(c for c in t.split()[0][len(root):] if c.isdigit()) or "0"),
    )
    if not contracts:
        ws["A1"] = "No data for %s" % name
        return

    all_dates = sorted(set().union(*(strip_data[t].index for t in contracts)))
    n_cols = 1 + len(contracts)

    title = "%s (%s) -- PX_LAST, end=%s" % (name, group, END_DATE)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws.cell(row=1, column=1, value=title)
    c.font = TITLE_FONT
    c.fill = GRP_FILLS.get(group, TITLE_FILL)
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 22

    c = ws.cell(row=2, column=1, value="Date")
    c.font = HDR_FONT; c.fill = HDR_FILL
    c.alignment = Alignment(horizontal="center")
    for j, tkr in enumerate(contracts, 2):
        num = "".join(ch for ch in tkr.split()[0][len(root):] if ch.isdigit())
        c = ws.cell(row=2, column=j, value="F%s" % num)
        c.font = HDR_FONT; c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center")

    for ri, dt in enumerate(all_dates, 3):
        ws.cell(row=ri, column=1, value=dt).number_format = "YYYY-MM-DD"
        ws.cell(row=ri, column=1).font = DATE_FONT
        for j, tkr in enumerate(contracts, 2):
            df = strip_data[tkr]
            cell = ws.cell(row=ri, column=j)
            if dt in df.index:
                v = df.loc[dt, "PX_LAST"]
                if pd.notna(v):
                    cell.value = float(v)
            cell.font = DATA_FONT
            cell.number_format = "#,##0.000"

    ws.column_dimensions["A"].width = 13
    for col in range(2, n_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 11
    ws.freeze_panes = "B3"
    print("    [sheet] '%s': %d rows, %d contracts" % (ws.title[:31], len(all_dates), len(contracts)))


def write_inventory_sheet(wb, sheet_name, ticker_map, data):
    """One sheet for a category of inventory tickers: Date | Ticker1 | Ticker2 ..."""
    ws = wb.create_sheet(title=sheet_name[:31])

    ordered = [t for t in ticker_map if t in data]
    if not ordered:
        ws["A1"] = "No data returned for %s" % sheet_name
        return

    all_dates = sorted(set().union(*(data[t].index for t in ordered)))
    n_cols = 1 + len(ordered)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws.cell(row=1, column=1,
                value="%s -- PX_LAST, %s to %s" % (sheet_name, START_DATE, END_DATE))
    c.font = TITLE_FONT; c.fill = TITLE_FILL
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 22

    c = ws.cell(row=2, column=1, value="Date")
    c.font = HDR_FONT; c.fill = HDR_FILL
    c.alignment = Alignment(horizontal="center")
    for j, tkr in enumerate(ordered, 2):
        label = ticker_map[tkr].replace(" [verify ticker]", "").replace(" [verify]", "")
        c = ws.cell(row=2, column=j, value=label)
        c.font = HDR_FONT; c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[2].height = 30

    for ri, dt in enumerate(all_dates, 3):
        fill = ALT_FILL if ri % 2 == 0 else None
        ws.cell(row=ri, column=1, value=dt).number_format = "YYYY-MM-DD"
        ws.cell(row=ri, column=1).font = DATE_FONT
        for j, tkr in enumerate(ordered, 2):
            cell = ws.cell(row=ri, column=j)
            df = data[tkr]
            if dt in df.index:
                v = df.loc[dt, "PX_LAST"]
                if pd.notna(v):
                    cell.value = float(v)
            cell.font = DATA_FONT
            cell.number_format = "#,##0.00"
            if fill:
                cell.fill = fill

    ws.column_dimensions["A"].width = 13
    for col in range(2, n_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 22
    ws.freeze_panes = "B3"
    print("    [sheet] '%s': %d rows, %d tickers" % (ws.title, len(all_dates), len(ordered)))


def write_readme(wb, title, rows):
    ws = wb.active
    ws.title = "README"
    meta = [
        (title, ""),
        ("", ""),
        ("Generated",        datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Start Date",       START_DATE),
        ("End Date",         END_DATE),
        ("Fields",           ", ".join(FIELDS)),
        ("Timeout/ticker",   "%ds; skip after %d consecutive empties"
                             % (TIMEOUT_PER_TICKER, EARLY_EXIT_EMPTY)),
        ("Reference paper",  "Gorton, Hayashi & Rouwenhorst (2013), Rev Finance"),
        ("", ""),
    ] + rows
    for r, (a, b) in enumerate(meta, 1):
        ws.cell(row=r, column=1, value=a).font = Font(name="Arial", bold=True, size=10)
        ws.cell(row=r, column=2, value=b).font = Font(name="Arial", size=10)
    ws.cell(row=1, column=1).font = Font(name="Arial", bold=True, size=13, color="0D3B66")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 65


def save_wb(wb, filename):
    path = os.path.join(SCRIPT_DIR, filename)
    wb.save(path)
    print("\n  Saved => %s" % path)
    print("  Sheets: %s\n" % ", ".join(wb.sheetnames))


# =============================================================================
# MAIN
# =============================================================================

def main():
    print("\n" + "=" * 70)
    print("  GHR REPLICATION — BLOOMBERG FULL DATA DOWNLOADER")
    print("  Start: %s   End: %s" % (START_DATE, END_DATE))
    print("  Timeout: %ds/ticker, early-exit after %d empty"
          % (TIMEOUT_PER_TICKER, EARLY_EXIT_EMPTY))
    print("=" * 70)

    con = connect()

    # ------------------------------------------------------------------
    # SECTION 1 — FUTURES PRICES
    # Download all strips; sort into one workbook per commodity group
    # ------------------------------------------------------------------

    # Collect strips by group
    group_data   = defaultdict(dict)   # {group: {root: {tkr: df}}}
    group_names  = defaultdict(dict)   # {group: {root: name}}

    print("\n" + "=" * 70)
    print("  SECTION 1 OF 2 — FUTURES STRIPS")
    print("=" * 70)

    for root, cfg in GHR_FUTURES.items():
        name  = cfg["name"]
        maxc  = cfg["max_contracts"]
        group = cfg["group"]
        print("\n--- %s ---" % name)
        strip = download_strip(con, root, maxc, name)
        group_data[group][root]  = strip
        group_names[group][root] = name

    # Write one workbook per group
    group_files = {
        "Metals":   "GHR_Futures_Metals.xlsx",
        "Grains":   "GHR_Futures_Grains.xlsx",
        "Softs":    "GHR_Futures_Softs.xlsx",
        "Meats":    "GHR_Futures_Meats.xlsx",
        "Energies": "GHR_Futures_Energies.xlsx",
    }
    group_readme = {
        "Metals": [
            ("LME Metals",  "LP,LA,LX,LN,LL,LT — F1-F27 (GHR: Cu,Al,Zn,Ni,Pb,Sn)"),
            ("Precious",    "PL,PA — Platinum and Palladium (GHR uses both)"),
            ("Reference",   "GC,SI,HG — Gold, Silver, COMEX Copper (not in GHR core)"),
        ],
        "Grains": [
            ("CBOT", "W,C,S,BO,SM,O — all 6 GHR grain commodities"),
            ("Note", "Generic contract: W1 Comdty = nearest active CBOT wheat contract"),
        ],
        "Softs": [
            ("ICE/CME", "CT,CC,KC,JO,LB — all 5 GHR soft commodities"),
            ("Note",    "Lumber (LB) is random-length lumber on CME"),
        ],
        "Meats": [
            ("CME",  "LC,LH,FC,DA,CB — Live Cattle, Lean Hogs, Feeder Cattle, Milk, Butter"),
            ("Note", "Pork Bellies (PB) delisted July 2011; limited data from 2010 onwards"),
        ],
        "Energies": [
            ("NYMEX/ICE", "CL,HO,RB,PN,NG,QL — all 6 GHR energy commodities"),
            ("Note",      "RB=RBOB replaces HU (unleaded gas) from 2006; CO,QS,GO added for reference"),
        ],
    }

    print("\n" + "=" * 70)
    print("  WRITING FUTURES WORKBOOKS")
    print("=" * 70)

    for group, filename in group_files.items():
        wb = Workbook()
        write_readme(wb, "GHR Replication — %s Futures" % group,
                     group_readme.get(group, []))
        strips_in_group = group_data.get(group, {})
        for root, strip in strips_in_group.items():
            name = group_names[group][root]
            ws = wb.create_sheet(title=name[:31])
            write_strip_sheet(ws, root, name, group, strip)
        save_wb(wb, filename)

    # ------------------------------------------------------------------
    # SECTION 2 — INVENTORY / WAREHOUSE STOCKS
    # ------------------------------------------------------------------

    print("=" * 70)
    print("  SECTION 2 OF 2 — INVENTORY / WAREHOUSE STOCKS")
    print("=" * 70)

    # Split INVENTORY_TICKERS into two logical buckets for separate sheets
    EXCHANGE_TICKERS = {k: v for k, v in INVENTORY_TICKERS.items()
                        if any(tag in k for tag in
                               ["LCSNSTOT","LASSNTOT","LMZSNTOT","LMNITON","LMPBTOT","LMSNTOT",
                                "LCSNCANC","LASCANWP","LMZSCANW","LMNICANW","LMPBCANW","LMSNCANW",
                                "LCSNOWRT","LASSNOWRT","LMZSOWRT","LMNIOWRT","LMPBOWRT","LMSNOWRT",
                                "SHFH","COMX","NYMX","ICTC"])}

    GOVT_TICKERS = {k: v for k, v in INVENTORY_TICKERS.items()
                    if k not in EXCHANGE_TICKERS}

    print("\n--- Exchange Warehouse Stocks (LME, SHFE, COMEX/NYMEX, ICE) ---")
    data_exchange = download_flat(con, EXCHANGE_TICKERS)

    print("\n--- Govt Data via Bloomberg (EIA energy, USDA grains, macro) ---")
    data_govt = download_flat(con, GOVT_TICKERS)

    # Write inventory workbooks
    print("\n" + "=" * 70)
    print("  WRITING INVENTORY WORKBOOKS")
    print("=" * 70)

    # Exchange inventory workbook — one sheet per category
    wb_ex = Workbook()
    write_readme(wb_ex, "GHR Replication — Exchange Inventory / Warehouse Stocks", [
        ("LME Total Stocks",          "Cu,Al,Zn,Ni,Pb,Sn — daily metric tons (Bloomberg Index)"),
        ("LME Cancelled Warrants",    "Cu,Al,Zn,Ni,Pb,Sn — daily MT (earmarked for withdrawal)"),
        ("LME On-Warrant Stocks",     "Cu,Al,Zn,Ni,Pb,Sn — daily MT (freely available)"),
        ("SHFE Stocks [verify]",      "Cu,Al,Zn,Ni,Pb,Sn,Rebar — weekly MT (verify tickers)"),
        ("COMEX/NYMEX [verify]",      "Au,Ag,HG Copper,Pt,Pd warehouse stocks"),
        ("ICE Certified Stocks",      "Cotton,Cocoa,Coffee C — daily certified stock levels"),
    ])

    # Sub-group for cleaner sheet organisation
    sub_groups = [
        ("LME Total Stocks",
         {k: v for k, v in EXCHANGE_TICKERS.items() if "STOT" in k or "TON" in k}),
        ("LME Cancelled Warrants",
         {k: v for k, v in EXCHANGE_TICKERS.items() if "CANC" in k or "CANW" in k or "CANWP" in k}),
        ("LME On-Warrant",
         {k: v for k, v in EXCHANGE_TICKERS.items() if "OWRT" in k}),
        ("SHFE Stocks",
         {k: v for k, v in EXCHANGE_TICKERS.items() if "SHFH" in k}),
        ("COMEX-NYMEX Stocks",
         {k: v for k, v in EXCHANGE_TICKERS.items() if "COMX" in k or "NYMX" in k}),
        ("ICE Certified Stocks",
         {k: v for k, v in EXCHANGE_TICKERS.items() if "ICTC" in k}),
    ]
    for sheet_name, tmap in sub_groups:
        write_inventory_sheet(wb_ex, sheet_name, tmap, data_exchange)
    save_wb(wb_ex, "GHR_Inventory_Exchange.xlsx")

    # Govt data workbook — EIA + USDA + macro
    wb_gov = Workbook()
    write_readme(wb_gov, "GHR Replication — Govt Data via Bloomberg (EIA, USDA, Macro)", [
        ("EIA Energy",       "US crude oil, distillate, gasoline, propane, nat gas, coal stocks"),
        ("USDA Grains",      "US wheat, corn, soybean stocks (quarterly Grain Stocks Report)"),
        ("Macro cross-ref",  "DXY, US 10Y yield, BCOM, GSCI for replication context"),
        ("Note",             "EIA/USDA data via Bloomberg Index tickers; periodicity weekly/monthly/quarterly"),
    ])

    eia_tickers  = {k: v for k, v in GOVT_TICKERS.items() if "DOE" in k}
    usda_tickers = {k: v for k, v in GOVT_TICKERS.items() if "USS" in k}
    macro_tickers= {k: v for k, v in GOVT_TICKERS.items()
                    if k not in eia_tickers and k not in usda_tickers}

    write_inventory_sheet(wb_gov, "EIA Energy Stocks",       eia_tickers,   data_govt)
    write_inventory_sheet(wb_gov, "USDA Grain Stocks",       usda_tickers,  data_govt)
    write_inventory_sheet(wb_gov, "Macro Cross-Reference",   macro_tickers, data_govt)
    save_wb(wb_gov, "GHR_Inventory_Govt.xlsx")

    # Disconnect
    try:
        con.stop()
    except Exception:
        pass

    print("=" * 70)
    print("  ALL DOWNLOADS COMPLETE")
    print("  Futures files:")
    for g, f in group_files.items():
        print("    %-12s => %s" % (g, f))
    print("  Inventory files:")
    print("    Exchange   => GHR_Inventory_Exchange.xlsx")
    print("    Govt/EIA   => GHR_Inventory_Govt.xlsx")
    print("=" * 70 + "\n")


if __name__ == "__main__":
    main()
