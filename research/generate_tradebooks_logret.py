"""
research/generate_tradebooks_logret.py
========================================
Formula-driven Excel tradebooks for the LOCKED Metals ex-ante strategy set
(log-return methodology, per project_risk_premia_research_framework memory
#14/#16), one workbook per product, 7 sheets each:
  Momentum | Carry_F1F3 | Carry_F1F13 | CarryMom_F1F3 | CarryMom_F1F13 | Value | EW_Portfolio

Every column from the ratio-continuous price onward is a LIVE EXCEL FORMULA
(not a pasted value) so the strategy logic can be independently audited/
recalculated in Excel -- same convention as scripts/generate_tradebooks.py
(dollar-PnL version), adapted here for log-return PnL and ratio (not
additive) continuous-price reconstruction. Date/F1_raw/F2_raw/Phase/curve
tenor prices are passed through as given data (Phase already encodes the
roll-calendar lookup from rolling_continuous.py -- re-deriving the calendar
itself in formulas was explicitly descoped, same as the existing script).

Each sheet's Performance Summary block shows BOTH the Python-computed
reference Return/Vol/IR (from research/engine.py, already verified against
common_engine.py) AND live Excel formulas computing the same statistics --
opening the file in Excel and comparing the two is the reconciliation check.
"""

from __future__ import annotations

import importlib
import os
import sys

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

_ASSET_CLASS = os.environ.get("TRADEBOOK_ASSET_CLASS", "metals")
cfg = importlib.import_module(f"configs.{_ASSET_CLASS}")

from engine import (carry_v1v2_composite_position, carry_v4_position,
                     log_return_metrics, momentum_composite_position, value_v1_position)

_OUT_DIR_NAME = "tradebooks_logret" if _ASSET_CLASS == "metals" else f"tradebooks_logret_{_ASSET_CLASS}"
OUT_DIR = os.path.join(cfg._REPO_ROOT, _OUT_DIR_NAME)
os.makedirs(OUT_DIR, exist_ok=True)

# ── Styling (mirrors scripts/generate_tradebooks.py's conventions) ──────────
_SECTION_FILL = PatternFill("solid", fgColor="1F1F1F")
_SECTION_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill("solid", fgColor="3D3D3D")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=9)
_METRIC_FILL = PatternFill("solid", fgColor="EFEFEF")
_METRIC_FONT = Font(size=9)
_TB_HEADER_FILL = PatternFill("solid", fgColor="2E5C8A")
_TB_HEADER_FONT = Font(bold=True, color="FFFFFF", size=9)
_ROLL_DAY_FILL = PatternFill("solid", fgColor="FF9F5A")
_F2_TRACKING_FILL = PatternFill("solid", fgColor="F5C265")
_BRIDGE_DAY_FILL = PatternFill("solid", fgColor="8FD3E8")
_REF_FILL = PatternFill("solid", fgColor="DDEBF7")
_LIVE_FILL = PatternFill("solid", fgColor="C6EFCE")


def _style(cell, fill=None, font=None, align=None):
    if fill: cell.fill = fill
    if font: cell.font = font
    if align: cell.alignment = align


def _data_start_row(n_metrics: int) -> int:
    """The Performance Summary block consumes a FIXED number of rows before
    the tradebook's own column-header row: 1 (title) + 1 (Metric/Ref/Live
    header) + n_metrics (one row per metric) + 1 (blank) + 3 (Roll/F2/Bridge
    legend) + 1 (blank) + 1 (TRADEBOOK title) = n_metrics + 8 rows, so the
    column-header row is n_metrics+9 and the first DATA row is n_metrics+10.
    Every row-formula function must anchor on THIS, not a hardcoded r=i+2 --
    that hardcoding was a real bug caught during verification (formulas were
    pointing at Performance-Summary-block cells instead of data rows)."""
    return n_metrics + 10


def _f1_cont_ratio_formula(r: int, i: int) -> str:
    """Ratio-continuous reconstruction from Phase+F1_raw+F2_raw (columns B,C,D
    at rows r), written into column E. Mirrors research/ratio_continuous.py's
    build_rolling_f1_ratio() phase logic exactly: F1_Tracking multiplies by
    the F1 ratio; Bridge multiplies by F1[t]/F2[t-1]; everything else (Roll
    day AND F2_Tracking share the same rule) multiplies by the F2 ratio."""
    if i == 0:
        return f"=B{r}"
    return (f'=IF(D{r}="F1_Tracking",E{r-1}*(B{r}/B{r-1}),'
            f'IF(D{r}="Bridge",E{r-1}*(B{r}/C{r-1}),'
            f'E{r-1}*(C{r}/C{r-1})))')


def _perf_summary_block(ws, ref_metrics: dict, live_formulas: dict, n_cols: int):
    """Writes a Performance Summary block with Python REFERENCE values (blue)
    next to LIVE Excel formulas computing the same thing (green) -- opening
    in Excel and comparing the two columns is the reconciliation check."""
    ws.append(["PERFORMANCE SUMMARY", "", ""])
    _style(ws.cell(ws.max_row, 1), fill=_SECTION_FILL, font=_SECTION_FONT)
    for c in (2, 3):
        _style(ws.cell(ws.max_row, c), fill=_SECTION_FILL)

    ws.append(["Metric", "Python Reference (engine.py)", "Live Excel Formula"])
    for c in (1, 2, 3):
        _style(ws.cell(ws.max_row, c), fill=_HEADER_FILL, font=_HEADER_FONT)

    for i, key in enumerate(ref_metrics):
        ws.append([key, ref_metrics[key], live_formulas.get(key, "")])
        fill = _METRIC_FILL if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        _style(ws.cell(ws.max_row, 1), fill=fill, font=_METRIC_FONT)
        _style(ws.cell(ws.max_row, 2), fill=_REF_FILL, font=_METRIC_FONT)
        _style(ws.cell(ws.max_row, 3), fill=_LIVE_FILL, font=_METRIC_FONT)

    ws.append(["", "", ""])
    ws.append(["Roll day (Roll_LTD-N)", "", ""])
    _style(ws.cell(ws.max_row, 1), fill=_ROLL_DAY_FILL, font=Font(size=9))
    ws.append(["F2 Tracking", "", ""])
    _style(ws.cell(ws.max_row, 1), fill=_F2_TRACKING_FILL, font=Font(size=9))
    ws.append(["Expiry/Bridge day", "", ""])
    _style(ws.cell(ws.max_row, 1), fill=_BRIDGE_DAY_FILL, font=Font(size=9))
    ws.append([""] * n_cols)

    ws.append(["TRADEBOOK"] + [""] * (n_cols - 1))
    _style(ws.cell(ws.max_row, 1), fill=_SECTION_FILL, font=_SECTION_FONT)


def _write_headers_and_rows(ws, headers: list[str], row_formula_fn, n_rows: int, phase_col_idx: int | None):
    ws.append(headers)
    hdr_row = ws.max_row
    for c in range(1, len(headers) + 1):
        _style(ws.cell(hdr_row, c), fill=_TB_HEADER_FILL, font=_TB_HEADER_FONT, align=Alignment(horizontal="center"))

    for i in range(n_rows):
        row_vals = row_formula_fn(i)
        ws.append(row_vals)
        if phase_col_idx is not None:
            phase_val = row_vals[phase_col_idx - 1]
            row_fill = None
            if isinstance(phase_val, str) and phase_val.startswith("Roll_LTD"):
                row_fill = _ROLL_DAY_FILL
            elif phase_val == "F2_Tracking":
                row_fill = _F2_TRACKING_FILL
            elif phase_val == "Bridge":
                row_fill = _BRIDGE_DAY_FILL
            if row_fill is not None:
                for c in range(1, len(headers) + 1):
                    ws.cell(ws.max_row, c).fill = row_fill

    for c in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 15
    ws.freeze_panes = ws.cell(hdr_row + 1, 1).coordinate
    return hdr_row


def _sign_pair_formula(col: str, r: int, i: int, fast: int, slow: int, row0: int) -> str:
    """SIGN(MA_fast - MA_slow) on column `col`, blank until the slow window
    is fully populated (mirrors pandas .rolling(window).mean()'s default
    min_periods=window -- no partial-window average). `row0` is the sheet's
    actual first DATA row (NOT a hardcoded 2 -- the Performance Summary
    block pushes data down well past row 2, a bug caught during
    verification), used to clamp the rolling-window range at the true start
    of the data rather than assuming row 2."""
    if i + 1 < slow:
        return ""
    slow_start = max(row0, r - slow + 1)
    fast_expr = f"{col}{r}" if fast == 1 else f"AVERAGE({col}{max(row0, r - fast + 1)}:{col}{r})"
    slow_expr = f"AVERAGE({col}{slow_start}:{col}{r})"
    return f"=SIGN({fast_expr}-{slow_expr})"


def _position_formula(sign_col: str, r: int, i: int, eff_shift: int) -> str:
    if i + 1 > eff_shift:
        return f'=IF({sign_col}{r - eff_shift}="",0,{sign_col}{r - eff_shift})'
    return 0


def _tc_cost_formula(pos_change_col: str, pos_col: str, phase_col: str, r: int, i: int, tc_bps: int) -> str:
    if i == 0:
        return f"={pos_change_col}{r}*({tc_bps}/10000/2)"
    return (f'=({pos_change_col}{r}+IF(AND(LEFT({phase_col}{r},8)="Roll_LTD",'
            f'{pos_col}{r}<>0,{pos_col}{r - 1}<>0,SIGN({pos_col}{r})=SIGN({pos_col}{r - 1})),1,0))'
            f'*({tc_bps}/10000/2)')


# ══════════════════════════════════════════════════════════════════
# SHEET WRITERS -- one per locked strategy
# ══════════════════════════════════════════════════════════════════

def write_momentum_sheet(wb, sheet_name: str, f1_series: pd.DataFrame, pairs, shift_n: int, tc_bps: int, ref_metrics: dict):
    ws = wb.create_sheet(sheet_name)
    n_rows = len(f1_series)
    eff_shift = shift_n + 1
    dates = f1_series.index.tolist()
    f1_raw = f1_series["F1_raw"].tolist()
    f2_raw = f1_series["F2_raw"].tolist()
    phase = f1_series["Phase"].tolist()

    d0 = _data_start_row(len(ref_metrics))
    d1 = d0 + n_rows - 1
    live = {
        "Return (annualized %)": f"=AVERAGE(S{d0}:S{d1})*252*100",
        "Vol (annualized %)": f"=STDEV(S{d0}:S{d1})*SQRT(252)*100",
        "IR (Sharpe)": f"=AVERAGE(S{d0}:S{d1})/STDEV(S{d0}:S{d1})*SQRT(252)",
        "N active days": f"=COUNT(S{d0}:S{d1})",
    }
    _perf_summary_block(ws, ref_metrics, live, n_cols=19)

    headers = ["Date", "F1_raw", "F2_raw", "Phase", "F1_cont_ratio", "LogPrice",
               f"Sign_{pairs[0][0]}_{pairs[0][1]}", f"Sign_{pairs[1][0]}_{pairs[1][1]}", f"Sign_{pairs[2][0]}_{pairs[2][1]}",
               "Composite_Raw", "Composite_Sign", "Position", "LogReturn", "Gross_PnL",
               "Position_Change", "TC_Cost", "Net_PnL", "Cum_Net_PnL", "Active_NetPnL"]

    def row_fn(i):
        r = d0 + i
        vals = [dates[i], f1_raw[i], f2_raw[i], phase[i], _f1_cont_ratio_formula(r, i),
                f"=IF(E{r}<=0,\"\",LN(E{r}))"]
        for (fast, slow) in pairs:
            vals.append(_sign_pair_formula("B", r, i, fast, slow, d0))
        vals.append(f'=IF(OR(G{r}="",H{r}="",I{r}=""),"",(G{r}+H{r}+I{r})/3)')
        vals.append(f'=IF(J{r}="","",SIGN(J{r}))')
        vals.append(_position_formula("K", r, i, eff_shift))
        vals.append("" if i == 0 else f"=F{r}-F{r-1}")
        vals.append(f'=IF(OR(L{r}="",M{r}=""),0,L{r}*M{r})')
        vals.append(f"=ABS(L{r})" if i == 0 else f"=ABS(L{r}-L{r-1})")
        vals.append(_tc_cost_formula("O", "L", "D", r, i, tc_bps))
        vals.append(f"=N{r}-P{r}")
        vals.append(f"=Q{r}" if i == 0 else f"=R{r-1}+Q{r}")
        vals.append(f"=IF(L{r}<>0,Q{r},\"\")")
        return vals

    _write_headers_and_rows(ws, headers, row_fn, n_rows, phase_col_idx=4)
    return ws


def write_carry_composite_sheet(wb, sheet_name: str, curve: pd.DataFrame, f1_series: pd.DataFrame,
                                 near: str, far: str, zscore_window: int, shift_n: int, tc_bps: int,
                                 ref_metrics: dict):
    """V1(Level)+V2(Z-score) composite Carry, on tenor pair (near, far)."""
    ws = wb.create_sheet(sheet_name)
    idx = f1_series.index
    n_rows = len(idx)
    eff_shift = shift_n + 1
    dates = idx.tolist()
    price_near = curve[near].reindex(idx).tolist()
    price_far = curve[far].reindex(idx).tolist()
    f1_raw = f1_series["F1_raw"].tolist()
    f2_raw = f1_series["F2_raw"].tolist()
    phase = f1_series["Phase"].tolist()

    d0 = _data_start_row(len(ref_metrics))
    d1 = d0 + n_rows - 1
    live = {
        "Return (annualized %)": f"=AVERAGE(U{d0}:U{d1})*252*100",
        "Vol (annualized %)": f"=STDEV(U{d0}:U{d1})*SQRT(252)*100",
        "IR (Sharpe)": f"=AVERAGE(U{d0}:U{d1})/STDEV(U{d0}:U{d1})*SQRT(252)",
        "N active days": f"=COUNT(U{d0}:U{d1})",
    }
    _perf_summary_block(ws, ref_metrics, live, n_cols=21)

    headers = ["Date", f"Price_{near}", f"Price_{far}", "F1_raw", "F2_raw", "Phase",
               "F1_cont_ratio", "LogPrice", "CarrySlope", "V1_Sign", "Z_Score", "V2_Sign",
               "Composite_Sign", "Position", "LogReturn", "Gross_PnL", "Position_Change",
               "TC_Cost", "Net_PnL", "Cum_Net_PnL", "Active_NetPnL"]

    def row_fn(i):
        r = d0 + i
        vals = [dates[i], price_near[i], price_far[i], f1_raw[i], f2_raw[i], phase[i]]
        # Ratio-continuous reconstruction (D=F1_raw, E=F2_raw, F=Phase, G=F1_cont_ratio in THIS sheet's layout)
        if i == 0:
            vals.append(f"=D{r}")
        else:
            vals.append(f'=IF(F{r}="F1_Tracking",G{r-1}*(D{r}/D{r-1}),'
                        f'IF(F{r}="Bridge",G{r-1}*(D{r}/E{r-1}),'
                        f'G{r-1}*(E{r}/E{r-1})))')
        vals.append(f"=IF(G{r}<=0,\"\",LN(G{r}))")  # H LogPrice
        vals.append(f'=IF(OR(B{r}="",C{r}=""),"",(B{r}-C{r})/B{r})')  # I CarrySlope
        vals.append(f'=IF(I{r}="","",SIGN(I{r}))')  # J V1_Sign
        if i + 1 < zscore_window:
            vals.append("")  # K Z_Score
        else:
            start = max(d0, r - zscore_window + 1)
            mean_e, std_e = f"AVERAGE(I{start}:I{r})", f"STDEV(I{start}:I{r})"
            vals.append(f'=IF(I{r}="","",IF({std_e}=0,"",(I{r}-{mean_e})/{std_e}))')
        vals.append(f'=IF(K{r}="","",SIGN(K{r}))')  # L V2_Sign
        vals.append(f'=IF(OR(J{r}="",L{r}=""),"",SIGN((J{r}+L{r})/2))')  # M Composite_Sign
        vals.append(_position_formula("M", r, i, eff_shift))  # N Position
        vals.append("" if i == 0 else f"=H{r}-H{r-1}")  # O LogReturn
        vals.append(f'=IF(OR(N{r}="",O{r}=""),0,N{r}*O{r})')  # P Gross_PnL
        vals.append(f"=ABS(N{r})" if i == 0 else f"=ABS(N{r}-N{r-1})")  # Q Position_Change
        vals.append(_tc_cost_formula("Q", "N", "F", r, i, tc_bps))  # R TC_Cost
        vals.append(f"=P{r}-R{r}")  # S Net_PnL
        vals.append(f"=S{r}" if i == 0 else f"=T{r-1}+S{r}")  # T Cum_Net_PnL
        vals.append(f"=IF(N{r}<>0,S{r},\"\")")  # U Active_NetPnL
        return vals

    _write_headers_and_rows(ws, headers, row_fn, n_rows, phase_col_idx=6)
    return ws


def write_carrymomentum_sheet(wb, sheet_name: str, curve: pd.DataFrame, f1_series: pd.DataFrame,
                               near: str, far: str, horizon: int, shift_n: int, tc_bps: int,
                               ref_metrics: dict):
    ws = wb.create_sheet(sheet_name)
    idx = f1_series.index
    n_rows = len(idx)
    eff_shift = shift_n + 1
    dates = idx.tolist()
    price_near = curve[near].reindex(idx).tolist()
    price_far = curve[far].reindex(idx).tolist()
    f1_raw = f1_series["F1_raw"].tolist()
    f2_raw = f1_series["F2_raw"].tolist()
    phase = f1_series["Phase"].tolist()

    d0 = _data_start_row(len(ref_metrics))
    d1 = d0 + n_rows - 1
    live = {
        "Return (annualized %)": f"=AVERAGE(S{d0}:S{d1})*252*100",
        "Vol (annualized %)": f"=STDEV(S{d0}:S{d1})*SQRT(252)*100",
        "IR (Sharpe)": f"=AVERAGE(S{d0}:S{d1})/STDEV(S{d0}:S{d1})*SQRT(252)",
        "N active days": f"=COUNT(S{d0}:S{d1})",
    }
    _perf_summary_block(ws, ref_metrics, live, n_cols=19)

    headers = ["Date", f"Price_{near}", f"Price_{far}", "F1_raw", "F2_raw", "Phase",
               "F1_cont_ratio", "LogPrice", "CarrySlope", f"CarrySlope_Lag{horizon}", "RawChange",
               "Signal_Sign", "Position", "LogReturn", "Gross_PnL", "Position_Change",
               "TC_Cost", "Net_PnL", "Cum_Net_PnL", "Active_NetPnL"]

    def row_fn(i):
        r = d0 + i
        vals = [dates[i], price_near[i], price_far[i], f1_raw[i], f2_raw[i], phase[i]]
        if i == 0:
            vals.append(f"=D{r}")
        else:
            vals.append(f'=IF(F{r}="F1_Tracking",G{r-1}*(D{r}/D{r-1}),'
                        f'IF(F{r}="Bridge",G{r-1}*(D{r}/E{r-1}),'
                        f'G{r-1}*(E{r}/E{r-1})))')
        vals.append(f"=IF(G{r}<=0,\"\",LN(G{r}))")  # H LogPrice
        vals.append(f'=IF(OR(B{r}="",C{r}=""),"",(B{r}-C{r})/B{r})')  # I CarrySlope
        vals.append("" if i < horizon else f"=I{r-horizon}")  # J CarrySlope_Lag
        vals.append(f'=IF(OR(I{r}="",J{r}=""),"",I{r}-J{r})')  # K RawChange
        vals.append(f'=IF(K{r}="","",SIGN(K{r}))')  # L Signal_Sign
        vals.append(_position_formula("L", r, i, eff_shift))  # M Position
        vals.append("" if i == 0 else f"=H{r}-H{r-1}")  # N LogReturn
        vals.append(f'=IF(OR(M{r}="",N{r}=""),0,M{r}*N{r})')  # O Gross_PnL
        vals.append(f"=ABS(M{r})" if i == 0 else f"=ABS(M{r}-M{r-1})")  # P Position_Change
        vals.append(_tc_cost_formula("P", "M", "F", r, i, tc_bps))  # Q TC_Cost
        vals.append(f"=O{r}-Q{r}")  # R Net_PnL
        vals.append(f"=R{r}" if i == 0 else f"=S{r-1}+R{r}")  # S Cum_Net_PnL -- fixed below
        return vals

    # Fix cumulative column letter references (R Net_PnL, S Cum_Net_PnL, T Active_NetPnL)
    def row_fn2(i):
        r = d0 + i
        vals = row_fn(i)
        vals[18] = f"=R{r}" if i == 0 else f"=S{r-1}+R{r}"  # S Cum_Net_PnL
        vals.append(f"=IF(M{r}<>0,R{r},\"\")")  # T Active_NetPnL
        return vals

    headers.append("Active_NetPnL")
    _write_headers_and_rows(ws, headers, row_fn2, n_rows, phase_col_idx=6)
    return ws


def _col_letter(n): 
    return get_column_letter(n)


def build_momentum_workbook(tc_bps: int = 5):
    """Momentum.xlsx: one sheet per Metals product (Copper/Aluminium/Lead/Zinc,
    each showing its own full price/signal/position/PnL detail) + an
    EW_Combined sheet that equal-weight-averages the 4 products' Net_PnL
    day-by-day (Level 1 of the Table-2-style construction). All 4 products
    are truncated to their COMMON date range first, so cross-sheet row
    references line up on the same date -- Copper's own data ends
    2025-12-31 while Aluminium/Lead/Zinc run to 2026-05-19, so naive
    row-by-row combination without this truncation would misalign dates."""
    products = cfg.PRODUCTS
    f1_series_by_product = {p: cfg.load_f1_series_logret(p) for p in products}

    common_start = max(s.index.min() for s in f1_series_by_product.values())
    common_end = min(s.index.max() for s in f1_series_by_product.values())
    print(f"Common date range across all {len(products)} products: {common_start.date()} to {common_end.date()}")

    for p in products:
        f1_series_by_product[p] = f1_series_by_product[p].loc[common_start:common_end]
    common_idx = _common_trading_index(f1_series_by_product, products)
    for p in products:
        f1_series_by_product[p] = f1_series_by_product[p].loc[common_idx]

    n_rows = len(f1_series_by_product[products[0]])
    for p in products:
        assert len(f1_series_by_product[p]) == n_rows, f"{p} row count mismatch after truncation"
        assert (f1_series_by_product[p].index == f1_series_by_product[products[0]].index).all(), \
            f"{p} date index mismatch after truncation"

    wb = Workbook()
    wb.remove(wb.active)

    net_pnl_col = "Q"   # Net_PnL column in the Momentum sheet layout (A..S, Q=17th=Net_PnL)
    active_col = "S"    # Active_NetPnL column

    # Python reference metrics computed on the SAME truncated common window,
    # using the actual ratio-continuous log_price (reused via
    # _rebuild_ratio_series so the config/calendar-file fix lives in one place).
    ref_metrics_by_product = {}
    combined_net_pnl = None
    for p in products:
        raw, log_price = _rebuild_ratio_series(p, common_start, common_end)
        # Intersection-filtered BEFORE computing the signal: the written
        # Excel sheet's own rolling-window formulas (AVERAGE over N *rows*,
        # not N *calendar days*) only ever see the rows actually written to
        # that sheet, so the Python reference metrics must be computed on
        # that exact same row set to reconcile against the Excel live
        # formulas -- computing the signal on a richer, un-intersected
        # calendar first would make the two disagree by construction. This
        # is a small, disclosed approximation for cross-exchange asset
        # classes (Energy/NGL): a handful of each product's own trading
        # days get dropped where another product in the set didn't trade
        # (see _common_trading_index docstring) -- fine for this
        # tradebook's formula-audit purpose, not used for the official
        # regime-table numbers (run_regime_table.py / statarb.py never
        # intersect across products this way).
        raw = raw.loc[common_idx]
        log_price = log_price.loc[common_idx]
        pos = momentum_composite_position(raw["F1_raw"], cfg.MOMENTUM_PAIRS, cfg.MOMENTUM_SHIFT_N)
        from engine import log_return_daily
        _, net = log_return_daily(pos, log_price, tc_bps, raw["Phase"])
        m = log_return_metrics(pos, log_price, tc_bps, raw["Phase"])
        # Active-day basis throughout (Return/Vol/IR all from the SAME
        # active-day-only sample), matching what the Excel Live Formula
        # column actually computes (AVERAGE/STDEV over Active_NetPnL, which
        # is blank -- not zero -- on inactive days) -- deliberately NOT the
        # legacy dollar-PnL convention's mixed basis (full-sample Return,
        # active-day Sharpe), which does not reconcile cleanly against a
        # single Excel column and caused a real false-alarm during
        # verification (2026-07-23) before being traced to this inconsistency.
        active_net = net[pos != 0]
        ref_metrics_by_product[p] = {
            "Return (annualized %)": round(active_net.mean() * 252 * 100, 4),
            "Vol (annualized %)": round(active_net.std(ddof=1) * np.sqrt(252) * 100, 4),
            "IR (Sharpe)": round(m["net"], 4),
            "N active days": m["nact"],
        }
        combined_net_pnl = net if combined_net_pnl is None else (combined_net_pnl + net)
        f1_series_by_product[p] = raw  # use the freshly-rebuilt, truncated, reanchored series consistently

    combined_net_pnl = combined_net_pnl / len(products)
    combo_ir = combined_net_pnl.mean() / combined_net_pnl.std() * np.sqrt(252)
    combo_ret = combined_net_pnl.mean() * 252
    combo_vol = combined_net_pnl.std() * np.sqrt(252)
    ew_ref_metrics = {
        "Return (annualized %)": round(combo_ret * 100, 4),
        "Vol (annualized %)": round(combo_vol * 100, 4),
        "IR (Sharpe)": round(combo_ir, 4),
        "N days": len(combined_net_pnl),
    }

    for p in products:
        write_momentum_sheet(wb, p, f1_series_by_product[p], cfg.MOMENTUM_PAIRS,
                              cfg.MOMENTUM_SHIFT_N, tc_bps, ref_metrics_by_product[p])

    # EW_Combined sheet: cross-references each product sheet's Net_PnL column,
    # equal-weight average day-by-day, cumulative + summary stats. Each
    # product sheet uses a 4-key ref_metrics dict (Return/Vol/IR/N active
    # days), so its own data starts at _data_start_row(4) -- computed
    # explicitly here rather than assumed, since this sheet's OWN data-start
    # row (from its own ew_ref_metrics dict) is a logically separate number
    # that only happens to coincide numerically.
    d0_product = _data_start_row(4)
    d0_ew = _data_start_row(len(ew_ref_metrics))
    d1_ew = d0_ew + n_rows - 1
    ws = wb.create_sheet("EW_Combined")
    live = {
        "Return (annualized %)": f"=AVERAGE(D{d0_ew}:D{d1_ew})*252*100",
        "Vol (annualized %)": f"=STDEV(D{d0_ew}:D{d1_ew})*SQRT(252)*100",
        "IR (Sharpe)": f"=AVERAGE(D{d0_ew}:D{d1_ew})/STDEV(D{d0_ew}:D{d1_ew})*SQRT(252)",
        "N days": f"=COUNT(D{d0_ew}:D{d1_ew})",
    }
    _perf_summary_block(ws, ew_ref_metrics, live, n_cols=4)
    headers = ["Date", "Combined_NetPnL_Formula", "Cum_Combined_NetPnL", "Combined_NetPnL"]
    dates = f1_series_by_product[products[0]].index.tolist()

    def row_fn(i):
        r_ew = d0_ew + i
        r_prod = d0_product + i
        refs = ",".join(f"{p}!{net_pnl_col}{r_prod}" for p in products)
        combo_formula = f"=AVERAGE({refs})"
        cum_formula = f"=B{r_ew}" if i == 0 else f"=C{r_ew-1}+B{r_ew}"
        return [dates[i], combo_formula, cum_formula, f"=B{r_ew}"]

    _write_headers_and_rows(ws, headers, row_fn, n_rows, phase_col_idx=None)

    out_path = os.path.join(OUT_DIR, "Momentum_Tradebook.xlsx")
    wb.save(out_path)
    print(f"Saved: {out_path}")
    print(f"EW Portfolio (Python reference): Return={ew_ref_metrics['Return (annualized %)']:.2f}%  "
          f"Vol={ew_ref_metrics['Vol (annualized %)']:.2f}%  IR={ew_ref_metrics['IR (Sharpe)']:.4f}")
    for p in products:
        print(f"  {p}: {ref_metrics_by_product[p]}")
    return out_path



def write_value_sheet(wb, sheet_name: str, curve: pd.DataFrame, f1_series: pd.DataFrame,
                       contract: str, lookback: int, threshold: float, shift_n: int, tc_bps: int,
                       ref_metrics: dict):
    """Value V1 MA-reversion on `contract` (default F8), lookback in trading
    days (default 1260 = 5yr), +-threshold band. min_periods mirrors
    value_v1_position exactly: max(lookback//2, min(lookback,60))."""
    ws = wb.create_sheet(sheet_name)
    idx = f1_series.index
    n_rows = len(idx)
    eff_shift = shift_n + 1
    dates = idx.tolist()
    price = curve[contract].reindex(idx).tolist()
    f1_raw = f1_series["F1_raw"].tolist()
    f2_raw = f1_series["F2_raw"].tolist()
    phase = f1_series["Phase"].tolist()
    min_periods = max(lookback // 2, min(lookback, 60))

    d0 = _data_start_row(len(ref_metrics))
    d1 = d0 + n_rows - 1
    live = {
        "Return (annualized %)": f"=AVERAGE(R{d0}:R{d1})*252*100",
        "Vol (annualized %)": f"=STDEV(R{d0}:R{d1})*SQRT(252)*100",
        "IR (Sharpe)": f"=AVERAGE(R{d0}:R{d1})/STDEV(R{d0}:R{d1})*SQRT(252)",
        "N active days": f"=COUNT(R{d0}:R{d1})",
    }
    _perf_summary_block(ws, ref_metrics, live, n_cols=18)

    headers = ["Date", f"Price_{contract}", "F1_raw", "F2_raw", "Phase", "F1_cont_ratio", "LogPrice",
               f"MA_{lookback}", "Deviation", "Signal", "Position", "LogReturn", "Gross_PnL",
               "Position_Change", "TC_Cost", "Net_PnL", "Cum_Net_PnL", "Active_NetPnL"]

    def row_fn(i):
        r = d0 + i
        vals = [dates[i], price[i], f1_raw[i], f2_raw[i], phase[i]]
        if i == 0:
            vals.append(f"=C{r}")
        else:
            vals.append(f'=IF(E{r}="F1_Tracking",F{r-1}*(C{r}/C{r-1}),'
                        f'IF(E{r}="Bridge",F{r-1}*(C{r}/D{r-1}),'
                        f'F{r-1}*(D{r}/D{r-1})))')
        vals.append(f"=IF(F{r}<=0,\"\",LN(F{r}))")  # G LogPrice
        if i + 1 < min_periods:
            vals.append("")  # H MA -- not enough history yet
        else:
            start = max(d0, r - lookback + 1)
            vals.append(f'=IF(COUNT(B{start}:B{r})>={min_periods},AVERAGE(B{start}:B{r}),"")')
        vals.append(f'=IF(OR(H{r}="",H{r}=0),"",(B{r}-H{r})/H{r})')  # I Deviation
        vals.append(f'=IF(I{r}="","",IF(I{r}<-{threshold},1,IF(I{r}>{threshold},-1,0)))')  # J Signal
        vals.append(_position_formula("J", r, i, eff_shift))  # K Position
        vals.append("" if i == 0 else f"=G{r}-G{r-1}")  # L LogReturn
        vals.append(f'=IF(OR(K{r}="",L{r}=""),0,K{r}*L{r})')  # M Gross_PnL
        vals.append(f"=ABS(K{r})" if i == 0 else f"=ABS(K{r}-K{r-1})")  # N Position_Change
        vals.append(_tc_cost_formula("N", "K", "E", r, i, tc_bps))  # O TC_Cost
        vals.append(f"=M{r}-O{r}")  # P Net_PnL
        vals.append(f"=P{r}" if i == 0 else f"=Q{r-1}+P{r}")  # Q Cum_Net_PnL
        vals.append(f"=IF(K{r}<>0,P{r},\"\")")  # R Active_NetPnL
        return vals

    _write_headers_and_rows(ws, headers, row_fn, n_rows, phase_col_idx=5)
    return ws


def _common_trading_index(f1_series_by_product, products):
    """Intersection of trading dates across all products, after boundary
    truncation. Metals/Precious Metals share one exchange calendar (LME/
    COMEX) so this is a no-op there. Energy/NGL mix exchanges with genuinely
    different regional holiday calendars (NYMEX vs ICE Brent/Gasoil vs Argus
    NGL) -- boundary truncation alone leaves different row counts per product
    (e.g. Energy: Brent/SingaporeGasoil 5289 rows vs WTI/RBOB/HeatingOil/
    NatGas 5160 vs FuelOil 4966, all within the same start/end window), so an
    explicit intersection is needed for the tradebook's row-aligned
    cross-sheet references to line up."""
    idx = f1_series_by_product[products[0]].index
    for p in products[1:]:
        idx = idx.intersection(f1_series_by_product[p].index)
    return idx.sort_values()


def _common_truncated_series(products):
    f1_series_by_product = {p: cfg.load_f1_series_logret(p) for p in products}
    common_start = max(s.index.min() for s in f1_series_by_product.values())
    common_end = min(s.index.max() for s in f1_series_by_product.values())
    for p in products:
        f1_series_by_product[p] = f1_series_by_product[p].loc[common_start:common_end]
    common_idx = _common_trading_index(f1_series_by_product, products)
    for p in products:
        f1_series_by_product[p] = f1_series_by_product[p].loc[common_idx]
    n_rows = len(f1_series_by_product[products[0]])
    for p in products:
        assert len(f1_series_by_product[p]) == n_rows
        assert (f1_series_by_product[p].index == f1_series_by_product[products[0]].index).all()
    return f1_series_by_product, common_start, common_end, n_rows, common_idx


def _rebuild_ratio_series(p, common_start, common_end):
    from ratio_continuous import get_metal_rolling_f1_ratio, reanchor_f1_continuous_ratio
    raw = get_metal_rolling_f1_ratio(cfg.METAL_CODES[p], futures_file=cfg.CURVE_FILE,
                                      calendar_file=cfg.CALENDAR_FILE, config=cfg.METALS_CONFIG, verbose=False)
    raw = raw.loc[common_start:common_end]
    raw = reanchor_f1_continuous_ratio(raw)
    log_price = np.log(raw["F1_continuous_ratio"].where(raw["F1_continuous_ratio"] > 0))
    return raw, log_price


def _ref_metrics_from_position(pos, log_price, phase, tc_bps):
    from engine import log_return_daily
    # Carry/Value positions come from `curve` (a separate load path from
    # `log_price`'s own `raw` DataFrame) -- its date index isn't guaranteed
    # identical to log_price's, unlike Momentum's pos (built directly off
    # `raw["F1_raw"]`). log_return_daily/log_return_metrics already reindex
    # internally, but this function's OWN active-day filter needs the same
    # alignment explicitly, or `net[pos != 0]` raises an unalignable-boolean-
    # -indexer error (caught during the Carry workbook run, 2026-07-24).
    pos = pos.reindex(log_price.index).fillna(0.0)
    _, net = log_return_daily(pos, log_price, tc_bps, phase)
    m = log_return_metrics(pos, log_price, tc_bps, phase)
    active_net = net[pos != 0]
    return {
        "Return (annualized %)": round(active_net.mean() * 252 * 100, 4),
        "Vol (annualized %)": round(active_net.std(ddof=1) * np.sqrt(252) * 100, 4),
        "IR (Sharpe)": round(m["net"], 4),
        "N active days": m["nact"],
    }, net


def _build_ew_sheet(wb, ew_ref_metrics, f1_series_by_product, products, n_rows, net_pnl_col):
    d0_product = _data_start_row(4)
    d0_ew = _data_start_row(len(ew_ref_metrics))
    d1_ew = d0_ew + n_rows - 1
    ws = wb.create_sheet("EW_Combined")
    live = {
        "Return (annualized %)": f"=AVERAGE(D{d0_ew}:D{d1_ew})*252*100",
        "Vol (annualized %)": f"=STDEV(D{d0_ew}:D{d1_ew})*SQRT(252)*100",
        "IR (Sharpe)": f"=AVERAGE(D{d0_ew}:D{d1_ew})/STDEV(D{d0_ew}:D{d1_ew})*SQRT(252)",
        "N days": f"=COUNT(D{d0_ew}:D{d1_ew})",
    }
    _perf_summary_block(ws, ew_ref_metrics, live, n_cols=4)
    headers = ["Date", "Combined_NetPnL_Formula", "Cum_Combined_NetPnL", "Combined_NetPnL"]
    dates = f1_series_by_product[products[0]].index.tolist()

    def row_fn(i):
        r_ew, r_prod = d0_ew + i, d0_product + i
        refs = ",".join(f"{p}!{net_pnl_col}{r_prod}" for p in products)
        return [dates[i], f"=AVERAGE({refs})", f"=B{r_ew}" if i == 0 else f"=C{r_ew-1}+B{r_ew}", f"=B{r_ew}"]

    _write_headers_and_rows(ws, headers, row_fn, n_rows, phase_col_idx=None)


def build_carry_workbook(near: str, far: str, tc_bps: int = 5):
    """Carry_{near}{far}_Tradebook.xlsx: V1+V2 composite Carry on tenor pair
    (near, far), one sheet per product + EW_Combined."""
    products = cfg.PRODUCTS
    f1_series_by_product, common_start, common_end, n_rows, common_idx = _common_truncated_series(products)
    curve_by_product = {p: cfg.load_curve(p).loc[common_start:common_end].loc[common_idx] for p in products}

    wb = Workbook()
    wb.remove(wb.active)
    net_pnl_col = "S"  # Carry composite sheet layout: Net_PnL is column S (19th)

    ref_metrics_by_product = {}
    combined_net_pnl = None
    for p in products:
        raw, log_price = _rebuild_ratio_series(p, common_start, common_end)
        raw = raw.loc[common_idx]
        log_price = log_price.loc[common_idx]
        curve = curve_by_product[p]
        pos = carry_v1v2_composite_position(curve, near, far, cfg.CARRY_ZSCORE_WINDOW, cfg.CARRY_SHIFT_N)
        ref_metrics_by_product[p], net = _ref_metrics_from_position(pos, log_price, raw["Phase"], tc_bps)
        combined_net_pnl = net if combined_net_pnl is None else (combined_net_pnl + net)
        f1_series_by_product[p] = raw
        curve_by_product[p] = curve

    combined_net_pnl = combined_net_pnl / len(products)
    ew_ref_metrics = {
        "Return (annualized %)": round(combined_net_pnl.mean() * 252 * 100, 4),
        "Vol (annualized %)": round(combined_net_pnl.std() * np.sqrt(252) * 100, 4),
        "IR (Sharpe)": round(combined_net_pnl.mean() / combined_net_pnl.std() * np.sqrt(252), 4),
        "N days": len(combined_net_pnl),
    }

    for p in products:
        write_carry_composite_sheet(wb, p, curve_by_product[p], f1_series_by_product[p], near, far,
                                     cfg.CARRY_ZSCORE_WINDOW, cfg.CARRY_SHIFT_N, tc_bps, ref_metrics_by_product[p])

    _build_ew_sheet(wb, ew_ref_metrics, f1_series_by_product, products, n_rows, net_pnl_col)

    out_path = os.path.join(OUT_DIR, f"Carry_{near}{far}_Tradebook.xlsx")
    wb.save(out_path)
    print(f"Saved: {out_path}")
    print(f"EW Portfolio: {ew_ref_metrics}")
    for p in products:
        print(f"  {p}: {ref_metrics_by_product[p]}")
    return out_path


def build_carrymom_workbook(near: str, far: str, tc_bps: int = 5):
    products = cfg.PRODUCTS
    f1_series_by_product, common_start, common_end, n_rows, common_idx = _common_truncated_series(products)
    curve_by_product = {p: cfg.load_curve(p).loc[common_start:common_end].loc[common_idx] for p in products}

    wb = Workbook()
    wb.remove(wb.active)
    net_pnl_col = "R"  # CarryMomentum sheet layout: Net_PnL is column R (18th)

    ref_metrics_by_product = {}
    combined_net_pnl = None
    for p in products:
        raw, log_price = _rebuild_ratio_series(p, common_start, common_end)
        raw = raw.loc[common_idx]
        log_price = log_price.loc[common_idx]
        curve = curve_by_product[p]
        pos = carry_v4_position(curve, cfg.CARRY_MOMENTUM_HORIZON, cfg.CARRY_MOMENTUM_SHIFT_N, near, far)
        ref_metrics_by_product[p], net = _ref_metrics_from_position(pos, log_price, raw["Phase"], tc_bps)
        combined_net_pnl = net if combined_net_pnl is None else (combined_net_pnl + net)
        f1_series_by_product[p] = raw
        curve_by_product[p] = curve

    combined_net_pnl = combined_net_pnl / len(products)
    ew_ref_metrics = {
        "Return (annualized %)": round(combined_net_pnl.mean() * 252 * 100, 4),
        "Vol (annualized %)": round(combined_net_pnl.std() * np.sqrt(252) * 100, 4),
        "IR (Sharpe)": round(combined_net_pnl.mean() / combined_net_pnl.std() * np.sqrt(252), 4),
        "N days": len(combined_net_pnl),
    }

    for p in products:
        write_carrymomentum_sheet(wb, p, curve_by_product[p], f1_series_by_product[p], near, far,
                                   cfg.CARRY_MOMENTUM_HORIZON, cfg.CARRY_MOMENTUM_SHIFT_N, tc_bps,
                                   ref_metrics_by_product[p])

    _build_ew_sheet(wb, ew_ref_metrics, f1_series_by_product, products, n_rows, net_pnl_col)

    out_path = os.path.join(OUT_DIR, f"CarryMom_{near}{far}_Tradebook.xlsx")
    wb.save(out_path)
    print(f"Saved: {out_path}")
    print(f"EW Portfolio: {ew_ref_metrics}")
    for p in products:
        print(f"  {p}: {ref_metrics_by_product[p]}")
    return out_path


def build_value_workbook(tc_bps: int = 5):
    products = cfg.PRODUCTS
    f1_series_by_product, common_start, common_end, n_rows, common_idx = _common_truncated_series(products)
    curve_by_product = {p: cfg.load_curve(p).loc[common_start:common_end].loc[common_idx] for p in products}

    wb = Workbook()
    wb.remove(wb.active)
    net_pnl_col = "P"  # Value sheet layout: Net_PnL is column P (16th)

    ref_metrics_by_product = {}
    combined_net_pnl = None
    for p in products:
        raw, log_price = _rebuild_ratio_series(p, common_start, common_end)
        raw = raw.loc[common_idx]
        log_price = log_price.loc[common_idx]
        curve = curve_by_product[p]
        pos = value_v1_position(curve, cfg.VALUE_CONTRACT, cfg.VALUE_LOOKBACK_DAYS, cfg.VALUE_THRESHOLD,
                                 cfg.VALUE_SHIFT_N)
        ref_metrics_by_product[p], net = _ref_metrics_from_position(pos, log_price, raw["Phase"], tc_bps)
        combined_net_pnl = net if combined_net_pnl is None else (combined_net_pnl + net)
        f1_series_by_product[p] = raw
        curve_by_product[p] = curve

    combined_net_pnl = combined_net_pnl / len(products)
    ew_ref_metrics = {
        "Return (annualized %)": round(combined_net_pnl.mean() * 252 * 100, 4),
        "Vol (annualized %)": round(combined_net_pnl.std() * np.sqrt(252) * 100, 4),
        "IR (Sharpe)": round(combined_net_pnl.mean() / combined_net_pnl.std() * np.sqrt(252), 4),
        "N days": len(combined_net_pnl),
    }

    for p in products:
        write_value_sheet(wb, p, curve_by_product[p], f1_series_by_product[p], cfg.VALUE_CONTRACT,
                           cfg.VALUE_LOOKBACK_DAYS, cfg.VALUE_THRESHOLD, cfg.VALUE_SHIFT_N, tc_bps,
                           ref_metrics_by_product[p])

    _build_ew_sheet(wb, ew_ref_metrics, f1_series_by_product, products, n_rows, net_pnl_col)

    out_path = os.path.join(OUT_DIR, "Value_Tradebook.xlsx")
    wb.save(out_path)
    print(f"Saved: {out_path}")
    print(f"EW Portfolio: {ew_ref_metrics}")
    for p in products:
        print(f"  {p}: {ref_metrics_by_product[p]}")
    return out_path


if __name__ == "__main__":
    print(f"Asset class: {_ASSET_CLASS}  (products: {cfg.PRODUCTS})")
    print(f"Carry tenor pairs: {cfg.CARRY_TENOR_PAIRS}")
    print(f"Output dir: {OUT_DIR}\n")

    build_momentum_workbook()
    for near, far in cfg.CARRY_TENOR_PAIRS:
        build_carry_workbook(near, far)
    for near, far in cfg.CARRY_TENOR_PAIRS:
        build_carrymom_workbook(near, far)
    build_value_workbook()
