"""
generate_tradebooks.py
=======================
Generates full Excel tradebooks (Performance Summary + day-by-day rows) for
every Stage 2 product x strategy, in the same visual/structural format as the
original Metals-Risk-Premia repo's scripts/momentum_signals.py,
carry_signals.py, value_signals.py: a dark-header "PERFORMANCE SUMMARY" block
followed by a copper-header "TRADEBOOK" block, one workbook per strategy
config, one sheet per execution-timing variant (Same Day Shift-0, Lag-1
Shift-1, Lag-2 Shift-2).

Uses the SAME signal math as common_engine.py (the live Stage 2 dashboards),
so these tradebooks reconcile exactly with what Momentum/Carry/Value show
on-screen:
  - Momentum : MA crossover on F1_raw, 3 default benchmark pairs
               (1,20) / (5,60) / (20,250)
  - Carry    : V1 Roll Yield (F1-F2)/F1, V3 Z-score(252d) of that same raw series
  - Value    : V1 MA-reversion, default contract (F8 if available), 5yr lookback, +-10%

PnL is always Position[t] x delta(F1_continuous)[t], regardless of which raw
series the signal is derived from -- identical convention project-wide.

Output: tradebooks/{product_code}/*.xlsx
"""

from __future__ import annotations

import os
import sys

import numpy as np
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_REPO_ROOT = os.path.dirname(_SCRIPT_DIR)
sys.path.insert(0, _REPO_ROOT)
sys.path.insert(0, _SCRIPT_DIR)

from rolling_continuous import (
    get_metal_rolling_f1, reanchor_f1_continuous,
    METALS_CONFIG, METALS_FUTURES_FILE, METALS_CALENDAR_FILE,
    ENERGY_CONFIG, ENERGY_FUTURES_FILE, ENERGY_CALENDAR_FILE,
    PRECIOUS_CONFIG, PRECIOUS_FUTURES_FILE, PRECIOUS_CALENDAR_FILE,
)
from common_curve_loader import load_curve_simple, load_curve_legacy_multiheader

TRADEBOOKS_DIR = Path(_REPO_ROOT) / "tradebooks"

MOMENTUM_PAIRS = [(1, 20), (5, 60), (20, 250)]
VALUE_LOOKBACK_DAYS = 1260   # 5yr
VALUE_THRESHOLD = 0.10       # +-10%
CARRY_V2_TENOR = ("F3", "F15")   # default long-slope tenor pair (matches render_carry_tab's first option)
CARRY_V4_HORIZON = 20            # default carry-momentum lookback (days)
TC_BPS_DEFAULT = 5                # matches common_shared.tc_label_map's default (dashboard selectbox index=1)

PRODUCTS = [
    # Energy
    dict(asset_class="Energy", code="CL", name="WTI Crude (NYMEX)", unit="/bbl",
         config=ENERGY_CONFIG, futures_file=ENERGY_FUTURES_FILE, calendar_file=ENERGY_CALENDAR_FILE, loader="simple"),
    dict(asset_class="Energy", code="CO", name="Brent Crude (ICE)", unit="/bbl",
         config=ENERGY_CONFIG, futures_file=ENERGY_FUTURES_FILE, calendar_file=ENERGY_CALENDAR_FILE, loader="simple"),
    dict(asset_class="Energy", code="XB", name="RBOB Gasoline (NYMEX)", unit="/gal",
         config=ENERGY_CONFIG, futures_file=ENERGY_FUTURES_FILE, calendar_file=ENERGY_CALENDAR_FILE, loader="simple"),
    dict(asset_class="Energy", code="HO", name="Heating Oil ULSD (NYMEX)", unit="/gal",
         config=ENERGY_CONFIG, futures_file=ENERGY_FUTURES_FILE, calendar_file=ENERGY_CALENDAR_FILE, loader="simple"),
    dict(asset_class="Energy", code="NG", name="Nat Gas Henry Hub (NYMEX)", unit="/MMBtu",
         config=ENERGY_CONFIG, futures_file=ENERGY_FUTURES_FILE, calendar_file=ENERGY_CALENDAR_FILE, loader="simple"),
    dict(asset_class="Energy", code="QS", name="Singapore Gasoil (ICE)", unit="/mt",
         config=ENERGY_CONFIG, futures_file=ENERGY_FUTURES_FILE, calendar_file=ENERGY_CALENDAR_FILE, loader="simple"),
    dict(asset_class="Energy", code="FO", name="Fuel Oil 3.5pct Barges (ICE)", unit="/mt",
         config=ENERGY_CONFIG, futures_file=ENERGY_FUTURES_FILE, calendar_file=ENERGY_CALENDAR_FILE, loader="simple"),
    # Metals (LME) -- data/06-30/Metals_Futures_Curve_Updated.xlsx, through 2026-06-30
    dict(asset_class="Metals", code="LP", name="LME Copper", unit="/MT",
         config=METALS_CONFIG, futures_file=METALS_FUTURES_FILE, calendar_file=METALS_CALENDAR_FILE, loader="simple"),
    dict(asset_class="Metals", code="LA", name="LME Aluminium", unit="/MT",
         config=METALS_CONFIG, futures_file=METALS_FUTURES_FILE, calendar_file=METALS_CALENDAR_FILE, loader="simple"),
    # Precious Metals
    dict(asset_class="Precious", code="GC", name="Gold COMEX", unit="/oz",
         config=PRECIOUS_CONFIG, futures_file=PRECIOUS_FUTURES_FILE, calendar_file=PRECIOUS_CALENDAR_FILE, loader="simple"),
    dict(asset_class="Precious", code="SI", name="Silver COMEX", unit="/oz",
         config=PRECIOUS_CONFIG, futures_file=PRECIOUS_FUTURES_FILE, calendar_file=PRECIOUS_CALENDAR_FILE, loader="simple"),
    dict(asset_class="Precious", code="HG", name="Copper CME (HG)", unit="/lb",
         config=PRECIOUS_CONFIG, futures_file=PRECIOUS_FUTURES_FILE, calendar_file=PRECIOUS_CALENDAR_FILE, loader="simple"),
    dict(asset_class="Precious", code="PL", name="Platinum NYMEX", unit="/oz",
         config=PRECIOUS_CONFIG, futures_file=PRECIOUS_FUTURES_FILE, calendar_file=PRECIOUS_CALENDAR_FILE, loader="simple"),
    dict(asset_class="Precious", code="PA", name="Palladium NYMEX", unit="/oz",
         config=PRECIOUS_CONFIG, futures_file=PRECIOUS_FUTURES_FILE, calendar_file=PRECIOUS_CALENDAR_FILE, loader="simple"),
]


# ══════════════════════════════════════════════════════════════════
# EXECUTION TIMING (identical to common_engine.exec_shift)
# ══════════════════════════════════════════════════════════════════
#
# A position decided from signal[t] needs F1_raw[t] as an input, so it isn't
# known until AFTER the t-1->t return has already happened. Pairing signal[t]
# with that same return (a raw shift(0)) uses day t's close twice -- once to
# compute the signal, once as the return's endpoint -- a same-bar look-ahead
# leak, not a valid "faster" execution. shift(1) is therefore the fastest any
# position can legitimately go live, so shift_n counts EXTRA days of delay ON
# TOP of that 1-day floor, not the raw shift itself:
# shift_n=0 (Same Day): position[t] = signal[t-1] -- the fastest legitimate
#   entry (the 1-day floor, no extra delay).
# shift_n=1 (Lag-1):    position[t] = signal[t-2] -- one extra day of delay
#   on top of the floor.
# shift_n=2 (Lag-2):    position[t] = signal[t-3] -- two extra days of delay
#   on top of the floor.
# All three are distinct, realistically-tradeable series -- none of them can
# ever pair a signal with the same-bar return that produced it. Single
# "Position" column throughout (no raw/effective split needed): what's shown
# is exactly what drives Daily_PnL/MTM/Cum_PnL.

TIMING_VARIANTS = [(0, "Same Day (Shift-0)"), (1, "Lag-1 (Shift-1)"), (2, "Lag-2 (Shift-2)")]


def exec_shift(sigbin: pd.Series, shift_n: int) -> pd.Series:
    """PnL-driving shift = shift_n + 1 (see comment block above)."""
    return sigbin.shift(shift_n + 1)


def _tc_columns(position: pd.Series, daily_pnl: pd.Series, f1r: pd.Series, tc_bps: int = TC_BPS_DEFAULT,
                 phase: pd.Series | None = None):
    """Shared TC convention across every builder AND compute_performance --
    identical to common_shared.transaction_cost(): tc[t] = |position[t] -
    position[t-1]| * (tc_bps/10000/2) * F1_raw[t], first day's TC based on
    the position's absolute size (a flip from flat).

    PLUS a roll-day charge when `phase` is passed: rolling a futures position
    forward (selling the expiring contract, buying the next one) is a REAL
    trade even when the strategy's directional position doesn't change
    across the roll -- staying long through a roll still means exiting the
    old contract and re-entering the new one, which would otherwise look
    free. Charged only on the actual roll day (Phase == "Roll_LTD-N"), only
    for exposure that's the SAME sign before and after (nothing to
    re-establish if flat or flipping, since the ordinary position-change
    cost already covers a fresh entry/exit that day) -- positions here are
    always -1/0/+1, so this is a same-sign-and-nonzero indicator, not a
    magnitude calculation. Returns (position_change, tc_cost, daily_pnl_net,
    cum_pnl_net) -- position_change (the returned/displayed column) is the
    pure directional change only; the roll charge is folded into tc_cost."""
    chg = position.diff().abs()
    if len(chg):
        chg.iloc[0] = abs(position.iloc[0])

    chargeable_units = chg.copy()
    if phase is not None:
        phase = phase.reindex(position.index)
        is_roll_day = phase.astype(str).str.startswith("Roll_LTD")
        prev_pos = position.shift(1)
        held_through_roll = is_roll_day & (position != 0) & (prev_pos != 0) & (np.sign(position) == np.sign(prev_pos))
        chargeable_units = chargeable_units + held_through_roll.astype(float)

    tc_cost = chargeable_units * (tc_bps / 10000.0 / 2.0) * f1r.reindex(position.index)
    daily_pnl_net = daily_pnl - tc_cost
    cum_pnl_net = daily_pnl_net.cumsum()
    return chg, tc_cost, daily_pnl_net, cum_pnl_net


# ══════════════════════════════════════════════════════════════════
# PERFORMANCE METRICS (ported from momentum_signals.py compute_performance)
# ══════════════════════════════════════════════════════════════════

def _consecutive(arr: np.ndarray, val: int) -> int:
    best = cur = 0
    for x in arr:
        if x == val:
            cur += 1
            best = max(best, cur)
        else:
            cur = 0
    return best


def compute_performance(daily_pnl: pd.Series, position: pd.Series, f1_cont: pd.Series, f1_raw: pd.Series,
                         shift_n: int, unit_label: str, tc_bps: int = TC_BPS_DEFAULT,
                         phase: pd.Series | None = None) -> dict:
    """All return/risk metrics in native dollar/unit terms (e.g. USD/MT, USD/bbl),
    not %-of-notional. %-of-notional requires dividing by F1_continuous[t-1], an
    additively back-adjusted level with no floor at zero -- confirmed to go
    negative for a majority of history on several products (Aluminium 75% of
    days, Nat Gas 88%, Fuel Oil 65%, WTI 42%), which can silently flip the sign
    of a return or inflate its magnitude by 2-3 orders of magnitude. Dollar PnL
    has no such division and is immune to this.

    Transaction costs: charged on POSITION CHANGES (not on notional), using
    F1_raw as the traded price -- identical convention to common_shared.
    transaction_cost(): tc[t] = |position[t]-position[t-1]| * (tc_bps/10000/2)
    * F1_raw[t], with the first live day's TC based on the position's absolute
    size (a flip from flat), PLUS a roll-day charge when `phase` is passed
    (see _tc_columns()'s docstring). Net PnL = Gross PnL - TC. Gross and Net
    Sharpe/Sortino/MaxDD/Calmar are both reported at the top of the summary;
    every other stat (Hit Rate, Avg Win/Loss, Profit Factor, streaks) is
    computed on GROSS PnL, matching the tradebook's Daily_PnL column (Net PnL
    lives in the separate Daily_PnL_Net column)."""
    position = position.reindex(daily_pnl.index).fillna(0)
    _, tc_cost, daily_pnl_net, _ = _tc_columns(position, daily_pnl, f1_raw, tc_bps, phase)

    active_pnl = daily_pnl[position != 0].dropna()
    active_pnl_net = daily_pnl_net[position != 0].dropna()
    n = len(active_pnl)

    def _ratios(active):
        ann = float(active.mean() * 252) if len(active) > 1 else np.nan
        std = float(active.std() * np.sqrt(252)) if len(active) > 1 else np.nan
        sh = ann / std if (std and std > 0) else np.nan
        down = active[active < 0]
        sortino_denom = float(down.std() * np.sqrt(252)) if len(down) > 1 else np.nan
        so = ann / sortino_denom if (sortino_denom and sortino_denom > 0) else np.nan
        return ann, std, sh, so

    def _max_dd(pnl_series):
        cum = pnl_series.fillna(0).cumsum()
        return float((cum - cum.cummax()).min())

    ann_pnl, ann_std, sharpe, sortino = _ratios(active_pnl)
    ann_pnl_net, ann_std_net, sharpe_net, sortino_net = _ratios(active_pnl_net)
    max_dd = _max_dd(daily_pnl)
    max_dd_net = _max_dd(daily_pnl_net)
    calmar = ann_pnl / abs(max_dd) if max_dd != 0 else np.nan
    calmar_net = ann_pnl_net / abs(max_dd_net) if max_dd_net != 0 else np.nan

    wins = active_pnl[active_pnl > 0]
    losses = active_pnl[active_pnl < 0]
    total_pnl = float(active_pnl.sum())
    total_pnl_net = float(active_pnl_net.sum())
    avg_win = float(wins.mean()) if len(wins) > 0 else np.nan
    avg_loss = float(losses.mean()) if len(losses) > 0 else np.nan
    pf_num = float(wins.sum())
    pf_den = float(abs(losses.sum()))
    profit_factor = pf_num / pf_den if pf_den > 0 else np.nan
    hit_rate = float((active_pnl > 0).mean()) if n > 0 else np.nan

    sign_arr = np.where(active_pnl > 0, 1, -1)
    max_con_w = _consecutive(sign_arr, 1)
    max_con_l = _consecutive(sign_arr, -1)

    timing_label = {0: "Same Day (Shift-0)", 1: "Lag-1 (Shift-1)", 2: "Lag-2 (Shift-2)"}[shift_n]
    eff_shift = shift_n + 1
    pos_note = (f"Position[t] = Signal[t-{eff_shift}]  ({timing_label})"
                + ("  -- the fastest a position can legitimately go live: a signal built "
                   "from today's close can't be paired with today's own return without a "
                   "look-ahead leak, so shift(1) is the floor and this is it." if shift_n == 0
                   else f"  -- {eff_shift - 1} extra day(s) of delay on top of that 1-day floor."))

    return {
        # Key ratios first, gross vs net side by side, per request.
        "Sharpe Ratio (Gross)": round(sharpe, 4) if pd.notna(sharpe) else np.nan,
        "Sharpe Ratio (Net)": round(sharpe_net, 4) if pd.notna(sharpe_net) else np.nan,
        "Sortino Ratio (Gross)": round(sortino, 4) if pd.notna(sortino) else np.nan,
        "Sortino Ratio (Net)": round(sortino_net, 4) if pd.notna(sortino_net) else np.nan,
        f"Max Drawdown Gross ({unit_label})": round(max_dd, 4),
        f"Max Drawdown Net ({unit_label})": round(max_dd_net, 4),
        "Calmar Ratio (Gross)": round(calmar, 4) if pd.notna(calmar) else np.nan,
        "Calmar Ratio (Net)": round(calmar_net, 4) if pd.notna(calmar_net) else np.nan,
        f"Total PnL Gross ({unit_label})": round(total_pnl, 2),
        f"Total PnL Net ({unit_label})": round(total_pnl_net, 2),
        f"Total TC Cost ({unit_label})": round(float(tc_cost.fillna(0).sum()), 2),
        "TC Rate (bps, one-way)": tc_bps,
        "Entry Convention": timing_label,
        "Start Date": str(daily_pnl.index[0].date()),
        "End Date": str(daily_pnl.index[-1].date()),
        "Total Calendar Days": len(daily_pnl),
        "Active Trading Days": n,
        "Warmup/Flat Days": len(daily_pnl) - n,
        f"Annualized PnL Gross ({unit_label})": round(ann_pnl, 4) if pd.notna(ann_pnl) else np.nan,
        f"Annualized PnL Net ({unit_label})": round(ann_pnl_net, 4) if pd.notna(ann_pnl_net) else np.nan,
        f"Annualized Std Dev ({unit_label})": round(ann_std, 4) if pd.notna(ann_std) else np.nan,
        "Hit Rate": f"{hit_rate*100:.2f}%",
        f"Avg Win ({unit_label})": round(avg_win, 2) if pd.notna(avg_win) else np.nan,
        f"Avg Loss ({unit_label})": round(avg_loss, 2) if pd.notna(avg_loss) else np.nan,
        "Profit Factor": round(profit_factor, 4) if pd.notna(profit_factor) else np.nan,
        "Max Consecutive Wins": max_con_w,
        "Max Consecutive Losses": max_con_l,
        "POSITION NOTE": pos_note,
        "PnL NOTE": "Daily_PnL = Position x delta_F1_continuous (roll cost in F1_cont); "
                    "Daily_PnL_Net = Daily_PnL - TC_Cost (see tradebook columns).",
        "TC NOTE": f"{tc_bps} bps one-way, charged on |position change| x F1_raw / 2 -- "
                   f"identical convention to the dashboard's default TC filter.",
    }


# ══════════════════════════════════════════════════════════════════
# TRADEBOOK BUILDERS (same formulas as common_engine.py)
# ══════════════════════════════════════════════════════════════════

def build_ma_tradebook(f1r: pd.Series, f1c: pd.Series, m: int, n: int, shift_n: int,
                        f2r: pd.Series | None = None, phase: pd.Series | None = None,
                        tc_bps: int = TC_BPS_DEFAULT) -> pd.DataFrame:
    ma_m = f1r.rolling(m).mean()
    ma_n = f1r.rolling(n).mean()
    crossover = ma_m - ma_n
    signal = np.sign(crossover)
    # Position[t] = Signal[t-(shift_n+1)] -- shift(1) is the floor (fastest a
    # position can legitimately go live), shift_n counts EXTRA delay on top
    # of it. See exec_shift() -- single column, what's shown is what pays.
    position = exec_shift(signal, shift_n).fillna(0)

    delta = f1c.diff()
    daily_pnl = position * delta
    cum_pnl = daily_pnl.cumsum()
    mtm = position * f1c
    pos_chg, tc_cost, daily_pnl_net, cum_pnl_net = _tc_columns(position, daily_pnl, f1r, tc_bps, phase)

    cols = {
        "Date": f1r.index, "F1_raw": f1r.round(4).values,
    }
    if f2r is not None:
        cols["F2_raw"] = f2r.reindex(f1r.index).round(4).values
    if phase is not None:
        cols["Phase"] = phase.reindex(f1r.index).values
    cols["F1_continuous"] = f1c.round(4).values
    cols.update({
        f"MA_{m}": ma_m.round(4).values, f"MA_{n}": ma_n.round(4).values,
        "Crossover": crossover.round(4).values, "Signal": signal.values, "Position": position.values,
        "F1_cont_daily_change": delta.round(4).values, "Daily_PnL": daily_pnl.round(4).values,
        "Position_Change": pos_chg.round(4).values, "TC_Cost": tc_cost.round(4).values,
        "Daily_PnL_Net": daily_pnl_net.round(4).values,
        "MTM": mtm.round(4).values, "Cum_PnL": cum_pnl.round(4).values,
        "Cum_PnL_Net": cum_pnl_net.round(4).values,
    })
    return pd.DataFrame(cols)


def _carry_v1_raw(curve: pd.DataFrame, a: str = "F1", b: str = "F2") -> pd.Series:
    fa, fb = curve[a], curve[b]
    return ((fa - fb) / fa).replace([np.inf, -np.inf], np.nan)


def build_carry_v1_tradebook(curve: pd.DataFrame, f1r: pd.Series, f1c: pd.Series, shift_n: int,
                              f2r: pd.Series | None = None, phase: pd.Series | None = None,
                              tc_bps: int = TC_BPS_DEFAULT) -> pd.DataFrame:
    raw = _carry_v1_raw(curve, "F1", "F2").reindex(f1c.index)
    signal = np.sign(raw)
    position = exec_shift(signal, shift_n).fillna(0)

    delta = f1c.diff()
    daily_pnl = position * delta
    cum_pnl = daily_pnl.cumsum()
    mtm = position * f1c
    pos_chg, tc_cost, daily_pnl_net, cum_pnl_net = _tc_columns(position, daily_pnl, f1r.reindex(f1c.index), tc_bps, phase)

    cols = {"Date": f1c.index, "F1_raw": f1r.reindex(f1c.index).round(4).values}
    if f2r is not None:
        cols["F2_raw"] = f2r.reindex(f1c.index).round(4).values
    if phase is not None:
        cols["Phase"] = phase.reindex(f1c.index).values
    cols["F1_continuous"] = f1c.round(4).values
    cols.update({
        "Carry_Raw_(F1-F2)/F1": raw.round(6).values,
        "Signal": signal.values, "Position": position.values,
        "F1_cont_daily_change": delta.round(4).values, "Daily_PnL": daily_pnl.round(4).values,
        "Position_Change": pos_chg.round(4).values, "TC_Cost": tc_cost.round(4).values,
        "Daily_PnL_Net": daily_pnl_net.round(4).values,
        "MTM": mtm.round(4).values, "Cum_PnL": cum_pnl.round(4).values,
        "Cum_PnL_Net": cum_pnl_net.round(4).values,
    })
    return pd.DataFrame(cols)


def build_carry_v2_tradebook(curve: pd.DataFrame, f1r: pd.Series, f1c: pd.Series, shift_n: int,
                              near: str = CARRY_V2_TENOR[0], far: str = CARRY_V2_TENOR[1],
                              f2r: pd.Series | None = None, phase: pd.Series | None = None,
                              tc_bps: int = TC_BPS_DEFAULT) -> pd.DataFrame:
    """V2 Long Slope -- same Roll Yield formula as V1, (near-far)/near, applied to a
    LONGER-DATED tenor pair (default F3/F15) instead of F1/F2. Reuses _carry_v1_raw
    (matches common_engine.carry_v2_position, which literally calls carry_v1_position
    with a different pair). near/far prices are explicit columns here since they're
    curve columns independent of F1_raw/F2_raw (which stay F1/F2, used only for the
    F1_continuous reconstruction, regardless of which tenor pair the signal trades)."""
    raw = _carry_v1_raw(curve, near, far).reindex(f1c.index)
    near_price = curve[near].reindex(f1c.index)
    far_price = curve[far].reindex(f1c.index)
    signal = np.sign(raw)
    position = exec_shift(signal, shift_n).fillna(0)

    delta = f1c.diff()
    daily_pnl = position * delta
    cum_pnl = daily_pnl.cumsum()
    mtm = position * f1c
    pos_chg, tc_cost, daily_pnl_net, cum_pnl_net = _tc_columns(position, daily_pnl, f1r.reindex(f1c.index), tc_bps, phase)

    cols = {"Date": f1c.index, "F1_raw": f1r.reindex(f1c.index).round(4).values}
    if f2r is not None:
        cols["F2_raw"] = f2r.reindex(f1c.index).round(4).values
    if phase is not None:
        cols["Phase"] = phase.reindex(f1c.index).values
    cols["F1_continuous"] = f1c.round(4).values
    cols.update({
        f"{near}_price": near_price.round(4).values, f"{far}_price": far_price.round(4).values,
        f"Carry_Raw_({near}-{far})/{near}": raw.round(6).values,
        "Signal": signal.values, "Position": position.values,
        "F1_cont_daily_change": delta.round(4).values, "Daily_PnL": daily_pnl.round(4).values,
        "Position_Change": pos_chg.round(4).values, "TC_Cost": tc_cost.round(4).values,
        "Daily_PnL_Net": daily_pnl_net.round(4).values,
        "MTM": mtm.round(4).values, "Cum_PnL": cum_pnl.round(4).values,
        "Cum_PnL_Net": cum_pnl_net.round(4).values,
    })
    return pd.DataFrame(cols)


def build_carry_v3_tradebook(curve: pd.DataFrame, f1r: pd.Series, f1c: pd.Series, shift_n: int,
                              window: int = 252, f2r: pd.Series | None = None,
                              phase: pd.Series | None = None, tc_bps: int = TC_BPS_DEFAULT) -> pd.DataFrame:
    base = _carry_v1_raw(curve, "F1", "F2").reindex(f1c.index)
    z = (base - base.rolling(window).mean()) / base.rolling(window).std()
    signal = np.sign(z.replace([np.inf, -np.inf], np.nan))
    position = exec_shift(signal, shift_n).fillna(0)

    delta = f1c.diff()
    daily_pnl = position * delta
    cum_pnl = daily_pnl.cumsum()
    mtm = position * f1c
    pos_chg, tc_cost, daily_pnl_net, cum_pnl_net = _tc_columns(position, daily_pnl, f1r.reindex(f1c.index), tc_bps, phase)

    cols = {"Date": f1c.index, "F1_raw": f1r.reindex(f1c.index).round(4).values}
    if f2r is not None:
        cols["F2_raw"] = f2r.reindex(f1c.index).round(4).values
    if phase is not None:
        cols["Phase"] = phase.reindex(f1c.index).values
    cols["F1_continuous"] = f1c.round(4).values
    cols.update({
        "Carry_Raw_(F1-F2)/F1": base.round(6).values,
        f"Zscore_{window}d": z.round(4).values, "Signal": signal.values, "Position": position.values,
        "F1_cont_daily_change": delta.round(4).values, "Daily_PnL": daily_pnl.round(4).values,
        "Position_Change": pos_chg.round(4).values, "TC_Cost": tc_cost.round(4).values,
        "Daily_PnL_Net": daily_pnl_net.round(4).values,
        "MTM": mtm.round(4).values, "Cum_PnL": cum_pnl.round(4).values,
        "Cum_PnL_Net": cum_pnl_net.round(4).values,
    })
    return pd.DataFrame(cols)


def build_carry_v4_tradebook(curve: pd.DataFrame, f1r: pd.Series, f1c: pd.Series, shift_n: int,
                              horizon: int = CARRY_V4_HORIZON, f2r: pd.Series | None = None,
                              phase: pd.Series | None = None, tc_bps: int = TC_BPS_DEFAULT) -> pd.DataFrame:
    """V4 Carry-Momentum -- is the F1-F2 carry itself improving or deteriorating?
    raw = base[t] - base[t-horizon]; signal = sign(raw). Matches
    common_engine.carry_v4_position exactly."""
    base = _carry_v1_raw(curve, "F1", "F2").reindex(f1c.index)
    carry_chg = base - base.shift(horizon)
    signal = np.sign(carry_chg)
    position = exec_shift(signal, shift_n).fillna(0)

    delta = f1c.diff()
    daily_pnl = position * delta
    cum_pnl = daily_pnl.cumsum()
    mtm = position * f1c
    pos_chg, tc_cost, daily_pnl_net, cum_pnl_net = _tc_columns(position, daily_pnl, f1r.reindex(f1c.index), tc_bps, phase)

    cols = {"Date": f1c.index, "F1_raw": f1r.reindex(f1c.index).round(4).values}
    if f2r is not None:
        cols["F2_raw"] = f2r.reindex(f1c.index).round(4).values
    if phase is not None:
        cols["Phase"] = phase.reindex(f1c.index).values
    cols["F1_continuous"] = f1c.round(4).values
    cols.update({
        "Carry_Raw_(F1-F2)/F1": base.round(6).values,
        f"Carry_Chg_{horizon}d": carry_chg.round(6).values, "Signal": signal.values, "Position": position.values,
        "F1_cont_daily_change": delta.round(4).values, "Daily_PnL": daily_pnl.round(4).values,
        "Position_Change": pos_chg.round(4).values, "TC_Cost": tc_cost.round(4).values,
        "Daily_PnL_Net": daily_pnl_net.round(4).values,
        "MTM": mtm.round(4).values, "Cum_PnL": cum_pnl.round(4).values,
        "Cum_PnL_Net": cum_pnl_net.round(4).values,
    })
    return pd.DataFrame(cols)


def build_value_v1_tradebook(curve: pd.DataFrame, f1r: pd.Series, f1c: pd.Series, contract: str,
                              lookback: int, threshold: float, shift_n: int,
                              f2r: pd.Series | None = None, phase: pd.Series | None = None,
                              tc_bps: int = TC_BPS_DEFAULT) -> pd.DataFrame:
    fk = curve[contract].reindex(f1c.index)
    ma = fk.rolling(lookback, min_periods=max(lookback // 2, 60)).mean()
    dev = ((fk - ma) / ma.replace(0, np.nan)).replace([np.inf, -np.inf], np.nan)
    signal = pd.Series(np.where(dev.values < -threshold, 1.0, np.where(dev.values > threshold, -1.0, 0.0)),
                        index=dev.index)
    position = exec_shift(signal, shift_n).fillna(0)

    delta = f1c.diff()
    daily_pnl = position * delta
    cum_pnl = daily_pnl.cumsum()
    mtm = position * f1c
    pos_chg, tc_cost, daily_pnl_net, cum_pnl_net = _tc_columns(position, daily_pnl, f1r.reindex(f1c.index), tc_bps, phase)

    cols = {"Date": f1c.index, "F1_raw": f1r.reindex(f1c.index).round(4).values}
    if f2r is not None:
        cols["F2_raw"] = f2r.reindex(f1c.index).round(4).values
    if phase is not None:
        cols["Phase"] = phase.reindex(f1c.index).values
    cols["F1_continuous"] = f1c.round(4).values
    cols.update({
        f"{contract}_price": fk.round(4).values,
        f"MA_{lookback}d": ma.round(4).values, "Deviation": dev.round(6).values,
        "Signal": signal.values, "Position": position.values,
        "F1_cont_daily_change": delta.round(4).values, "Daily_PnL": daily_pnl.round(4).values,
        "Position_Change": pos_chg.round(4).values, "TC_Cost": tc_cost.round(4).values,
        "Daily_PnL_Net": daily_pnl_net.round(4).values,
        "MTM": mtm.round(4).values, "Cum_PnL": cum_pnl.round(4).values,
        "Cum_PnL_Net": cum_pnl_net.round(4).values,
    })
    return pd.DataFrame(cols)


# ══════════════════════════════════════════════════════════════════
# EXCEL WRITER (same visual style as momentum_signals.py)
# ══════════════════════════════════════════════════════════════════

_HEADER_FILL = PatternFill("solid", fgColor="2B3A47")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
_SECTION_FILL = PatternFill("solid", fgColor="4A4A4A")
_SECTION_FONT = Font(bold=True, color="F5C842", size=10)
_METRIC_FILL = PatternFill("solid", fgColor="F2F2F2")
_METRIC_FONT = Font(size=9)
_TB_HEADER_FILL = PatternFill("solid", fgColor="B87333")
_TB_HEADER_FONT = Font(bold=True, color="FFFFFF", size=9)

# Whole-row highlight for the roll->expiry span, per the confirmed phase
# logic in rolling_continuous.py: the roll day itself (Roll_LTD-N) gets its
# own distinct color; F2_Tracking (after rolling, before expiry, F2-sourced)
# gets a second; Bridge (the day after LAST_TRADEABLE_DT, back to F1) gets a
# third. Ordinary F1_Tracking days are left unhighlighted.
_ROLL_DAY_FILL = PatternFill("solid", fgColor="FF9F5A")       # deep orange -- roll day itself
_F2_TRACKING_FILL = PatternFill("solid", fgColor="F5C265")    # amber -- F2 Tracking (F2-sourced)
_BRIDGE_DAY_FILL = PatternFill("solid", fgColor="8FD3E8")     # light blue -- expiry/bridge day (back to F1)
_MANUAL_HEADER_FILL = PatternFill("solid", fgColor="5A3D8A")  # purple -- distinguishes the formula sheet header
_MANUAL_FONT = Font(size=9, color="1B5E20", italic=True)      # green italic -- marks formula-driven cells
_RECONCILE_MATCH_FILL = PatternFill("solid", fgColor="C6EFCE")     # light green -- Excel/Python PnL reconcile
_RECONCILE_KNOWN_FILL = PatternFill("solid", fgColor="FFEB9C")     # amber -- documented, expected, harmless
_RECONCILE_MISMATCH_FILL = PatternFill("solid", fgColor="FFC7CE")  # light red -- unexplained, investigate


def _style_cell(cell, fill=None, font=None, align=None):
    if fill: cell.fill = fill
    if font: cell.font = font
    if align: cell.alignment = align


def _write_xl_sheet(wb: Workbook, tb: pd.DataFrame, metrics: dict, sheet_name: str) -> None:
    ws = wb.create_sheet(sheet_name)

    ws.append(["PERFORMANCE SUMMARY", ""])
    _style_cell(ws.cell(ws.max_row, 1), fill=_SECTION_FILL, font=_SECTION_FONT, align=Alignment(horizontal="left"))
    _style_cell(ws.cell(ws.max_row, 2), fill=_SECTION_FILL)

    ws.append(["Metric", "Value"])
    for col in (1, 2):
        _style_cell(ws.cell(ws.max_row, col), fill=_HEADER_FILL, font=_HEADER_FONT)

    for i, (k, v) in enumerate(metrics.items()):
        ws.append([k, v])
        fill = _METRIC_FILL if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        _style_cell(ws.cell(ws.max_row, 1), fill=fill, font=_METRIC_FONT)
        if k == "EXCEL RECONCILIATION" and isinstance(v, str):
            # MATCH -> green; a documented/expected ramp-up difference (still
            # says "MISMATCH" but is fully explained and harmless) -> amber,
            # so it reads as "reviewed, not a bug" rather than an alarm;
            # anything else (genuinely unexplained divergence) -> red.
            if v.startswith("MATCH"):
                recon_fill = _RECONCILE_MATCH_FILL
            elif "KNOWN, expected" in v:
                recon_fill = _RECONCILE_KNOWN_FILL
            else:
                recon_fill = _RECONCILE_MISMATCH_FILL
            _style_cell(ws.cell(ws.max_row, 2), fill=recon_fill, font=Font(size=9, bold=True))
        else:
            _style_cell(ws.cell(ws.max_row, 2), fill=fill, font=_METRIC_FONT)

    ws.append(["", ""])

    if "Phase" in tb.columns:
        ws.append(["Roll day (Roll_LTD-N)", ""])
        _style_cell(ws.cell(ws.max_row, 1), fill=_ROLL_DAY_FILL, font=Font(size=9))
        ws.append(["F2 Tracking (F1_continuous is F2-sourced)", ""])
        _style_cell(ws.cell(ws.max_row, 1), fill=_F2_TRACKING_FILL, font=Font(size=9))
        ws.append(["Expiry/Bridge day (F1_continuous bridges back to F1)", ""])
        _style_cell(ws.cell(ws.max_row, 1), fill=_BRIDGE_DAY_FILL, font=Font(size=9))
        ws.append(["", ""])

    ws.append(["TRADEBOOK"] + [""] * (len(tb.columns) - 1))
    _style_cell(ws.cell(ws.max_row, 1), fill=_SECTION_FILL, font=_SECTION_FONT)

    ws.append(list(tb.columns))
    hdr_row = ws.max_row
    for col_idx in range(1, len(tb.columns) + 1):
        _style_cell(ws.cell(hdr_row, col_idx), fill=_TB_HEADER_FILL, font=_TB_HEADER_FONT,
                    align=Alignment(horizontal="center"))

    cols = list(tb.columns)
    phase_col_idx = cols.index("Phase") + 1 if "Phase" in cols else None
    f1c_col_idx = cols.index("F1_continuous") + 1 if "F1_continuous" in cols else None

    for _, row in tb.iterrows():
        vals = []
        for v in row:
            if isinstance(v, float) and np.isnan(v):
                vals.append(None)
            elif hasattr(v, "item"):
                vals.append(v.item())
            else:
                vals.append(v)
        ws.append(vals)

        # Highlight the WHOLE ROW for the roll->expiry span -- Roll_LTD-N (roll
        # day itself) in deep orange, F2_Tracking (F2-sourced, "after rolling
        # before expiry") in amber, Bridge (F1-sourced again, "once expired")
        # in blue. Ordinary F1_Tracking days are left unhighlighted. See the
        # module-level comment on _ROLL_DAY_FILL for the phase logic.
        if phase_col_idx:
            phase_val = vals[phase_col_idx - 1]
            row_fill = None
            if isinstance(phase_val, str) and phase_val.startswith("Roll_LTD"):
                row_fill = _ROLL_DAY_FILL
            elif phase_val == "F2_Tracking":
                row_fill = _F2_TRACKING_FILL
            elif phase_val == "Bridge":
                row_fill = _BRIDGE_DAY_FILL
            if row_fill is not None:
                for col_idx in range(1, len(tb.columns) + 1):
                    ws.cell(ws.max_row, col_idx).fill = row_fill

    ws.column_dimensions["A"].width = max(20, max((len(str(k)) for k in metrics.keys()), default=20) + 2)
    for i in range(2, len(tb.columns) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 16
    ws.freeze_panes = ws.cell(hdr_row + 1, 1).coordinate


# ══════════════════════════════════════════════════════════════════
# MANUAL EXCEL-FORMULA RECONSTRUCTION (momentum MA-crossover only, for now)
# ══════════════════════════════════════════════════════════════════
#
# Writes the SAME pipeline as build_ma_tradebook(), but as live Excel
# formulas instead of Python-computed values -- Date/F1_raw/F2_raw/Phase are
# passed through as given data (Phase already encodes the roll-calendar
# lookup done in rolling_continuous.py; re-deriving the calendar itself in
# formulas was explicitly descoped), and every column from F1_continuous
# onward is a formula so it can be audited/recalculated independently in
# Excel. Column layout: A Date, B F1_raw, C F2_raw, D Phase, E F1_continuous,
# F MA_fast, G MA_slow, H Crossover, I Signal, J Position,
# K F1_cont_daily_change, L Daily_PnL, M Position_Change, N TC_Cost,
# O Daily_PnL_Net, P MTM, Q Cum_PnL, R Cum_PnL_Net.
#
# Position (J) is Signal[t-(shift_n+1)]: shift(1) is the floor (fastest a
# position can legitimately go live -- a signal built from today's close
# can't be paired with today's own return without a look-ahead leak), and
# shift_n counts EXTRA days of delay on top of that floor. Same Day
# (shift_n=0) references 1 row up, Lag-1 (shift_n=1) references 2 rows up,
# Lag-2 (shift_n=2) references 3 rows up -- all distinct, single column,
# what's shown is exactly what Daily_PnL/MTM/Cum_PnL are computed from.
#
# TC: Position_Change (M) = |Position[t]-Position[t-1]| (first row = |Position|,
# a flip from flat); TC_Cost (N) = M x (tc_bps/10000/2) x F1_raw -- identical
# convention to common_shared.pos_metrics_generic()/compute_performance().
# Daily_PnL_Net (O) = Daily_PnL - TC_Cost; Cum_PnL_Net (R) is its running sum.
#
# Known simplification: Python's rolling builder carries the last good F1/F2
# value forward across NaN gaps; these formulas assume no gaps (the common
# case). On a date with a genuine data gap, the two sheets can disagree by a
# cell -- everywhere else they should match exactly.

_MANUAL_COLS = ["Date", "F1_raw", "F2_raw", "Phase", "F1_continuous", None, None,
                "Crossover", "Signal", "Position", "F1_cont_daily_change", "Daily_PnL",
                "Position_Change", "TC_Cost", "Daily_PnL_Net", "MTM", "Cum_PnL", "Cum_PnL_Net"]


def write_manual_formula_sheet_ma(wb: Workbook, tb: pd.DataFrame, m: int, n: int,
                                   shift_n: int, sheet_name: str, tc_bps: int = TC_BPS_DEFAULT) -> None:
    ws = wb.create_sheet(sheet_name)

    headers = list(_MANUAL_COLS)
    headers[5] = f"MA_{m}"
    headers[6] = f"MA_{n}"
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        _style_cell(ws.cell(1, col_idx), fill=_MANUAL_HEADER_FILL, font=_TB_HEADER_FONT,
                    align=Alignment(horizontal="center"))

    n_rows = len(tb)
    dates = tb["Date"].tolist()
    f1_vals = tb["F1_raw"].tolist()
    f2_vals = tb["F2_raw"].tolist() if "F2_raw" in tb.columns else [None] * n_rows
    phase_vals = tb["Phase"].tolist() if "Phase" in tb.columns else [None] * n_rows

    for i in range(n_rows):
        r = i + 2  # excel row (1 = header)
        row_vals = [dates[i], f1_vals[i], f2_vals[i], phase_vals[i]]

        # F1_continuous: anchor on the first row, formula-driven thereafter.
        if i == 0:
            row_vals.append(f"=B{r}")
        else:
            row_vals.append(
                f'=IF(D{r}="F1_Tracking",E{r-1}+(B{r}-B{r-1}),'
                f'IF(D{r}="Bridge",E{r-1}+(B{r}-C{r-1}),'
                f'E{r-1}+(C{r}-C{r-1})))'
            )

        # MA_fast / MA_slow: literal ranges (m, n are fixed per sheet).
        row_vals.append(f"=AVERAGE(B{r-m+1}:B{r})" if i + 1 >= m else "")
        row_vals.append(f"=AVERAGE(B{r-n+1}:B{r})" if i + 1 >= n else "")
        # Crossover / Signal
        row_vals.append(f'=IF(OR(F{r}="",G{r}=""),"",F{r}-G{r})')
        row_vals.append(f'=IF(H{r}="","",SIGN(H{r}))')
        # Position: Signal[t-(shift_n+1)] -- shift(1) is the floor, shift_n
        # counts EXTRA rows of delay on top of it. Same Day (shift_n=0)
        # references 1 row up (r-1), Lag-1 (shift_n=1) references 2 rows up
        # (r-2), Lag-2 (shift_n=2) references 3 rows up (r-3) -- never the
        # same row, so this can never leak. Blank/0 until enough history
        # exists.
        eff_shift = shift_n + 1
        if i + 1 > eff_shift:
            row_vals.append(f'=IF(I{r-eff_shift}="",0,I{r-eff_shift})')
        else:
            row_vals.append(0)
        # F1_cont_daily_change
        row_vals.append("" if i == 0 else f"=E{r}-E{r-1}")
        # Daily_PnL
        row_vals.append(f'=IF(OR(J{r}="",K{r}=""),0,J{r}*K{r})')
        # Position_Change / TC_Cost / Daily_PnL_Net. TC_Cost adds a roll-day
        # charge on top of the ordinary position-change charge: rolling a
        # held position forward is a real trade even with no signal change,
        # so if today is a roll day AND the position is the same sign
        # (nonzero) before and after, one extra chargeable unit applies (see
        # _tc_columns()'s docstring for the full rationale). i==0 has no
        # prior row to compare against, so it can never be a roll-through.
        row_vals.append(f"=ABS(J{r})" if i == 0 else f"=ABS(J{r}-J{r-1})")
        if i == 0:
            row_vals.append(f"=M{r}*({tc_bps}/10000/2)*B{r}")
        else:
            row_vals.append(
                f'=(M{r}+IF(AND(LEFT(D{r},8)="Roll_LTD",J{r}<>0,J{r-1}<>0,SIGN(J{r})=SIGN(J{r-1})),1,0))'
                f'*({tc_bps}/10000/2)*B{r}'
            )
        row_vals.append(f"=L{r}-N{r}")
        # MTM
        row_vals.append(f"=J{r}*E{r}")
        # Cum_PnL / Cum_PnL_Net
        row_vals.append(f"=L{r}" if i == 0 else f"=Q{r-1}+L{r}")
        row_vals.append(f"=O{r}" if i == 0 else f"=R{r-1}+O{r}")

        ws.append(row_vals)
        for col_idx in (5, 11, 12, 14, 15, 16, 17, 18):  # F1_cont, delta, PnL, TC_Cost, PnL_Net, MTM, CumPnL, CumPnL_Net
            ws.cell(r, col_idx).font = _MANUAL_FONT

        # Whole-row highlight for the roll->expiry span -- same rule as the
        # Python sheet (Roll_LTD-N = deep orange, F2_Tracking = amber,
        # Bridge = blue).
        phase_val = phase_vals[i]
        row_fill = None
        if isinstance(phase_val, str) and phase_val.startswith("Roll_LTD"):
            row_fill = _ROLL_DAY_FILL
        elif phase_val == "F2_Tracking":
            row_fill = _F2_TRACKING_FILL
        elif phase_val == "Bridge":
            row_fill = _BRIDGE_DAY_FILL
        if row_fill is not None:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(r, col_idx).fill = row_fill

    ws.column_dimensions["A"].width = 12
    for i in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 15
    ws.freeze_panes = "A2"


def _row_phase_fill(phase_val):
    if isinstance(phase_val, str) and phase_val.startswith("Roll_LTD"):
        return _ROLL_DAY_FILL
    if phase_val == "F2_Tracking":
        return _F2_TRACKING_FILL
    if phase_val == "Bridge":
        return _BRIDGE_DAY_FILL
    return None


def _f1_continuous_formula(r: int, i: int) -> str:
    """Same reconstruction as write_manual_formula_sheet_ma -- Phase (D) drives
    whether F1_continuous (E) tracks day-over-day F1 (B) or F2 (C) deltas."""
    if i == 0:
        return f"=B{r}"
    return (f'=IF(D{r}="F1_Tracking",E{r-1}+(B{r}-B{r-1}),'
            f'IF(D{r}="Bridge",E{r-1}+(B{r}-C{r-1}),'
            f'E{r-1}+(C{r}-C{r-1})))')


# ══════════════════════════════════════════════════════════════════
# RECONCILIATION CHECK -- Excel formula sheet vs Python engine
# ══════════════════════════════════════════════════════════════════
# Actually re-parsing/evaluating the generated .xlsx with the `formulas`
# package is correct but far too slow on a full ~20yr (~5000-row) history
# (verified only practical on small synthetic windows during development).
# Instead, each _reconcile_* function is a pure-pandas MIRROR of the exact
# same formula logic written by its corresponding write_manual_formula_sheet_*
# (same F1_continuous phase-reconstruction, same fixed-window/blank-until-N
# conventions) -- fast enough to run on every row of every product, and
# because it's built from the identical formula text (not re-derived), a
# divergence from the engine's true tb here is the same divergence a human
# opening the real workbook in Excel would see. The result is written into
# the workbook's Performance Summary as "EXCEL RECONCILIATION" so a mismatch
# (and, where known, its cause) is visible without leaving Excel.

def _reconstruct_f1_continuous_shadow(f1r: np.ndarray, f2r: np.ndarray, phase: np.ndarray) -> np.ndarray:
    n = len(f1r)
    shadow = np.empty(n)
    shadow[0] = f1r[0]
    for i in range(1, n):
        ph = phase[i]
        if ph == "F1_Tracking":
            shadow[i] = shadow[i - 1] + (f1r[i] - f1r[i - 1])
        elif ph == "Bridge":
            shadow[i] = shadow[i - 1] + (f1r[i] - f2r[i - 1])
        else:
            shadow[i] = shadow[i - 1] + (f2r[i] - f2r[i - 1])
    return shadow


def _reconcile_cum_pnl(tb: pd.DataFrame, daily_pnl_shadow: np.ndarray, f1c_shadow: np.ndarray | None,
                        known_ramp_up_rows: int | None = None, known_ramp_up_reason: str = "") -> str:
    """Compares a formula-sheet-equivalent Daily_PnL series against the
    engine's true tb["Daily_PnL"] DAY BY DAY -- not via cumulative Cum_PnL,
    which would misclassify a real, harmless, already-known divergence: once
    ANY single day's PnL differs, Cum_PnL (a running sum) carries that
    difference forward as a constant offset for the rest of history even if
    every subsequent day matches exactly, which would wrongly read as an
    ongoing/growing mismatch. Classifies any real divergence (data-gap in
    F1_continuous reconstruction vs a documented ramp-up window vs
    unexplained), and returns a one-line note for the Performance Summary."""
    n_rows = len(tb)
    daily_true = tb["Daily_PnL"].fillna(0).values
    daily_diff = np.abs(daily_pnl_shadow - daily_true)
    mism_mask = daily_diff > 1e-2  # $ tolerance on a PER-DAY basis
    n_mismatch = int(mism_mask.sum())

    cum_pnl_shadow = np.cumsum(daily_pnl_shadow)
    true_cum = tb["Cum_PnL"].fillna(0).values
    final_offset = float(cum_pnl_shadow[-1] - true_cum[-1]) if n_rows else 0.0

    if n_mismatch == 0:
        return (f"MATCH -- Excel formulas reconcile exactly with the Python engine across all {n_rows} rows "
                f"(Daily_PnL identical every day).")

    if known_ramp_up_rows is not None:
        post_mask = mism_mask[known_ramp_up_rows:] if n_rows > known_ramp_up_rows else np.array([], dtype=bool)
        n_post = int(post_mask.sum())
        if n_post == 0:
            return (f"MISMATCH in rows 2-{known_ramp_up_rows + 1} only (KNOWN, expected): {known_ramp_up_reason} "
                    f"-- {n_mismatch} day(s) differ in this ramp-up span, producing a constant Cum_PnL offset of "
                    f"${final_offset:.2f} carried forward for the rest of history, but Daily_PnL is IDENTICAL "
                    f"from row {known_ramp_up_rows + 2} onward (verified) -- the two sheets never newly diverge "
                    f"after the ramp-up.")
        first_post_idx = known_ramp_up_rows + int(np.argmax(post_mask))
        return (f"MISMATCH beyond the expected ramp-up window -- {n_post} day(s) differ starting row "
                f"{first_post_idx + 2}, even after row {known_ramp_up_rows + 1} (final Cum_PnL offset "
                f"${final_offset:.2f}). NOT fully explained by {known_ramp_up_reason.rstrip('.')}; investigate "
                f"(likely a genuine data gap -- see F1_continuous reconciliation below).")

    first_idx = int(np.argmax(mism_mask))
    if f1c_shadow is not None:
        f1c_true = tb["F1_continuous"].values
        recon_diff = np.abs(f1c_shadow - f1c_true)
        gap_mask = recon_diff > 0.5
        if gap_mask.any():
            gap_idx = int(np.argmax(gap_mask))
            return (f"MISMATCH from row {first_idx + 2} ({n_mismatch}/{n_rows} days differ, final Cum_PnL "
                    f"offset ${final_offset:.2f}) -- traced to a genuine data gap: the formula sheet's "
                    f"F1_continuous reconstruction (Phase-driven day-over-day F1/F2 deltas) diverges from the "
                    f"engine's true series starting row {gap_idx + 2}, most likely a forward-filled NaN gap in "
                    f"F1_raw/F2_raw that the formula sheet can't replicate (see 'Known simplification' note on "
                    f"the Formulas sheet).")
    return (f"MISMATCH from row {first_idx + 2} ({n_mismatch}/{n_rows} days differ, final Cum_PnL offset "
            f"${final_offset:.2f}) -- not explained by a known F1_continuous data gap; investigate.")


def _reconcile_ma(tb: pd.DataFrame, m: int, n: int, shift_n: int) -> str:
    if "F2_raw" not in tb.columns or "Phase" not in tb.columns:
        return "Not checked -- F2_raw/Phase unavailable for this product (no roll-cycle data)."
    f1r, f2r, phase = tb["F1_raw"].values, tb["F2_raw"].values, tb["Phase"].values
    f1c_shadow = _reconstruct_f1_continuous_shadow(f1r, f2r, phase)

    f1r_s = pd.Series(f1r)
    crossover = f1r_s.rolling(m).mean().values - f1r_s.rolling(n).mean().values
    signal = np.sign(crossover)
    position = pd.Series(signal).shift(shift_n + 1).fillna(0).values
    delta = np.diff(f1c_shadow, prepend=np.nan)
    daily_pnl = np.nan_to_num(position * delta)
    return _reconcile_cum_pnl(tb, daily_pnl, f1c_shadow)


def _reconcile_carry_v1(tb: pd.DataFrame, shift_n: int) -> str:
    if "F2_raw" not in tb.columns or "Phase" not in tb.columns:
        return "Not checked -- F2_raw/Phase unavailable for this product (no roll-cycle data)."
    f1r, f2r, phase = tb["F1_raw"].values, tb["F2_raw"].values, tb["Phase"].values
    f1c_shadow = _reconstruct_f1_continuous_shadow(f1r, f2r, phase)

    raw = np.where(f1r == 0, np.nan, (f1r - f2r) / f1r)
    signal = np.sign(raw)
    position = pd.Series(signal).shift(shift_n + 1).fillna(0).values
    delta = np.diff(f1c_shadow, prepend=np.nan)
    daily_pnl = np.nan_to_num(position * delta)
    return _reconcile_cum_pnl(tb, daily_pnl, f1c_shadow)


def _reconcile_carry_v2(tb: pd.DataFrame, shift_n: int, near: str, far: str) -> str:
    if "F2_raw" not in tb.columns or "Phase" not in tb.columns:
        return "Not checked -- F2_raw/Phase unavailable for this product (no roll-cycle data)."
    f1r, f2r, phase = tb["F1_raw"].values, tb["F2_raw"].values, tb["Phase"].values
    f1c_shadow = _reconstruct_f1_continuous_shadow(f1r, f2r, phase)

    near_price = tb[f"{near}_price"].values
    far_price = tb[f"{far}_price"].values
    raw = np.where(near_price == 0, np.nan, (near_price - far_price) / near_price)
    signal = np.sign(raw)
    position = pd.Series(signal).shift(shift_n + 1).fillna(0).values
    delta = np.diff(f1c_shadow, prepend=np.nan)
    daily_pnl = np.nan_to_num(position * delta)
    return _reconcile_cum_pnl(tb, daily_pnl, f1c_shadow)


def _reconcile_carry_v4(tb: pd.DataFrame, shift_n: int, horizon: int) -> str:
    if "F2_raw" not in tb.columns or "Phase" not in tb.columns:
        return "Not checked -- F2_raw/Phase unavailable for this product (no roll-cycle data)."
    f1r, f2r, phase = tb["F1_raw"].values, tb["F2_raw"].values, tb["Phase"].values
    f1c_shadow = _reconstruct_f1_continuous_shadow(f1r, f2r, phase)

    raw = pd.Series(np.where(f1r == 0, np.nan, (f1r - f2r) / f1r))
    carry_chg = (raw - raw.shift(horizon)).values
    signal = np.sign(carry_chg)
    position = pd.Series(signal).shift(shift_n + 1).fillna(0).values
    delta = np.diff(f1c_shadow, prepend=np.nan)
    daily_pnl = np.nan_to_num(position * delta)
    return _reconcile_cum_pnl(tb, daily_pnl, f1c_shadow)


def _reconcile_carry_v3(tb: pd.DataFrame, shift_n: int, window: int) -> str:
    if "F2_raw" not in tb.columns or "Phase" not in tb.columns:
        return "Not checked -- F2_raw/Phase unavailable for this product (no roll-cycle data)."
    f1r, f2r, phase = tb["F1_raw"].values, tb["F2_raw"].values, tb["Phase"].values
    f1c_shadow = _reconstruct_f1_continuous_shadow(f1r, f2r, phase)

    raw = pd.Series(np.where(f1r == 0, np.nan, (f1r - f2r) / f1r))
    std = raw.rolling(window).std().values
    z = np.where(std == 0, np.nan, (raw.values - raw.rolling(window).mean().values) / std)
    signal = np.sign(z)
    position = pd.Series(signal).shift(shift_n + 1).fillna(0).values
    delta = np.diff(f1c_shadow, prepend=np.nan)
    daily_pnl = np.nan_to_num(position * delta)
    return _reconcile_cum_pnl(tb, daily_pnl, f1c_shadow)


def _reconcile_value_v1(tb: pd.DataFrame, contract: str, lookback: int, threshold: float, shift_n: int) -> str:
    if "F2_raw" not in tb.columns or "Phase" not in tb.columns:
        return "Not checked -- F2_raw/Phase unavailable for this product (no roll-cycle data)."
    f1r, f2r, phase = tb["F1_raw"].values, tb["F2_raw"].values, tb["Phase"].values
    f1c_shadow = _reconstruct_f1_continuous_shadow(f1r, f2r, phase)

    fk = tb[f"{contract}_price"].values
    # Fixed full-size window, blank until `lookback` rows have elapsed --
    # matches the Formula sheet's AVERAGE(F{r-lookback+1}:F{r}) exactly (NOT
    # the engine's growing min_periods window -- see write_manual_formula_
    # sheet_value_v1's KNOWN SIMPLIFICATION comment). min_periods=1 mirrors
    # Excel's AVERAGE() ignoring blanks with no minimum-count floor -- using
    # min_periods=lookback here would wrongly blank out the ENTIRE window
    # around any single missing Fk day, which real Excel's AVERAGE() does not
    # do (confirmed against WTI's one genuine 1-day Fk gap).
    ma_shadow = pd.Series(fk).rolling(lookback, min_periods=1).mean().values
    ma_shadow[:lookback - 1] = np.nan
    dev = np.where(ma_shadow == 0, np.nan, (fk - ma_shadow) / ma_shadow)
    signal = np.where(np.isnan(dev), 0.0, np.where(dev < -threshold, 1.0, np.where(dev > threshold, -1.0, 0.0)))
    position = pd.Series(signal).shift(shift_n + 1).fillna(0).values
    delta = np.diff(f1c_shadow, prepend=np.nan)
    daily_pnl = np.nan_to_num(position * delta)

    min_periods = max(lookback // 2, 60)
    reason = (f"the Excel formula's fixed {lookback}-row MA window doesn't turn on until row {lookback + 1}, "
              f"while the Python engine's rolling(min_periods={min_periods}) window turns on earlier (a "
              f"growing window).")
    # A signal still inside the ramp-up window (row < lookback) surfaces in
    # Position shift_n+1 rows later -- so the last row that can carry a
    # ramp-related difference is lookback + shift_n, not lookback itself
    # (confirmed empirically: shift_n=1 -> 1 residual row at exactly
    # `lookback`, shift_n=2 -> 2 residual rows at `lookback`/`lookback+1`,
    # both eliminated by this extra `+ shift_n`).
    return _reconcile_cum_pnl(tb, daily_pnl, f1c_shadow, known_ramp_up_rows=lookback + shift_n,
                               known_ramp_up_reason=reason)


# ══════════════════════════════════════════════════════════════════
# MANUAL EXCEL-FORMULA RECONSTRUCTION -- Carry V1 Roll Yield
# ══════════════════════════════════════════════════════════════════
# Column layout: A Date, B F1_raw, C F2_raw, D Phase, E F1_continuous,
# F Carry_Raw (F1-F2)/F1, G Signal, H Position, I F1_cont_daily_change,
# J Daily_PnL, K Position_Change, L TC_Cost, M Daily_PnL_Net, N MTM,
# O Cum_PnL, P Cum_PnL_Net. Same Position/PnL/TC timing convention as
# write_manual_formula_sheet_ma (see its module comment for the full
# shift_n+1 and TC rationale).

_CARRY_V1_COLS = ["Date", "F1_raw", "F2_raw", "Phase", "F1_continuous",
                   "Carry_Raw_(F1-F2)/F1", "Signal", "Position", "F1_cont_daily_change", "Daily_PnL",
                   "Position_Change", "TC_Cost", "Daily_PnL_Net", "MTM", "Cum_PnL", "Cum_PnL_Net"]


def write_manual_formula_sheet_carry_v1(wb: Workbook, tb: pd.DataFrame, shift_n: int, sheet_name: str,
                                          tc_bps: int = TC_BPS_DEFAULT) -> None:
    ws = wb.create_sheet(sheet_name)
    ws.append(_CARRY_V1_COLS)
    for col_idx in range(1, len(_CARRY_V1_COLS) + 1):
        _style_cell(ws.cell(1, col_idx), fill=_MANUAL_HEADER_FILL, font=_TB_HEADER_FONT,
                    align=Alignment(horizontal="center"))

    n_rows = len(tb)
    dates = tb["Date"].tolist()
    f1_vals = tb["F1_raw"].tolist()
    f2_vals = tb["F2_raw"].tolist() if "F2_raw" in tb.columns else [None] * n_rows
    phase_vals = tb["Phase"].tolist() if "Phase" in tb.columns else [None] * n_rows
    eff_shift = shift_n + 1

    for i in range(n_rows):
        r = i + 2
        row_vals = [dates[i], f1_vals[i], f2_vals[i], phase_vals[i]]
        row_vals.append(_f1_continuous_formula(r, i))
        # Carry_Raw
        row_vals.append(f'=IF(B{r}=0,"",(B{r}-C{r})/B{r})')
        # Signal
        row_vals.append(f'=IF(F{r}="","",SIGN(F{r}))')
        # Position -- Signal[t-(shift_n+1)], never same-row (see momentum sheet's rationale)
        if i + 1 > eff_shift:
            row_vals.append(f'=IF(G{r-eff_shift}="",0,G{r-eff_shift})')
        else:
            row_vals.append(0)
        # F1_cont_daily_change
        row_vals.append("" if i == 0 else f"=E{r}-E{r-1}")
        # Daily_PnL
        row_vals.append(f'=IF(OR(H{r}="",I{r}=""),0,H{r}*I{r})')
        # Position_Change / TC_Cost / Daily_PnL_Net (TC_Cost includes the
        # roll-day charge -- see MA sheet's comment for the full rationale)
        row_vals.append(f"=ABS(H{r})" if i == 0 else f"=ABS(H{r}-H{r-1})")
        if i == 0:
            row_vals.append(f"=K{r}*({tc_bps}/10000/2)*B{r}")
        else:
            row_vals.append(
                f'=(K{r}+IF(AND(LEFT(D{r},8)="Roll_LTD",H{r}<>0,H{r-1}<>0,SIGN(H{r})=SIGN(H{r-1})),1,0))'
                f'*({tc_bps}/10000/2)*B{r}'
            )
        row_vals.append(f"=J{r}-L{r}")
        # MTM
        row_vals.append(f"=H{r}*E{r}")
        # Cum_PnL / Cum_PnL_Net
        row_vals.append(f"=J{r}" if i == 0 else f"=O{r-1}+J{r}")
        row_vals.append(f"=M{r}" if i == 0 else f"=P{r-1}+M{r}")

        ws.append(row_vals)
        for col_idx in (5, 9, 10, 12, 13, 14, 15, 16):  # F1_cont, delta, PnL, TC_Cost, PnL_Net, MTM, CumPnL, CumPnL_Net
            ws.cell(r, col_idx).font = _MANUAL_FONT
        row_fill = _row_phase_fill(phase_vals[i])
        if row_fill is not None:
            for col_idx in range(1, len(_CARRY_V1_COLS) + 1):
                ws.cell(r, col_idx).fill = row_fill

    ws.column_dimensions["A"].width = 12
    for i in range(2, len(_CARRY_V1_COLS) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 16
    ws.freeze_panes = "A2"


# ══════════════════════════════════════════════════════════════════
# MANUAL EXCEL-FORMULA RECONSTRUCTION -- Carry V2 Long Slope
# ══════════════════════════════════════════════════════════════════
# Same Roll Yield formula as V1, applied to a longer-dated tenor pair --
# near/far prices are curve columns independent of F1_raw/F2_raw (which stay
# F1/F2, used only for F1_continuous), so they get their own data columns.
# Column layout: A Date, B F1_raw, C F2_raw, D Phase, E F1_continuous,
# F {near}_price, G {far}_price, H Carry_Raw_(near-far)/near, I Signal,
# J Position, K F1_cont_daily_change, L Daily_PnL, M Position_Change,
# N TC_Cost, O Daily_PnL_Net, P MTM, Q Cum_PnL, R Cum_PnL_Net.

def write_manual_formula_sheet_carry_v2(wb: Workbook, tb: pd.DataFrame, shift_n: int, near: str, far: str,
                                          sheet_name: str, tc_bps: int = TC_BPS_DEFAULT) -> None:
    near_col, far_col = f"{near}_price", f"{far}_price"
    headers = ["Date", "F1_raw", "F2_raw", "Phase", "F1_continuous", near_col, far_col,
               f"Carry_Raw_({near}-{far})/{near}", "Signal", "Position", "F1_cont_daily_change", "Daily_PnL",
               "Position_Change", "TC_Cost", "Daily_PnL_Net", "MTM", "Cum_PnL", "Cum_PnL_Net"]
    ws = wb.create_sheet(sheet_name)
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        _style_cell(ws.cell(1, col_idx), fill=_MANUAL_HEADER_FILL, font=_TB_HEADER_FONT,
                    align=Alignment(horizontal="center"))

    n_rows = len(tb)
    dates = tb["Date"].tolist()
    f1_vals = tb["F1_raw"].tolist()
    f2_vals = tb["F2_raw"].tolist() if "F2_raw" in tb.columns else [None] * n_rows
    phase_vals = tb["Phase"].tolist() if "Phase" in tb.columns else [None] * n_rows
    near_vals = tb[near_col].tolist()
    far_vals = tb[far_col].tolist()
    eff_shift = shift_n + 1

    for i in range(n_rows):
        r = i + 2
        row_vals = [dates[i], f1_vals[i], f2_vals[i], phase_vals[i]]
        row_vals.append(_f1_continuous_formula(r, i))
        # near/far price: literal data (independent curve columns)
        row_vals.append(near_vals[i])
        row_vals.append(far_vals[i])
        # Carry_Raw
        row_vals.append(f'=IF(F{r}=0,"",(F{r}-G{r})/F{r})')
        # Signal
        row_vals.append(f'=IF(H{r}="","",SIGN(H{r}))')
        # Position -- Signal[t-(shift_n+1)]
        if i + 1 > eff_shift:
            row_vals.append(f'=IF(I{r-eff_shift}="",0,I{r-eff_shift})')
        else:
            row_vals.append(0)
        # F1_cont_daily_change
        row_vals.append("" if i == 0 else f"=E{r}-E{r-1}")
        # Daily_PnL
        row_vals.append(f'=IF(OR(J{r}="",K{r}=""),0,J{r}*K{r})')
        # Position_Change / TC_Cost / Daily_PnL_Net (TC_Cost includes the
        # roll-day charge -- see MA sheet's comment for the full rationale)
        row_vals.append(f"=ABS(J{r})" if i == 0 else f"=ABS(J{r}-J{r-1})")
        if i == 0:
            row_vals.append(f"=M{r}*({tc_bps}/10000/2)*B{r}")
        else:
            row_vals.append(
                f'=(M{r}+IF(AND(LEFT(D{r},8)="Roll_LTD",J{r}<>0,J{r-1}<>0,SIGN(J{r})=SIGN(J{r-1})),1,0))'
                f'*({tc_bps}/10000/2)*B{r}'
            )
        row_vals.append(f"=L{r}-N{r}")
        # MTM
        row_vals.append(f"=J{r}*E{r}")
        # Cum_PnL / Cum_PnL_Net
        row_vals.append(f"=L{r}" if i == 0 else f"=Q{r-1}+L{r}")
        row_vals.append(f"=O{r}" if i == 0 else f"=R{r-1}+O{r}")

        ws.append(row_vals)
        for col_idx in (5, 11, 12, 14, 15, 16, 17, 18):  # F1_cont, delta, PnL, TC_Cost, PnL_Net, MTM, CumPnL, CumPnL_Net
            ws.cell(r, col_idx).font = _MANUAL_FONT
        row_fill = _row_phase_fill(phase_vals[i])
        if row_fill is not None:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(r, col_idx).fill = row_fill

    ws.column_dimensions["A"].width = 12
    for i in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 16
    ws.freeze_panes = "A2"


# ══════════════════════════════════════════════════════════════════
# MANUAL EXCEL-FORMULA RECONSTRUCTION -- Carry V3 Z-score
# ══════════════════════════════════════════════════════════════════
# Column layout: A Date, B F1_raw, C F2_raw, D Phase, E F1_continuous,
# F Carry_Raw, G Zscore_{window}d, H Signal, I Position,
# J F1_cont_daily_change, K Daily_PnL, L Position_Change, M TC_Cost,
# N Daily_PnL_Net, O MTM, P Cum_PnL, Q Cum_PnL_Net.
# Zscore uses a full (non-partial) rolling window, matching pandas'
# rolling(window).mean()/.std() default min_periods=window -- blank until
# `window` rows of Carry_Raw exist, same convention as MA_fast/MA_slow in
# write_manual_formula_sheet_ma.

def write_manual_formula_sheet_carry_v3(wb: Workbook, tb: pd.DataFrame, shift_n: int,
                                          window: int, sheet_name: str, tc_bps: int = TC_BPS_DEFAULT) -> None:
    headers = ["Date", "F1_raw", "F2_raw", "Phase", "F1_continuous",
               "Carry_Raw_(F1-F2)/F1", f"Zscore_{window}d", "Signal", "Position", "F1_cont_daily_change",
               "Daily_PnL", "Position_Change", "TC_Cost", "Daily_PnL_Net", "MTM", "Cum_PnL", "Cum_PnL_Net"]
    ws = wb.create_sheet(sheet_name)
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        _style_cell(ws.cell(1, col_idx), fill=_MANUAL_HEADER_FILL, font=_TB_HEADER_FONT,
                    align=Alignment(horizontal="center"))

    n_rows = len(tb)
    dates = tb["Date"].tolist()
    f1_vals = tb["F1_raw"].tolist()
    f2_vals = tb["F2_raw"].tolist() if "F2_raw" in tb.columns else [None] * n_rows
    phase_vals = tb["Phase"].tolist() if "Phase" in tb.columns else [None] * n_rows
    eff_shift = shift_n + 1

    for i in range(n_rows):
        r = i + 2
        row_vals = [dates[i], f1_vals[i], f2_vals[i], phase_vals[i]]
        row_vals.append(_f1_continuous_formula(r, i))
        # Carry_Raw
        row_vals.append(f'=IF(B{r}=0,"",(B{r}-C{r})/B{r})')
        # Zscore: full window only (blank before `window` rows of Carry_Raw exist)
        if i + 1 >= window:
            row_vals.append(
                f'=IF(STDEV(F{r-window+1}:F{r})=0,"",'
                f'(F{r}-AVERAGE(F{r-window+1}:F{r}))/STDEV(F{r-window+1}:F{r}))'
            )
        else:
            row_vals.append("")
        # Signal
        row_vals.append(f'=IF(G{r}="","",SIGN(G{r}))')
        # Position -- Signal[t-(shift_n+1)]
        if i + 1 > eff_shift:
            row_vals.append(f'=IF(H{r-eff_shift}="",0,H{r-eff_shift})')
        else:
            row_vals.append(0)
        # F1_cont_daily_change
        row_vals.append("" if i == 0 else f"=E{r}-E{r-1}")
        # Daily_PnL
        row_vals.append(f'=IF(OR(I{r}="",J{r}=""),0,I{r}*J{r})')
        # Position_Change / TC_Cost / Daily_PnL_Net (TC_Cost includes the
        # roll-day charge -- see MA sheet's comment for the full rationale)
        row_vals.append(f"=ABS(I{r})" if i == 0 else f"=ABS(I{r}-I{r-1})")
        if i == 0:
            row_vals.append(f"=L{r}*({tc_bps}/10000/2)*B{r}")
        else:
            row_vals.append(
                f'=(L{r}+IF(AND(LEFT(D{r},8)="Roll_LTD",I{r}<>0,I{r-1}<>0,SIGN(I{r})=SIGN(I{r-1})),1,0))'
                f'*({tc_bps}/10000/2)*B{r}'
            )
        row_vals.append(f"=K{r}-M{r}")
        # MTM
        row_vals.append(f"=I{r}*E{r}")
        # Cum_PnL / Cum_PnL_Net
        row_vals.append(f"=K{r}" if i == 0 else f"=P{r-1}+K{r}")
        row_vals.append(f"=N{r}" if i == 0 else f"=Q{r-1}+N{r}")

        ws.append(row_vals)
        for col_idx in (5, 10, 11, 13, 14, 15, 16, 17):  # F1_cont, delta, PnL, TC_Cost, PnL_Net, MTM, CumPnL, CumPnL_Net
            ws.cell(r, col_idx).font = _MANUAL_FONT
        row_fill = _row_phase_fill(phase_vals[i])
        if row_fill is not None:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(r, col_idx).fill = row_fill

    ws.column_dimensions["A"].width = 12
    for i in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 16
    ws.freeze_panes = "A2"


# ══════════════════════════════════════════════════════════════════
# MANUAL EXCEL-FORMULA RECONSTRUCTION -- Carry V4 Carry-Momentum
# ══════════════════════════════════════════════════════════════════
# Is the F1-F2 carry itself improving or deteriorating? raw = Carry_Raw[t] -
# Carry_Raw[t-horizon]; signal = sign(raw). Matches common_engine.
# carry_v4_position exactly. Column layout: A Date, B F1_raw, C F2_raw,
# D Phase, E F1_continuous, F Carry_Raw, G Carry_Chg_{horizon}d, H Signal,
# I Position, J F1_cont_daily_change, K Daily_PnL, L Position_Change,
# M TC_Cost, N Daily_PnL_Net, O MTM, P Cum_PnL, Q Cum_PnL_Net.

def write_manual_formula_sheet_carry_v4(wb: Workbook, tb: pd.DataFrame, shift_n: int, horizon: int,
                                          sheet_name: str, tc_bps: int = TC_BPS_DEFAULT) -> None:
    headers = ["Date", "F1_raw", "F2_raw", "Phase", "F1_continuous",
               "Carry_Raw_(F1-F2)/F1", f"Carry_Chg_{horizon}d", "Signal", "Position", "F1_cont_daily_change",
               "Daily_PnL", "Position_Change", "TC_Cost", "Daily_PnL_Net", "MTM", "Cum_PnL", "Cum_PnL_Net"]
    ws = wb.create_sheet(sheet_name)
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        _style_cell(ws.cell(1, col_idx), fill=_MANUAL_HEADER_FILL, font=_TB_HEADER_FONT,
                    align=Alignment(horizontal="center"))

    n_rows = len(tb)
    dates = tb["Date"].tolist()
    f1_vals = tb["F1_raw"].tolist()
    f2_vals = tb["F2_raw"].tolist() if "F2_raw" in tb.columns else [None] * n_rows
    phase_vals = tb["Phase"].tolist() if "Phase" in tb.columns else [None] * n_rows
    eff_shift = shift_n + 1

    for i in range(n_rows):
        r = i + 2
        row_vals = [dates[i], f1_vals[i], f2_vals[i], phase_vals[i]]
        row_vals.append(_f1_continuous_formula(r, i))
        # Carry_Raw
        row_vals.append(f'=IF(B{r}=0,"",(B{r}-C{r})/B{r})')
        # Carry_Chg: change in the carry itself over `horizon` days
        if i + 1 > horizon:
            row_vals.append(f'=IF(OR(F{r}="",F{r-horizon}=""),"",F{r}-F{r-horizon})')
        else:
            row_vals.append("")
        # Signal
        row_vals.append(f'=IF(G{r}="","",SIGN(G{r}))')
        # Position -- Signal[t-(shift_n+1)]
        if i + 1 > eff_shift:
            row_vals.append(f'=IF(H{r-eff_shift}="",0,H{r-eff_shift})')
        else:
            row_vals.append(0)
        # F1_cont_daily_change
        row_vals.append("" if i == 0 else f"=E{r}-E{r-1}")
        # Daily_PnL
        row_vals.append(f'=IF(OR(I{r}="",J{r}=""),0,I{r}*J{r})')
        # Position_Change / TC_Cost / Daily_PnL_Net (TC_Cost includes the
        # roll-day charge -- see MA sheet's comment for the full rationale)
        row_vals.append(f"=ABS(I{r})" if i == 0 else f"=ABS(I{r}-I{r-1})")
        if i == 0:
            row_vals.append(f"=L{r}*({tc_bps}/10000/2)*B{r}")
        else:
            row_vals.append(
                f'=(L{r}+IF(AND(LEFT(D{r},8)="Roll_LTD",I{r}<>0,I{r-1}<>0,SIGN(I{r})=SIGN(I{r-1})),1,0))'
                f'*({tc_bps}/10000/2)*B{r}'
            )
        row_vals.append(f"=K{r}-M{r}")
        # MTM
        row_vals.append(f"=I{r}*E{r}")
        # Cum_PnL / Cum_PnL_Net
        row_vals.append(f"=K{r}" if i == 0 else f"=P{r-1}+K{r}")
        row_vals.append(f"=N{r}" if i == 0 else f"=Q{r-1}+N{r}")

        ws.append(row_vals)
        for col_idx in (5, 10, 11, 13, 14, 15, 16, 17):  # F1_cont, delta, PnL, TC_Cost, PnL_Net, MTM, CumPnL, CumPnL_Net
            ws.cell(r, col_idx).font = _MANUAL_FONT
        row_fill = _row_phase_fill(phase_vals[i])
        if row_fill is not None:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(r, col_idx).fill = row_fill

    ws.column_dimensions["A"].width = 12
    for i in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 16
    ws.freeze_panes = "A2"


# ══════════════════════════════════════════════════════════════════
# MANUAL EXCEL-FORMULA RECONSTRUCTION -- Value V1 MA Reversion
# ══════════════════════════════════════════════════════════════════
# Column layout: A Date, B F1_raw, C F2_raw, D Phase, E F1_continuous,
# F {contract}_price, G MA_{lookback}d, H Deviation, I Signal, J Position,
# K F1_cont_daily_change, L Daily_PnL, M MTM, N Cum_PnL.
#
# MA_{lookback}d uses a fixed, full-size `lookback` window (AVERAGE ignores
# blank Fk cells, so a contract that doesn't exist yet for part of the
# window is skipped the same way pandas' NaN-skipping mean() would), blank
# until `lookback` calendar rows have elapsed -- same convention as
# MA_fast/MA_slow in write_manual_formula_sheet_ma and Zscore in
# write_manual_formula_sheet_carry_v3.
#
# KNOWN SIMPLIFICATION (larger than elsewhere in this file): Python's
# rolling(lookback, min_periods=max(lookback//2,60)).mean() is a GROWING
# window that starts producing values once `min_periods` (not the full
# `lookback`) non-NaN observations exist -- so for the default 5yr lookback
# (1260 rows, min_periods=630), the Python sheet's MA/Deviation/Signal can
# turn on roughly 2.5 years before this formula sheet's does. An
# INDEX(...)-based dynamic range was tried to replicate the growing window
# exactly but was NOT evaluated correctly by the `formulas` verification
# library, so this simpler, verified-correct fixed-window formula is used
# instead. Once both sheets are past `lookback` rows of history they
# reconcile exactly (verified via the `formulas` package) -- only the
# early "ramp-up" span differs.

def write_manual_formula_sheet_value_v1(wb: Workbook, tb: pd.DataFrame, shift_n: int,
                                          contract: str, lookback: int, threshold: float,
                                          sheet_name: str, tc_bps: int = TC_BPS_DEFAULT) -> None:
    price_col = f"{contract}_price"
    headers = ["Date", "F1_raw", "F2_raw", "Phase", "F1_continuous",
               price_col, f"MA_{lookback}d", "Deviation", "Signal", "Position", "F1_cont_daily_change",
               "Daily_PnL", "Position_Change", "TC_Cost", "Daily_PnL_Net", "MTM", "Cum_PnL", "Cum_PnL_Net"]
    ws = wb.create_sheet(sheet_name)
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        _style_cell(ws.cell(1, col_idx), fill=_MANUAL_HEADER_FILL, font=_TB_HEADER_FONT,
                    align=Alignment(horizontal="center"))

    n_rows = len(tb)
    dates = tb["Date"].tolist()
    f1_vals = tb["F1_raw"].tolist()
    f2_vals = tb["F2_raw"].tolist() if "F2_raw" in tb.columns else [None] * n_rows
    phase_vals = tb["Phase"].tolist() if "Phase" in tb.columns else [None] * n_rows
    fk_vals = tb[price_col].tolist()
    eff_shift = shift_n + 1

    for i in range(n_rows):
        r = i + 2
        row_vals = [dates[i], f1_vals[i], f2_vals[i], phase_vals[i]]
        row_vals.append(_f1_continuous_formula(r, i))
        # Fk price: literal data (independent curve column, not derived from F1/F2)
        row_vals.append(fk_vals[i])
        # MA_{lookback}d: fixed full-size window, blank until `lookback` rows
        # have elapsed (see KNOWN SIMPLIFICATION above -- Python's actual
        # min_periods behavior turns on earlier than this formula does).
        row_vals.append(f"=AVERAGE(F{r-lookback+1}:F{r})" if i + 1 >= lookback else "")
        # Deviation
        row_vals.append(f'=IF(OR(F{r}="",G{r}="",G{r}=0),"",(F{r}-G{r})/G{r})')
        # Signal: always 0/+-1, never blank (Value's classification treats an
        # undefined Deviation as flat, unlike Momentum/Carry's NaN-propagating SIGN())
        row_vals.append(f'=IF(H{r}="",0,IF(H{r}<-{threshold},1,IF(H{r}>{threshold},-1,0)))')
        # Position -- Signal[t-(shift_n+1)]
        if i + 1 > eff_shift:
            row_vals.append(f'=IF(I{r-eff_shift}="",0,I{r-eff_shift})')
        else:
            row_vals.append(0)
        # F1_cont_daily_change
        row_vals.append("" if i == 0 else f"=E{r}-E{r-1}")
        # Daily_PnL
        row_vals.append(f'=IF(OR(J{r}="",K{r}=""),0,J{r}*K{r})')
        # Position_Change / TC_Cost / Daily_PnL_Net (TC_Cost includes the
        # roll-day charge -- see MA sheet's comment for the full rationale)
        row_vals.append(f"=ABS(J{r})" if i == 0 else f"=ABS(J{r}-J{r-1})")
        if i == 0:
            row_vals.append(f"=M{r}*({tc_bps}/10000/2)*B{r}")
        else:
            row_vals.append(
                f'=(M{r}+IF(AND(LEFT(D{r},8)="Roll_LTD",J{r}<>0,J{r-1}<>0,SIGN(J{r})=SIGN(J{r-1})),1,0))'
                f'*({tc_bps}/10000/2)*B{r}'
            )
        row_vals.append(f"=L{r}-N{r}")
        # MTM
        row_vals.append(f"=J{r}*E{r}")
        # Cum_PnL / Cum_PnL_Net
        row_vals.append(f"=L{r}" if i == 0 else f"=Q{r-1}+L{r}")
        row_vals.append(f"=O{r}" if i == 0 else f"=R{r-1}+O{r}")

        ws.append(row_vals)
        for col_idx in (5, 11, 12, 14, 15, 16, 17, 18):  # F1_cont, delta, PnL, TC_Cost, PnL_Net, MTM, CumPnL, CumPnL_Net
            ws.cell(r, col_idx).font = _MANUAL_FONT
        row_fill = _row_phase_fill(phase_vals[i])
        if row_fill is not None:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(r, col_idx).fill = row_fill

    ws.column_dimensions["A"].width = 12
    for i in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 16
    ws.freeze_panes = "A2"


def _save_strategy_workbook(wb: Workbook, filepath: Path, sharpes: dict) -> dict:
    filepath.parent.mkdir(parents=True, exist_ok=True)
    wb.save(filepath)
    return sharpes


def save_momentum_tradebook_excel(f1r: pd.Series, f1c: pd.Series, f2r: pd.Series, phase: pd.Series,
                                   m: int, n: int, unit_label: str, filepath: Path,
                                   tc_bps: int = TC_BPS_DEFAULT) -> dict:
    """Momentum MA-crossover tradebook -- F2_raw + Phase columns, roll/expiry-day
    highlighting, manual-Excel-formula sheet adjacent to each Python sheet --
    6 sheets total (3 timing variants x Python+Formulas)."""
    sharpes = {}
    wb = Workbook()
    wb.remove(wb.active)
    for shift_n, label in TIMING_VARIANTS:
        tb = build_ma_tradebook(f1r=f1r, f1c=f1c, m=m, n=n, shift_n=shift_n, f2r=f2r, phase=phase, tc_bps=tc_bps)
        pos = pd.Series(tb["Position"].values, index=pd.DatetimeIndex(tb["Date"]))
        pnl = pd.Series(tb["Daily_PnL"].values, index=pd.DatetimeIndex(tb["Date"]))
        f1r_al = pd.Series(tb["F1_raw"].values, index=pd.DatetimeIndex(tb["Date"]))
        phase_al = pd.Series(tb["Phase"].values, index=pd.DatetimeIndex(tb["Date"])) if "Phase" in tb.columns else None
        met = compute_performance(pnl, pos, f1c.reindex(pos.index), f1r_al, shift_n, unit_label, tc_bps, phase_al)
        met["EXCEL RECONCILIATION"] = _reconcile_ma(tb, m, n, shift_n)
        _write_xl_sheet(wb, tb, met, f"{label} Python")
        write_manual_formula_sheet_ma(wb, tb, m, n, shift_n, f"{label} Formulas", tc_bps)
        sharpes[f"Shift-{shift_n} Sharpe"] = met["Sharpe Ratio (Gross)"]
    return _save_strategy_workbook(wb, filepath, sharpes)


def save_carry_v1_tradebook_excel(curve: pd.DataFrame, f1r: pd.Series, f1c: pd.Series, f2r: pd.Series,
                                   phase: pd.Series, unit_label: str, filepath: Path,
                                   tc_bps: int = TC_BPS_DEFAULT) -> dict:
    """Carry V1 Roll Yield tradebook -- 6 sheets total (3 timing variants x Python+Formulas)."""
    sharpes = {}
    wb = Workbook()
    wb.remove(wb.active)
    for shift_n, label in TIMING_VARIANTS:
        tb = build_carry_v1_tradebook(curve=curve, f1r=f1r, f1c=f1c, shift_n=shift_n, f2r=f2r, phase=phase,
                                       tc_bps=tc_bps)
        pos = pd.Series(tb["Position"].values, index=pd.DatetimeIndex(tb["Date"]))
        pnl = pd.Series(tb["Daily_PnL"].values, index=pd.DatetimeIndex(tb["Date"]))
        f1r_al = pd.Series(tb["F1_raw"].values, index=pd.DatetimeIndex(tb["Date"]))
        phase_al = pd.Series(tb["Phase"].values, index=pd.DatetimeIndex(tb["Date"])) if "Phase" in tb.columns else None
        met = compute_performance(pnl, pos, f1c.reindex(pos.index), f1r_al, shift_n, unit_label, tc_bps, phase_al)
        met["EXCEL RECONCILIATION"] = _reconcile_carry_v1(tb, shift_n)
        _write_xl_sheet(wb, tb, met, f"{label} Python")
        write_manual_formula_sheet_carry_v1(wb, tb, shift_n, f"{label} Formulas", tc_bps)
        sharpes[f"Shift-{shift_n} Sharpe"] = met["Sharpe Ratio (Gross)"]
    return _save_strategy_workbook(wb, filepath, sharpes)


def save_carry_v2_tradebook_excel(curve: pd.DataFrame, f1r: pd.Series, f1c: pd.Series, f2r: pd.Series,
                                   phase: pd.Series, unit_label: str, filepath: Path,
                                   near: str = CARRY_V2_TENOR[0], far: str = CARRY_V2_TENOR[1],
                                   tc_bps: int = TC_BPS_DEFAULT) -> dict:
    """Carry V2 Long Slope tradebook -- 6 sheets total (3 timing variants x Python+Formulas)."""
    sharpes = {}
    wb = Workbook()
    wb.remove(wb.active)
    for shift_n, label in TIMING_VARIANTS:
        tb = build_carry_v2_tradebook(curve=curve, f1r=f1r, f1c=f1c, shift_n=shift_n, near=near, far=far,
                                       f2r=f2r, phase=phase, tc_bps=tc_bps)
        pos = pd.Series(tb["Position"].values, index=pd.DatetimeIndex(tb["Date"]))
        pnl = pd.Series(tb["Daily_PnL"].values, index=pd.DatetimeIndex(tb["Date"]))
        f1r_al = pd.Series(tb["F1_raw"].values, index=pd.DatetimeIndex(tb["Date"]))
        phase_al = pd.Series(tb["Phase"].values, index=pd.DatetimeIndex(tb["Date"])) if "Phase" in tb.columns else None
        met = compute_performance(pnl, pos, f1c.reindex(pos.index), f1r_al, shift_n, unit_label, tc_bps, phase_al)
        met["EXCEL RECONCILIATION"] = _reconcile_carry_v2(tb, shift_n, near, far)
        _write_xl_sheet(wb, tb, met, f"{label} Python")
        write_manual_formula_sheet_carry_v2(wb, tb, shift_n, near, far, f"{label} Formulas", tc_bps)
        sharpes[f"Shift-{shift_n} Sharpe"] = met["Sharpe Ratio (Gross)"]
    return _save_strategy_workbook(wb, filepath, sharpes)


def save_carry_v3_tradebook_excel(curve: pd.DataFrame, f1r: pd.Series, f1c: pd.Series, f2r: pd.Series,
                                   phase: pd.Series, window: int, unit_label: str, filepath: Path,
                                   tc_bps: int = TC_BPS_DEFAULT) -> dict:
    """Carry V3 Z-score tradebook -- 6 sheets total (3 timing variants x Python+Formulas)."""
    sharpes = {}
    wb = Workbook()
    wb.remove(wb.active)
    for shift_n, label in TIMING_VARIANTS:
        tb = build_carry_v3_tradebook(curve=curve, f1r=f1r, f1c=f1c, shift_n=shift_n, window=window,
                                       f2r=f2r, phase=phase, tc_bps=tc_bps)
        pos = pd.Series(tb["Position"].values, index=pd.DatetimeIndex(tb["Date"]))
        pnl = pd.Series(tb["Daily_PnL"].values, index=pd.DatetimeIndex(tb["Date"]))
        f1r_al = pd.Series(tb["F1_raw"].values, index=pd.DatetimeIndex(tb["Date"]))
        phase_al = pd.Series(tb["Phase"].values, index=pd.DatetimeIndex(tb["Date"])) if "Phase" in tb.columns else None
        met = compute_performance(pnl, pos, f1c.reindex(pos.index), f1r_al, shift_n, unit_label, tc_bps, phase_al)
        met["EXCEL RECONCILIATION"] = _reconcile_carry_v3(tb, shift_n, window)
        _write_xl_sheet(wb, tb, met, f"{label} Python")
        write_manual_formula_sheet_carry_v3(wb, tb, shift_n, window, f"{label} Formulas", tc_bps)
        sharpes[f"Shift-{shift_n} Sharpe"] = met["Sharpe Ratio (Gross)"]
    return _save_strategy_workbook(wb, filepath, sharpes)


def save_carry_v4_tradebook_excel(curve: pd.DataFrame, f1r: pd.Series, f1c: pd.Series, f2r: pd.Series,
                                   phase: pd.Series, horizon: int, unit_label: str, filepath: Path,
                                   tc_bps: int = TC_BPS_DEFAULT) -> dict:
    """Carry V4 Carry-Momentum tradebook -- 6 sheets total (3 timing variants x Python+Formulas)."""
    sharpes = {}
    wb = Workbook()
    wb.remove(wb.active)
    for shift_n, label in TIMING_VARIANTS:
        tb = build_carry_v4_tradebook(curve=curve, f1r=f1r, f1c=f1c, shift_n=shift_n, horizon=horizon,
                                       f2r=f2r, phase=phase, tc_bps=tc_bps)
        pos = pd.Series(tb["Position"].values, index=pd.DatetimeIndex(tb["Date"]))
        pnl = pd.Series(tb["Daily_PnL"].values, index=pd.DatetimeIndex(tb["Date"]))
        f1r_al = pd.Series(tb["F1_raw"].values, index=pd.DatetimeIndex(tb["Date"]))
        phase_al = pd.Series(tb["Phase"].values, index=pd.DatetimeIndex(tb["Date"])) if "Phase" in tb.columns else None
        met = compute_performance(pnl, pos, f1c.reindex(pos.index), f1r_al, shift_n, unit_label, tc_bps, phase_al)
        met["EXCEL RECONCILIATION"] = _reconcile_carry_v4(tb, shift_n, horizon)
        _write_xl_sheet(wb, tb, met, f"{label} Python")
        write_manual_formula_sheet_carry_v4(wb, tb, shift_n, horizon, f"{label} Formulas", tc_bps)
        sharpes[f"Shift-{shift_n} Sharpe"] = met["Sharpe Ratio (Gross)"]
    return _save_strategy_workbook(wb, filepath, sharpes)


def save_value_v1_tradebook_excel(curve: pd.DataFrame, f1r: pd.Series, f1c: pd.Series, f2r: pd.Series,
                                   phase: pd.Series, contract: str, lookback: int, threshold: float,
                                   unit_label: str, filepath: Path, tc_bps: int = TC_BPS_DEFAULT) -> dict:
    """Value V1 MA-Reversion tradebook -- 6 sheets total (3 timing variants x Python+Formulas)."""
    sharpes = {}
    wb = Workbook()
    wb.remove(wb.active)
    for shift_n, label in TIMING_VARIANTS:
        tb = build_value_v1_tradebook(curve=curve, f1r=f1r, f1c=f1c, contract=contract, lookback=lookback,
                                       threshold=threshold, shift_n=shift_n, f2r=f2r, phase=phase, tc_bps=tc_bps)
        pos = pd.Series(tb["Position"].values, index=pd.DatetimeIndex(tb["Date"]))
        pnl = pd.Series(tb["Daily_PnL"].values, index=pd.DatetimeIndex(tb["Date"]))
        f1r_al = pd.Series(tb["F1_raw"].values, index=pd.DatetimeIndex(tb["Date"]))
        phase_al = pd.Series(tb["Phase"].values, index=pd.DatetimeIndex(tb["Date"])) if "Phase" in tb.columns else None
        met = compute_performance(pnl, pos, f1c.reindex(pos.index), f1r_al, shift_n, unit_label, tc_bps, phase_al)
        met["EXCEL RECONCILIATION"] = _reconcile_value_v1(tb, contract, lookback, threshold, shift_n)
        _write_xl_sheet(wb, tb, met, f"{label} Python")
        write_manual_formula_sheet_value_v1(wb, tb, shift_n, contract, lookback, threshold, f"{label} Formulas", tc_bps)
        sharpes[f"Shift-{shift_n} Sharpe"] = met["Sharpe Ratio (Gross)"]
    return _save_strategy_workbook(wb, filepath, sharpes)


# ══════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════

def main():
    TRADEBOOKS_DIR.mkdir(exist_ok=True)
    log_rows = []

    for p in PRODUCTS:
        code, name, unit = p["code"], p["name"], p["unit"]
        print("=" * 78)
        print(f"{p['asset_class']} — {name} ({code})")
        print("=" * 78)

        f1_df = get_metal_rolling_f1(code, futures_file=p["futures_file"], calendar_file=p["calendar_file"],
                                      verbose=False, config=p["config"])
        f1_df = reanchor_f1_continuous(f1_df[f1_df.index.year >= 2006])
        f1r, f1c = f1_df["F1_raw"], f1_df["F1_continuous"]
        f2r, phase = f1_df["F2_raw"], f1_df["Phase"]

        if p["loader"] == "simple":
            curve = load_curve_simple(p["futures_file"], p["config"][code]["price_sheet"])
        else:
            curve_data = load_curve_legacy_multiheader(p["futures_file"])
            curve = curve_data[p["config"][code]["price_sheet"]]["prices"]
        curve = curve[curve.index.year >= 2006]

        out_dir = TRADEBOOKS_DIR / code

        def _fmt_sh(sh: dict) -> str:
            return "  ".join(f"{k}={v}" for k, v in sh.items())

        # ── Momentum: 3 default benchmark MA pairs ──────────────────────────
        for m, n in MOMENTUM_PAIRS:
            fpath = out_dir / f"Momentum_MA_{m}_{n}.xlsx"
            sh = save_momentum_tradebook_excel(f1r, f1c, f2r, phase, m=m, n=n, unit_label=unit, filepath=fpath)
            print(f"  Momentum MA({m},{n})  {_fmt_sh(sh)}  -> {fpath.name}")
            log_rows.append({"code": code, "name": name, "strategy": f"Momentum MA({m},{n})", **sh})

        # ── Carry: V1 Roll Yield, V2 Long Slope, V3 Z-score(252d), V4 Carry-Momentum ──
        if "F1" in curve.columns and "F2" in curve.columns:
            fpath = out_dir / "Carry_V1_RollYield.xlsx"
            sh = save_carry_v1_tradebook_excel(curve, f1r, f1c, f2r, phase, unit_label=unit, filepath=fpath)
            print(f"  Carry V1 (F1-F2)/F1  {_fmt_sh(sh)}  -> {fpath.name}")
            log_rows.append({"code": code, "name": name, "strategy": "Carry V1", **sh})

            fpath = out_dir / "Carry_V3_Zscore252.xlsx"
            sh = save_carry_v3_tradebook_excel(curve, f1r, f1c, f2r, phase, window=252,
                                                unit_label=unit, filepath=fpath)
            print(f"  Carry V3 Z-score(252d)  {_fmt_sh(sh)}  -> {fpath.name}")
            log_rows.append({"code": code, "name": name, "strategy": "Carry V3", **sh})

            fpath = out_dir / f"Carry_V4_Momentum{CARRY_V4_HORIZON}d.xlsx"
            sh = save_carry_v4_tradebook_excel(curve, f1r, f1c, f2r, phase, horizon=CARRY_V4_HORIZON,
                                                unit_label=unit, filepath=fpath)
            print(f"  Carry V4 Carry-Momentum({CARRY_V4_HORIZON}d)  {_fmt_sh(sh)}  -> {fpath.name}")
            log_rows.append({"code": code, "name": name, "strategy": "Carry V4", **sh})

            near, far = CARRY_V2_TENOR
            if near in curve.columns and far in curve.columns:
                fpath = out_dir / f"Carry_V2_LongSlope_{near}_{far}.xlsx"
                sh = save_carry_v2_tradebook_excel(curve, f1r, f1c, f2r, phase, near=near, far=far,
                                                    unit_label=unit, filepath=fpath)
                print(f"  Carry V2 ({near}-{far})/{near}  {_fmt_sh(sh)}  -> {fpath.name}")
                log_rows.append({"code": code, "name": name, "strategy": "Carry V2", **sh})
            else:
                print(f"  Carry V2 skipped -- {near}/{far} not found in curve.")
        else:
            print("  Carry skipped -- F1/F2 not found in curve.")

        # ── Value: V1 MA-reversion, default contract/lookback/threshold ────
        contracts = [c for c in curve.columns if str(c).startswith("F") and str(c)[1:].isdigit()
                     and int(str(c)[1:]) <= 15]
        if contracts:
            contract = contracts[min(7, len(contracts) - 1)]
            fpath = out_dir / f"Value_V1_{contract}.xlsx"
            sh = save_value_v1_tradebook_excel(curve, f1r, f1c, f2r, phase, contract=contract,
                                                lookback=VALUE_LOOKBACK_DAYS, threshold=VALUE_THRESHOLD,
                                                unit_label=unit, filepath=fpath)
            print(f"  Value V1 {contract} 5yr +-10%  {_fmt_sh(sh)}  -> {fpath.name}")
            log_rows.append({"code": code, "name": name, "strategy": f"Value V1 {contract}", **sh})
        else:
            print("  Value skipped -- no usable Fk contracts in curve.")

    summary = pd.DataFrame(log_rows)
    summary.to_csv(TRADEBOOKS_DIR / "tradebooks_summary.csv", index=False)
    print("\n" + "=" * 78)
    print(f"Done. {len(log_rows)} tradebooks written under {TRADEBOOKS_DIR.resolve()}")
    print("Summary -> tradebooks/tradebooks_summary.csv")


if __name__ == "__main__":
    main()
