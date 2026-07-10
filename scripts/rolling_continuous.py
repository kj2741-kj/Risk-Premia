"""
rolling_continuous.py
=====================
Generic continuous rolling F1 price-series builder for commodity futures.

Rolls N trading days before LAST_TRADEABLE_DT (configurable 1–10, default 5).

Rolling logic (return-based / additive stitching):
---------------------------------------------------
  F1_Tracking (normal days, holding front-month F1):
      F1_cont[t] = F1_cont[t-1] + (F1_raw[t] - F1_raw[t-1])

  Roll day (N trading days before LAST_TRADEABLE_DT):
      Switch from F1 to F2. Track F2 delta on roll day.
      F1_cont[t] = F1_cont[t-1] + (F2[t] - F2[t-1])

  F2_Tracking (between roll day and data-file contract switch):
      F1_cont[t] = F1_cont[t-1] + (F2[t] - F2[t-1])

  Bridge (day after LAST_TRADEABLE_DT, when the data file switches F1
  to the new contract):
      F1_cont[t] = F1_cont[t-1] + (F1[t] - F2[t-1])
      Then resume normal F1 tracking.

Calendar dates are snapped to the nearest prior trading day when they
don't appear in the price data (holiday mismatches, data gaps).

Usage:
------
    from rolling_continuous import get_metal_rolling_f1

    df = get_metal_rolling_f1("LP")                   # LME Copper, default 5 days before expiry
    df = get_metal_rolling_f1("LP", roll_day=3)       # roll 3 trading days before expiry
    df = get_metal_rolling_f1("CL", roll_day=7,       # WTI, 7 days before expiry
             config=ENERGY_CONFIG,
             futures_file=ENERGY_FUTURES_FILE,
             calendar_file=ENERGY_CALENDAR_FILE)
"""

from __future__ import annotations

import os
import numpy as np
import pandas as pd

_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_REPO_ROOT = os.path.dirname(_SCRIPT_DIR)
DATA_DIR = os.path.join(_REPO_ROOT, "data")
OUTPUTS_DIR = os.path.join(_REPO_ROOT, "outputs")

DEFAULT_ROLL_DAY = 5

# ── Default file paths (can be overridden at call time) ───────────────────────
DEFAULT_FUTURES_FILE  = os.path.join(DATA_DIR, "Metals Futures Curve.csv")
DEFAULT_CALENDAR_FILE = os.path.join(DATA_DIR, "expiry_calendars_20260526.xlsx")

# ── Metal configuration ───────────────────────────────────────────────────────
# Extend this dict to add more metals.
# f1_col / f2_col: 0-based column index in the price sheet (after iloc).
# data_start_row: first row of actual price data (0-based, after header rows).
METAL_CONFIG: dict[str, dict] = {
    "LP": {
        "name"          : "LME Copper",
        "price_sheet"   : "Copper LME",
        "calendar_sheet": "LP - LME Copper",
        "f1_col"        : 1,
        "f2_col"        : 4,
        "data_start_row": 3,
    },
    "LA": {
        "name"          : "LME Aluminium",
        "price_sheet"   : "ALuminium LME",
        "calendar_sheet": "LA - LME Aluminium",
        "f1_col"        : 1,
        "f2_col"        : 4,
        "data_start_row": 3,
    },
    "LX": {
        "name"          : "LME Zinc",
        "price_sheet"   : "Zinc LME",
        "calendar_sheet": "LX - LME Zinc",
        "f1_col"        : 1,
        "f2_col"        : 4,
        "data_start_row": 3,
    },
    "LN": {
        "name"          : "LME Nickel",
        "price_sheet"   : "Nickel LME",
        "calendar_sheet": "LN - LME Nickel",
        "f1_col"        : 1,
        "f2_col"        : 4,
        "data_start_row": 3,
    },
    "LL": {
        "name"          : "LME Lead",
        "price_sheet"   : "Lead LME",
        "calendar_sheet": "LL - LME Lead",
        "f1_col"        : 1,
        "f2_col"        : 4,
        "data_start_row": 3,
    },
    "LT": {
        "name"          : "LME Tin",
        "price_sheet"   : "Tin LME",
        "calendar_sheet": "LT - Custom (LT)",
        "f1_col"        : 1,
        "f2_col"        : 4,
        "data_start_row": 3,
    },
    "GC": {
        "name"          : "COMEX Gold",
        "price_sheet"   : "Gold COMEX",
        "calendar_sheet": "GC - COMEX Gold",
        "f1_col"        : 1,
        "f2_col"        : 4,
        "data_start_row": 3,
    },
    "SI": {
        "name"          : "COMEX Silver",
        "price_sheet"   : "Silver COMEX",
        "calendar_sheet": "SI - COMEX Silver",
        "f1_col"        : 1,
        "f2_col"        : 4,
        "data_start_row": 3,
    },
    "PL": {
        "name"          : "COMEX Platinum",
        "price_sheet"   : "Platinum COMEX",
        "calendar_sheet": "PL - COMEX Platinum",
        "f1_col"        : 1,
        "f2_col"        : 4,
        "data_start_row": 3,
    },
    "PA": {
        "name"          : "COMEX Palladium",
        "price_sheet"   : "Palladium COMEX",
        "calendar_sheet": "PA - COMEX Palladium",
        "f1_col"        : 1,
        "f2_col"        : 4,
        "data_start_row": 3,
    },
}

# ── Stage 2 configs (data/06-30/*.xlsx -- simple single-header-row format:
#    title row, then 'Date','F1','F2',... header row, data from row 2) ────────
# f1_col=1/f2_col=2/data_start_row=2 throughout since there's no Volume/OI
# interleaving in this newer format, unlike METAL_CONFIG's legacy layout above.
#
# IMPORTANT: use these WITH the matching futures_file/calendar_file overrides,
# e.g. get_metal_rolling_f1("CL", futures_file=ENERGY_FUTURES_FILE,
#                           calendar_file=ENERGY_CALENDAR_FILE)
# -- the DEFAULT_FUTURES_FILE/DEFAULT_CALENDAR_FILE module constants above are
# for METAL_CONFIG (Metals Futures Curve.csv) only.

ENERGY_FUTURES_FILE  = os.path.join(DATA_DIR, "06-30", "Energy_Futures_Updated.xlsx")
ENERGY_CALENDAR_FILE = os.path.join(DATA_DIR, "06-30", "expiry_calendars_20260701.xlsx")

# LME base metals, refreshed through 2026-06-30 (README sheet inside the
# workbook confirms "End Date 20260630"). Supersedes METAL_CONFIG's LP/LA/LX/
# LN/LL/LT entries above, which point at the stale "Metals Futures Curve.csv"
# (Copper LME sheet there stops 2025-12-31; the rest stop ~2026-05-19/20).
# Paired with the matching 2026-07-01-vintage calendar (same file Energy uses)
# rather than the older expiry_calendars_20260526.xlsx, whose LME sheet names
# ("LP - LME Copper") don't match this calendar's naming ("LP - Copper (LME)").
METALS_FUTURES_FILE  = os.path.join(DATA_DIR, "06-30", "Metals_Futures_Curve_Updated.xlsx")
METALS_CALENDAR_FILE = os.path.join(DATA_DIR, "06-30", "expiry_calendars_20260701.xlsx")

METALS_CONFIG: dict[str, dict] = {
    "LP": {
        "name": "LME Copper", "price_sheet": "Copper LME",
        "calendar_sheet": "LP - Copper (LME)",
        "f1_col": 1, "f2_col": 2, "data_start_row": 2,
    },
    "LA": {
        "name": "LME Aluminium", "price_sheet": "Aluminium LME",
        "calendar_sheet": "LA - Aluminium (LME)",
        "f1_col": 1, "f2_col": 2, "data_start_row": 2,
    },
    "LX": {
        "name": "LME Zinc", "price_sheet": "Zinc LME",
        "calendar_sheet": "LX - Zinc (LME)",
        "f1_col": 1, "f2_col": 2, "data_start_row": 2,
    },
    "LN": {
        "name": "LME Nickel", "price_sheet": "Nickel LME",
        "calendar_sheet": "LN - Nickel (LME)",
        "f1_col": 1, "f2_col": 2, "data_start_row": 2,
    },
    "LL": {
        "name": "LME Lead", "price_sheet": "Lead LME",
        "calendar_sheet": "LL - Lead (LME)",
        "f1_col": 1, "f2_col": 2, "data_start_row": 2,
    },
    "LT": {
        "name": "LME Tin", "price_sheet": "Tin LME",
        "calendar_sheet": "LT - Tin (LME)",
        "f1_col": 1, "f2_col": 2, "data_start_row": 2,
    },
}

ENERGY_CONFIG: dict[str, dict] = {
    "CL": {
        "name": "WTI Crude (NYMEX)", "price_sheet": "WTI Crude (NYMEX)",
        "calendar_sheet": "CL - WTI Crude Oil (NYMEX)",
        "f1_col": 1, "f2_col": 2, "data_start_row": 2,
    },
    "CO": {
        "name": "Brent Crude (ICE)", "price_sheet": "Brent Crude (ICE)",
        "calendar_sheet": "CO - Brent Crude Oil (ICE)",
        "f1_col": 1, "f2_col": 2, "data_start_row": 2,
    },
    "XB": {
        "name": "RBOB Gasoline (NYMEX)", "price_sheet": "RBOB Gasoline (NYMEX)",
        "calendar_sheet": "XB - RBOB Gasoline (NYMEX)",
        "f1_col": 1, "f2_col": 2, "data_start_row": 2,
    },
    "HO": {
        "name": "Heating Oil ULSD (NYMEX)", "price_sheet": "Heating Oil ULSD (NYMEX)",
        "calendar_sheet": "HO - ULSD - Heating Oil (NYMEX)",
        "f1_col": 1, "f2_col": 2, "data_start_row": 2,
    },
    "NG": {
        "name": "Nat Gas Henry Hub (NYMEX)", "price_sheet": "Nat Gas Henry Hub (NYMEX)",
        "calendar_sheet": "NG - Natural Gas Henry Hub (NYM",  # sheet name genuinely truncated at 31 chars
        "f1_col": 1, "f2_col": 2, "data_start_row": 2,
    },
    "QS": {
        "name": "Singapore Gasoil (ICE)", "price_sheet": "Singapore Gasoil (ICE)",
        "calendar_sheet": "QS - Gasoil (ICE)",
        "f1_col": 1, "f2_col": 2, "data_start_row": 2,
    },
    "FO": {
        "name": "Fuel Oil 3.5pct Barges (ICE)", "price_sheet": "Fuel Oil 3.5pct Barges (ICE)",
        "calendar_sheet": "FO - Fuel Oil 3.5% Barges FOB R",
        "f1_col": 1, "f2_col": 2, "data_start_row": 2,
    },
    "SJ": {
        "name": "Singapore Jet Kerosene (ICE)", "price_sheet": "Singapore Jet Kerosene (ICE)",
        "calendar_sheet": "SJ - Jet-Kerosene Cargoes CIF N",
        "f1_col": 1, "f2_col": 2, "data_start_row": 2,
    },
    "NFY": {
        "name": "Naphtha CIF NWE Platts (ICE)", "price_sheet": "Naphtha CIF NWE Platts (ICE)",
        "calendar_sheet": "NFY - Naphtha cif NWE Cargoes (",
        "f1_col": 1, "f2_col": 2, "data_start_row": 2,
    },
    # NOTE: "GO" (ICE Gasoil London) is in the price file but has NO matching
    # sheet in expiry_calendars_20260701.xlsx -- cannot build F1_continuous for
    # it without a roll calendar. Excluded until a calendar is available.
}

PRECIOUS_FUTURES_FILE  = os.path.join(DATA_DIR, "06-30", "Precious_Metals_Futures_Updated.xlsx")
PRECIOUS_CALENDAR_FILE = os.path.join(DATA_DIR, "06-30", "expiry_calendars_20260701.xlsx")

PRECIOUS_CONFIG: dict[str, dict] = {
    "GC": {
        "name": "Gold COMEX", "price_sheet": "Gold COMEX",
        "calendar_sheet": "GC - Gold (COMEX)",
        "f1_col": 1, "f2_col": 2, "data_start_row": 2,
    },
    "SI": {
        "name": "Silver COMEX", "price_sheet": "Silver COMEX",
        "calendar_sheet": "SI - Silver (COMEX)",
        "f1_col": 1, "f2_col": 2, "data_start_row": 2,
    },
    "HG": {
        "name": "Copper CME (HG)", "price_sheet": "Copper CME (HG)",
        "calendar_sheet": "HG - Copper (COMEX)",
        "f1_col": 1, "f2_col": 2, "data_start_row": 2,
    },
    "PL": {
        "name": "Platinum NYMEX", "price_sheet": "Platinum NYMEX",
        "calendar_sheet": "PL - Platinum (NYMEX)",
        "f1_col": 1, "f2_col": 2, "data_start_row": 2,
    },
    "PA": {
        "name": "Palladium NYMEX", "price_sheet": "Palladium NYMEX",
        "calendar_sheet": "PA - Palladium (NYMEX)",
        "f1_col": 1, "f2_col": 2, "data_start_row": 2,
    },
}

NGL_FUTURES_FILE  = os.path.join(DATA_DIR, "06-30", "NGL_Futures_Updated.xlsx")
NGL_CALENDAR_FILE = os.path.join(DATA_DIR, "06-30", "expiry_calendars_20260701.xlsx")

# Calendar sheet names for NGL/petrochemical tickers use an em dash (—,
# not a hyphen) and are truncated at Excel's 31-char sheet-name limit -- copied
# verbatim from expiry_calendars_20260701.xlsx's own sheet names, which
# independently confirm the CAP=Ethane/BAP=Propane/DAE=Butane/PCW=Ethylene
# mapping used to correct NGL_Futures_Updated.xlsx's price-sheet names
# on 2026-07-10 (see that workbook's README "CORRECTION NOTE").
NGL_CONFIG: dict[str, dict] = {
    "CAP": {
        "name": "Ethane (Mt Belvieu)", "price_sheet": "Ethane Argus",
        "calendar_sheet": "CAP - Ethane — Mt Belvieu Swap ",
        "f1_col": 1, "f2_col": 2, "data_start_row": 2,
    },
    "BAP": {
        "name": "Propane (Mt Belvieu)", "price_sheet": "Propane Argus",
        "calendar_sheet": "BAP - Propane — Mt Belvieu Swap",
        "f1_col": 1, "f2_col": 2, "data_start_row": 2,
    },
    "DAE": {
        "name": "Butane (Mt Belvieu)", "price_sheet": "Butane Argus",
        "calendar_sheet": "DAE - Butane — Mt Belvieu Swap ",
        "f1_col": 1, "f2_col": 2, "data_start_row": 2,
    },
    "IBD": {
        "name": "Isobutane (Mt Belvieu)", "price_sheet": "Isobutane Argus",
        "calendar_sheet": "IBD - Isobutane — Mt Belvieu Sw",
        "f1_col": 1, "f2_col": 2, "data_start_row": 2,
    },
    "PCW": {
        "name": "Ethylene (Mt Belvieu)", "price_sheet": "Ethylene Argus",
        "calendar_sheet": "PCW - Ethylene — Mt Belvieu Fut",
        "f1_col": 1, "f2_col": 2, "data_start_row": 2,
    },
    "PGP": {
        "name": "Propylene (Polymer Grade)", "price_sheet": "Propylene Argus",
        "calendar_sheet": "PGP - Propylene — Polymer Grade",
        "f1_col": 1, "f2_col": 2, "data_start_row": 2,
    },
}


# ─────────────────────────────────────────────────────────────────────────────
# 1. Data loaders
# ─────────────────────────────────────────────────────────────────────────────

def load_metal_prices(
    filepath: str,
    sheet_name: str,
    f1_col: int = 1,
    f2_col: int = 4,
    data_start_row: int = 3,
) -> pd.DataFrame:
    """
    Load F1 and F2 raw closing prices from the futures curve workbook.

    Parameters
    ----------
    filepath       : Path to the futures curve file (xlsx or xlsx disguised as csv).
    sheet_name     : Worksheet name, e.g. "Copper LME".
    f1_col         : 0-based column index for F1 price (default 1).
    f2_col         : 0-based column index for F2 price (default 4).
    data_start_row : First row of price data, 0-based (default 3, skipping 3 header rows).

    Returns
    -------
    DataFrame indexed by Date with columns: F1_raw, F2_raw.
    """
    raw   = pd.read_excel(filepath, sheet_name=sheet_name, header=None)
    dates = pd.to_datetime(raw.iloc[data_start_row:, 0], errors="coerce")
    f1    = pd.to_numeric(raw.iloc[data_start_row:, f1_col], errors="coerce")
    f2    = pd.to_numeric(raw.iloc[data_start_row:, f2_col], errors="coerce")

    df = pd.DataFrame({"Date": dates.values, "F1_raw": f1.values, "F2_raw": f2.values})
    df["Date"] = pd.to_datetime(df["Date"])
    df = (df.dropna(subset=["Date"])
            .set_index("Date")
            .sort_index())
    return df


def load_metal_calendar(
    filepath: str,
    sheet_name: str,
) -> pd.DataFrame:
    """
    Load roll and expiry calendar for a metal.

    Uses LAST_TRADEABLE_DT as the roll date (holiday-adjusted) and
    FUT_DLV_DT_LAST as the expiry anchor.

    Returns
    -------
    DataFrame with columns: Contract, roll_date, expiry_date  (both pd.Timestamp).
    """
    cal = pd.read_excel(filepath, sheet_name=sheet_name)
    cal["roll_date"]   = pd.to_datetime(cal["LAST_TRADEABLE_DT"], errors="coerce")
    cal["expiry_date"] = pd.to_datetime(cal["FUT_DLV_DT_LAST"],   errors="coerce")
    cal = (cal.dropna(subset=["roll_date", "expiry_date"])
              .sort_values("roll_date")
              .reset_index(drop=True))
    return cal[["Contract", "roll_date", "expiry_date"]]


# ─────────────────────────────────────────────────────────────────────────────
# 2. Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _snap_to_trading_day(dt: pd.Timestamp, sorted_trading: list[pd.Timestamp]) -> pd.Timestamp | None:
    """Return dt if it's a trading day, else the nearest prior trading day."""
    dt = dt.normalize()
    for t in reversed(sorted_trading):
        if t <= dt:
            return t
    return None


def _compute_roll_dates(
    calendar: pd.DataFrame,
    sorted_trading: list[pd.Timestamp],
    price_date_set: set[pd.Timestamp],
    n_days_before: int,
) -> tuple[dict[pd.Timestamp, str], set[pd.Timestamp], dict[pd.Timestamp, str]]:
    """
    For each contract, compute the roll date = N trading days before
    (snapped) LAST_TRADEABLE_DT.

    Returns
    -------
    roll_map       : {roll_date: contract_name}
    ltd_set        : set of snapped LTD dates (for bridge detection)
    ltd_to_contract: {snapped_ltd: contract_name}
    """
    roll_map = {}
    ltd_set = set()
    ltd_to_contract = {}

    for _, r in calendar.iterrows():
        ltd_raw = r["roll_date"].normalize()

        # Snap LTD to nearest prior trading day if not in price data
        snapped_ltd = _snap_to_trading_day(ltd_raw, sorted_trading)
        if snapped_ltd is None:
            continue

        ltd_set.add(snapped_ltd)
        ltd_to_contract[snapped_ltd] = r["Contract"]

        # Walk back N trading days from the snapped LTD
        ltd_idx = None
        for j, t in enumerate(sorted_trading):
            if t == snapped_ltd:
                ltd_idx = j
                break
        if ltd_idx is None:
            continue

        roll_idx = ltd_idx - n_days_before
        if roll_idx < 0:
            continue

        roll_date = sorted_trading[roll_idx]
        roll_map[roll_date] = r["Contract"]

    return roll_map, ltd_set, ltd_to_contract


# ─────────────────────────────────────────────────────────────────────────────
# 3. Core rolling algorithm
# ─────────────────────────────────────────────────────────────────────────────

def build_rolling_f1(
    prices: pd.DataFrame,
    calendar: pd.DataFrame,
    roll_day: int = DEFAULT_ROLL_DAY,
) -> pd.DataFrame:
    """
    Build a continuous rolling F1 price series using return-based stitching.
    Rolls N trading days before each LAST_TRADEABLE_DT.

    Parameters
    ----------
    prices   : DataFrame indexed by Date with columns F1_raw, F2_raw.
    calendar : DataFrame with columns Contract, roll_date, expiry_date.
    roll_day : How many trading days before LAST_TRADEABLE_DT to roll (1–10, default 5).

    Returns
    -------
    DataFrame with columns:
        F1_raw, F2_raw, F1_continuous,
        Phase, is_roll_date, is_bridge_date, active_contract
    """
    if not 1 <= roll_day <= 10:
        raise ValueError(f"roll_day must be 1–10, got {roll_day}")

    trading_dates = prices.index.normalize()
    price_date_set = set(trading_dates)
    sorted_trading = sorted(price_date_set)

    roll_map, ltd_set, ltd_to_contract = _compute_roll_dates(
        calendar, sorted_trading, price_date_set, n_days_before=roll_day,
    )
    roll_set = set(roll_map.keys())

    f1_arr = prices["F1_raw"].values
    f2_arr = prices["F2_raw"].values
    n = len(trading_dates)

    f1_cont = np.full(n, np.nan)
    phase_labels = np.full(n, "", dtype=object)
    is_roll = np.zeros(n, dtype=bool)
    is_bridge = np.zeros(n, dtype=bool)
    active_cont = np.full(n, "", dtype=object)

    in_f2_phase = False
    prev_f1 = np.nan
    prev_f2 = np.nan
    current_contract = "—"

    for i, d in enumerate(trading_dates):
        f1v = f1_arr[i]
        f2v = f2_arr[i]

        # ── Bridge detection: day after LAST_TRADEABLE_DT ───────────────
        # Fires when the calendar says the contract has rolled, regardless
        # of NaN state — must end F2 phase every month.
        if in_f2_phase:
            yesterday = trading_dates[i - 1] if i > 0 else None
            if yesterday is not None and yesterday in ltd_set:
                f1_prev = f1_cont[i - 1] if i > 0 else np.nan
                if not np.isnan(f1v) and not np.isnan(prev_f2) and not np.isnan(f1_prev):
                    f1_cont[i] = f1_prev + (f1v - prev_f2)
                elif not np.isnan(f1v) and not np.isnan(prev_f1) and not np.isnan(f1_prev):
                    f1_cont[i] = f1_prev + (f1v - prev_f1)
                elif not np.isnan(f1_prev):
                    f1_cont[i] = f1_prev
                else:
                    f1_cont[i] = f1v if not np.isnan(f1v) else 0.0

                phase_labels[i] = "Bridge"
                is_bridge[i] = True
                if yesterday in ltd_to_contract:
                    current_contract = ltd_to_contract[yesterday]
                active_cont[i] = current_contract
                prev_f1 = f1v
                prev_f2 = f2v
                in_f2_phase = False
                continue

        # ── Roll day (N trading days before LTD) ───────────────────────
        if d in roll_set:
            f1_prev = f1_cont[i - 1] if i > 0 else np.nan
            if i == 0 or np.isnan(f1_prev):
                f1_cont[i] = f1v
            elif not np.isnan(f2v) and not np.isnan(prev_f2):
                f1_cont[i] = f1_prev + (f2v - prev_f2)
            elif not np.isnan(f1v) and not np.isnan(prev_f1):
                f1_cont[i] = f1_prev + (f1v - prev_f1)
            else:
                f1_cont[i] = f1_prev

            phase_labels[i] = f"Roll_LTD-{roll_day}"
            is_roll[i] = True
            active_cont[i] = roll_map[d]
            current_contract = roll_map[d]
            prev_f1 = f1v
            prev_f2 = f2v
            in_f2_phase = True
            continue

        # ── F2 tracking ─────────────────────────────────────────────────
        if in_f2_phase:
            f1_prev = f1_cont[i - 1] if i > 0 else np.nan
            if not np.isnan(f2v) and not np.isnan(prev_f2) and not np.isnan(f1_prev):
                f1_cont[i] = f1_prev + (f2v - prev_f2)
            elif not np.isnan(f1_prev):
                f1_cont[i] = f1_prev
            else:
                f1_cont[i] = f1v

            phase_labels[i] = "F2_Tracking"
            active_cont[i] = current_contract
            prev_f1 = f1v
            prev_f2 = f2v
            continue

        # ── Normal F1 tracking ──────────────────────────────────────────
        if i == 0 or np.isnan(f1_cont[i - 1]):
            f1_cont[i] = f1v
        elif not np.isnan(f1v) and not np.isnan(prev_f1):
            f1_cont[i] = f1_cont[i - 1] + (f1v - prev_f1)
        else:
            f1_cont[i] = f1_cont[i - 1]

        phase_labels[i] = "F1_Tracking"
        active_cont[i] = current_contract
        prev_f1 = f1v
        prev_f2 = f2v

    out = prices[["F1_raw", "F2_raw"]].copy()
    out["F1_continuous"] = f1_cont
    out["Phase"] = phase_labels
    out["is_roll_date"] = is_roll
    out["is_bridge_date"] = is_bridge
    out["active_contract"] = active_cont
    return out


# ─────────────────────────────────────────────────────────────────────────────
# 4. High-level wrapper
# ─────────────────────────────────────────────────────────────────────────────

def get_metal_rolling_f1(
    metal_code: str,
    futures_file: str  = DEFAULT_FUTURES_FILE,
    calendar_file: str = DEFAULT_CALENDAR_FILE,
    verbose: bool      = True,
    config: dict | None = None,
    roll_day: int = DEFAULT_ROLL_DAY,
) -> pd.DataFrame:
    """
    End-to-end loader + builder for any configured product.

    Parameters
    ----------
    metal_code    : Key from `config` (defaults to METAL_CONFIG), e.g. "LP" for
                    LME Copper, or "CL" with config=ENERGY_CONFIG for WTI.
    futures_file  : Path to the futures curve workbook.
    calendar_file : Path to the expiry calendars workbook.
    verbose       : Print progress messages if True.
    config        : Which config registry to look `metal_code` up in. Defaults
                    to METAL_CONFIG. MUST be passed explicitly for Energy/
                    Precious Stage-2 codes -- several codes (GC, SI, PL, PA)
                    exist in BOTH METAL_CONFIG (legacy Metals Futures Curve.csv,
                    f1_col/f2_col=1/4) and PRECIOUS_CONFIG (data/06-30 simple
                    format, f1_col/f2_col=1/2) with different column layouts,
                    so relying on a default here would silently read the wrong
                    columns for one of the two callers.
    roll_day      : How many trading days before LAST_TRADEABLE_DT to roll
                    (1–10, default 5).

    Returns
    -------
    DataFrame with columns: F1_raw, F2_raw, F1_continuous, Phase,
                            is_roll_date, is_bridge_date, active_contract.
    """
    if config is None:
        config = METAL_CONFIG
    if metal_code not in config:
        raise ValueError(
            f"Unknown code '{metal_code}' for this config. "
            f"Available: {list(config.keys())}"
        )

    cfg = config[metal_code]
    if verbose:
        print(f"[{metal_code}] Loading prices from sheet '{cfg['price_sheet']}' ...")

    prices = load_metal_prices(
        filepath       = futures_file,
        sheet_name     = cfg["price_sheet"],
        f1_col         = cfg["f1_col"],
        f2_col         = cfg["f2_col"],
        data_start_row = cfg["data_start_row"],
    )

    if verbose:
        print(f"[{metal_code}] {len(prices)} trading days "
              f"({prices.index[0].date()} -> {prices.index[-1].date()})")
        print(f"[{metal_code}] Loading calendar from sheet '{cfg['calendar_sheet']}' ...")

    cal = load_metal_calendar(
        filepath   = calendar_file,
        sheet_name = cfg["calendar_sheet"],
    )

    if verbose:
        print(f"[{metal_code}] {len(cal)} contracts in calendar")
        print(f"[{metal_code}] Building rolling F1 (LTD-{roll_day}) ...")

    result = build_rolling_f1(prices, cal, roll_day=roll_day)

    if verbose:
        rolls   = result["is_roll_date"].sum()
        bridges = result["is_bridge_date"].sum()
        print(f"[{metal_code}] Done: {rolls} roll events, {bridges} bridge events")

    return result


def reanchor_f1_continuous(f1_df: pd.DataFrame) -> pd.DataFrame:
    """Re-anchors F1_continuous so its FIRST row equals F1_raw's first row.

    build_rolling_f1() anchors F1_continuous to F1_raw on the first day of
    the FULL underlying price history (which may start years before any
    analysis window we actually use), then every subsequent value is a
    running sum of day-over-day deltas. Once a caller slices the result to a
    shorter analysis window (e.g. `f1_df[f1_df.index.year >= 2006]`), that
    original anchor point is no longer the first visible row -- so
    F1_continuous[first visible day] silently carries forward whatever net
    roll-adjustment accumulated over ALL the sliced-off history before it,
    and no longer equals F1_raw[first visible day].

    This is purely a display/level convention -- PnL and Sharpe only ever
    use day-over-day deltas (a constant shift cancels exactly in diff()), so
    this has zero effect on any backtest result. It exists so the series is
    anchored to the actual analysis window's first day (matching what the
    manual Excel-formula tradebooks reconstruct from Phase+F1_raw+F2_raw,
    which have no visibility into pre-window history and so anchor there by
    construction) -- call this immediately after slicing a get_metal_rolling_f1
    (or get_rolling_f1_5td) result to a start date, in every caller, so the
    dashboard and the tradebook generator never show two different
    F1_continuous levels for the same window.

    Anchors off the first row where BOTH F1_continuous and F1_raw are
    non-NaN, not blindly row 0 -- a handful of products (e.g. the NGL
    tickers) have 1-2 missing F1 quotes at the very start of their sliced
    window, and iloc[0] landing on one of those would make offset NaN,
    which then poisons every row (anything minus NaN is NaN)."""
    f1_df = f1_df.copy()
    valid = f1_df["F1_continuous"].notna() & f1_df["F1_raw"].notna()
    if not valid.any():
        return f1_df
    anchor = f1_df.loc[valid].iloc[0]
    offset = anchor["F1_continuous"] - anchor["F1_raw"]
    f1_df["F1_continuous"] = f1_df["F1_continuous"] - offset
    return f1_df


# ─────────────────────────────────────────────────────────────────────────────
# 5. Generate verification Excel files when run directly
# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import argparse
    from pathlib import Path
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    parser = argparse.ArgumentParser(description="Generate verification Excel files for rolling continuous series.")
    parser.add_argument("--roll-day", type=int, default=DEFAULT_ROLL_DAY,
                        choices=range(1, 11), metavar="N",
                        help="Roll N trading days before LAST_TRADEABLE_DT (1–10, default 5)")
    args = parser.parse_args()
    rd = args.roll_day

    OUTPUT_DIR = Path(_REPO_ROOT) / f"verification_ltd{rd}"
    OUTPUT_DIR.mkdir(exist_ok=True)

    ALL_PRODUCTS = [
        ("LP", "LME_Copper",         METAL_CONFIG,    DEFAULT_FUTURES_FILE,  DEFAULT_CALENDAR_FILE),
        ("LA", "LME_Aluminium",      METAL_CONFIG,    DEFAULT_FUTURES_FILE,  DEFAULT_CALENDAR_FILE),
        ("CL", "WTI_Crude",          ENERGY_CONFIG,   ENERGY_FUTURES_FILE,   ENERGY_CALENDAR_FILE),
        ("CO", "Brent_Crude",        ENERGY_CONFIG,   ENERGY_FUTURES_FILE,   ENERGY_CALENDAR_FILE),
        ("XB", "RBOB_Gasoline",      ENERGY_CONFIG,   ENERGY_FUTURES_FILE,   ENERGY_CALENDAR_FILE),
        ("HO", "Heating_Oil",        ENERGY_CONFIG,   ENERGY_FUTURES_FILE,   ENERGY_CALENDAR_FILE),
        ("NG", "Nat_Gas",            ENERGY_CONFIG,   ENERGY_FUTURES_FILE,   ENERGY_CALENDAR_FILE),
        ("QS", "Singapore_Gasoil",   ENERGY_CONFIG,   ENERGY_FUTURES_FILE,   ENERGY_CALENDAR_FILE),
        ("FO", "Fuel_Oil",           ENERGY_CONFIG,   ENERGY_FUTURES_FILE,   ENERGY_CALENDAR_FILE),
        ("GC", "Gold_COMEX",         PRECIOUS_CONFIG, PRECIOUS_FUTURES_FILE, PRECIOUS_CALENDAR_FILE),
        ("SI", "Silver_COMEX",       PRECIOUS_CONFIG, PRECIOUS_FUTURES_FILE, PRECIOUS_CALENDAR_FILE),
        ("HG", "Copper_CME",         PRECIOUS_CONFIG, PRECIOUS_FUTURES_FILE, PRECIOUS_CALENDAR_FILE),
        ("PL", "Platinum_NYMEX",     PRECIOUS_CONFIG, PRECIOUS_FUTURES_FILE, PRECIOUS_CALENDAR_FILE),
        ("PA", "Palladium_NYMEX",    PRECIOUS_CONFIG, PRECIOUS_FUTURES_FILE, PRECIOUS_CALENDAR_FILE),
        ("CAP", "Ethane",            NGL_CONFIG,      NGL_FUTURES_FILE,      NGL_CALENDAR_FILE),
        ("BAP", "Propane",           NGL_CONFIG,      NGL_FUTURES_FILE,      NGL_CALENDAR_FILE),
        ("DAE", "Butane",            NGL_CONFIG,      NGL_FUTURES_FILE,      NGL_CALENDAR_FILE),
        ("IBD", "Isobutane",         NGL_CONFIG,      NGL_FUTURES_FILE,      NGL_CALENDAR_FILE),
        ("PCW", "Ethylene",          NGL_CONFIG,      NGL_FUTURES_FILE,      NGL_CALENDAR_FILE),
        ("PGP", "Propylene",         NGL_CONFIG,      NGL_FUTURES_FILE,      NGL_CALENDAR_FILE),
    ]

    HEADER_FILL  = PatternFill("solid", fgColor="2B3A47")
    HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
    ROLL_FILL    = PatternFill("solid", fgColor="FFF3CD")   # yellow
    BRIDGE_FILL  = PatternFill("solid", fgColor="D4EDDA")   # green
    F2_FILL      = PatternFill("solid", fgColor="E8EAF6")   # blue

    for code, label, config, futures_file, calendar_file in ALL_PRODUCTS:
        name = config[code]["name"]
        print(f"\n{'='*70}")
        print(f"{name} ({code}) — roll LTD-{rd}")
        print(f"{'='*70}")

        df = get_metal_rolling_f1(code, config=config, futures_file=futures_file,
                                  calendar_file=calendar_file, verbose=True, roll_day=rd)

        rolls   = df["is_roll_date"].sum()
        bridges = df["is_bridge_date"].sum()
        print(f"\n  F1_raw     : {df['F1_raw'].iloc[0]:.2f} -> {df['F1_raw'].iloc[-1]:.2f}")
        print(f"  F1_cont    : {df['F1_continuous'].iloc[0]:.2f} -> {df['F1_continuous'].iloc[-1]:.2f}")
        print(f"  Diff (end) : {df['F1_continuous'].iloc[-1] - df['F1_raw'].iloc[-1]:.2f}")
        print(f"  Days <= 0  : {(df['F1_continuous'] <= 0).sum()}")
        print(f"  Rolls={rolls}, Bridges={bridges}")

        out_path = OUTPUT_DIR / f"{code}_{label}_LTD-{rd}_rolling.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = f"{code} Continuous"

        cols = ["Date", "F1_raw", "F2_raw", "F1_continuous", "Phase",
                "is_roll_date", "is_bridge_date", "active_contract",
                "dF1", "dF2", "dF1_cont"]
        ws.append(cols)
        for c in range(1, len(cols) + 1):
            cell = ws.cell(1, c)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")

        df_out = df.reset_index()
        df_out["dF1"]      = df_out["F1_raw"].diff()
        df_out["dF2"]      = df_out["F2_raw"].diff()
        df_out["dF1_cont"] = df_out["F1_continuous"].diff()

        for _, row in df_out.iterrows():
            vals = [
                row["Date"].strftime("%Y-%m-%d") if hasattr(row["Date"], "strftime") else str(row["Date"]),
                round(row["F1_raw"], 4) if pd.notna(row["F1_raw"]) else None,
                round(row["F2_raw"], 4) if pd.notna(row["F2_raw"]) else None,
                round(row["F1_continuous"], 4) if pd.notna(row["F1_continuous"]) else None,
                row["Phase"],
                row["is_roll_date"],
                row["is_bridge_date"],
                row["active_contract"],
                round(row["dF1"], 4) if pd.notna(row["dF1"]) else None,
                round(row["dF2"], 4) if pd.notna(row["dF2"]) else None,
                round(row["dF1_cont"], 4) if pd.notna(row["dF1_cont"]) else None,
            ]
            ws.append(vals)
            r = ws.max_row
            if row["is_roll_date"]:
                for c in range(1, len(cols) + 1):
                    ws.cell(r, c).fill = ROLL_FILL
            elif row["is_bridge_date"]:
                for c in range(1, len(cols) + 1):
                    ws.cell(r, c).fill = BRIDGE_FILL
            elif row["Phase"] == "F2_Tracking":
                for c in range(1, len(cols) + 1):
                    ws.cell(r, c).fill = F2_FILL

        for c in range(1, len(cols) + 1):
            ws.column_dimensions[get_column_letter(c)].width = 16
        ws.freeze_panes = "A2"

        wb.save(out_path)
        print(f"  Saved -> {out_path}")

    print(f"\nDone. All {len(ALL_PRODUCTS)} files in {OUTPUT_DIR.resolve()}")
