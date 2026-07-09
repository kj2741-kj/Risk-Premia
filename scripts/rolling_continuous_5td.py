"""
rolling_continuous_5td.py
==========================
Continuous rolling F1 price-series builder — rolls on the Nth trading day
of every month (configurable, default N=5).

Rolling logic (return-based / additive stitching):
---------------------------------------------------
On every trading day the continuous series advances by the daily change
of whichever raw contract the strategy is currently "holding":

  Normal days (holding front-month F1):
      F1_cont[t] = F1_cont[t-1] + (F1_raw[t] - F1_raw[t-1])

  Roll day (Nth trading day of the month):
      We switch from the old F1 to the new F2.
      F1_cont[t] = F1_cont[t-1] + (F2_raw[t] - F2_raw[t-1])
      (Track F2's move on the roll day itself — the position
       is transferred at yesterday's close, so today's PnL
       comes from the new contract.)

  F2 Tracking (between roll day and data-file contract switch):
      F1_cont[t] = F1_cont[t-1] + (F2[t] - F2[t-1])

  Bridge (day after LAST_TRADEABLE_DT, when the data file switches F1
  to the new contract):
      F1_cont[t] = F1_cont[t-1] + (F1[t] - F2[t-1])
      Then resume normal F1 tracking.

Usage:
------
    from rolling_continuous_5td import get_rolling_f1

    df = get_rolling_f1("LP", roll_day=5)             # LME Copper, 5th TD
    df = get_rolling_f1("CL", roll_day=3,             # WTI, 3rd TD
             config=ENERGY_CONFIG,
             futures_file=ENERGY_FUTURES_FILE,
             calendar_file=ENERGY_CALENDAR_FILE)
"""

from __future__ import annotations

import os
import numpy as np
import pandas as pd

_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_REPO_ROOT = os.path.dirname(_SCRIPT_DIR)
DATA_DIR = os.path.join(_REPO_ROOT, "data")

from rolling_continuous import (
    load_metal_prices, load_metal_calendar,
    METAL_CONFIG, DEFAULT_FUTURES_FILE, DEFAULT_CALENDAR_FILE,
    ENERGY_CONFIG, ENERGY_FUTURES_FILE, ENERGY_CALENDAR_FILE,
    PRECIOUS_CONFIG, PRECIOUS_FUTURES_FILE, PRECIOUS_CALENDAR_FILE,
    METALS_CONFIG, METALS_FUTURES_FILE, METALS_CALENDAR_FILE,
)

DEFAULT_ROLL_DAY = 5


def _nth_trading_days(trading_dates: pd.DatetimeIndex, n: int) -> set[pd.Timestamp]:
    """Return the set of Nth-trading-day-of-month dates."""
    result = set()
    current_ym = None
    count = 0
    for d in sorted(trading_dates):
        ym = (d.year, d.month)
        if ym != current_ym:
            current_ym = ym
            count = 0
        count += 1
        if count == n:
            result.add(d.normalize())
    return result


def build_rolling_f1(
    prices: pd.DataFrame,
    calendar: pd.DataFrame,
    roll_day: int = DEFAULT_ROLL_DAY,
) -> pd.DataFrame:
    """
    Build a continuous rolling F1 price series using the Nth trading day
    of each month as the roll trigger.

    Parameters
    ----------
    prices   : DataFrame indexed by Date with columns F1_raw, F2_raw.
    calendar : DataFrame with columns Contract, roll_date, expiry_date.
               roll_date = LAST_TRADEABLE_DT (used only to detect when the
               data file switches F1 from old to new contract).
    roll_day : Which trading day of the month to roll on (1–10, default 5).

    Returns
    -------
    DataFrame with columns:
        F1_raw, F2_raw, F1_continuous,
        Phase, is_roll_date, is_bridge_date, active_contract
    """
    if not 1 <= roll_day <= 10:
        raise ValueError(f"roll_day must be 1–10, got {roll_day}")

    trading_dates = prices.index.normalize()
    price_date_set = set(trading_dates)
    roll_set = _nth_trading_days(trading_dates, roll_day)

    # Snap each LAST_TRADEABLE_DT to the nearest prior trading day when the
    # exact calendar date isn't in the price data (holiday mismatch, data gap).
    sorted_trading = sorted(price_date_set)
    ltd_set = set()
    ltd_to_contract = {}
    for _, r in calendar.iterrows():
        ltd = r["roll_date"].normalize()
        if ltd in price_date_set:
            snapped = ltd
        else:
            candidates = [t for t in sorted_trading if t < ltd]
            snapped = candidates[-1] if candidates else None
        if snapped is not None:
            ltd_set.add(snapped)
            ltd_to_contract[snapped] = r["Contract"]

    f1_arr = prices["F1_raw"].values
    f2_arr = prices["F2_raw"].values
    n = len(trading_dates)

    f1_cont = np.full(n, np.nan)
    phase_labels = np.full(n, "", dtype=object)
    is_roll = np.zeros(n, dtype=bool)
    is_bridge = np.zeros(n, dtype=bool)
    active_cont = np.full(n, "", dtype=object)

    in_f2_phase = False
    prev_f1 = np.nan
    prev_f2 = np.nan
    current_contract = "—"

    for i, d in enumerate(trading_dates):
        f1v = f1_arr[i]
        f2v = f2_arr[i]

        # ── Bridge detection: day after LAST_TRADEABLE_DT ───────────────
        # Always fires when the calendar says the contract has rolled,
        # regardless of NaN state — must end F2 phase every month.
        if in_f2_phase:
            yesterday = trading_dates[i - 1] if i > 0 else None
            if yesterday is not None and yesterday in ltd_set:
                f1_prev = f1_cont[i - 1] if i > 0 else np.nan
                if not np.isnan(f1v) and not np.isnan(prev_f2) and not np.isnan(f1_prev):
                    f1_cont[i] = f1_prev + (f1v - prev_f2)
                elif not np.isnan(f1v) and not np.isnan(prev_f1) and not np.isnan(f1_prev):
                    f1_cont[i] = f1_prev + (f1v - prev_f1)
                elif not np.isnan(f1_prev):
                    f1_cont[i] = f1_prev
                else:
                    f1_cont[i] = f1v if not np.isnan(f1v) else 0.0

                phase_labels[i] = "Bridge"
                is_bridge[i] = True
                if yesterday in ltd_to_contract:
                    current_contract = ltd_to_contract[yesterday]
                active_cont[i] = current_contract
                prev_f1 = f1v
                prev_f2 = f2v
                in_f2_phase = False
                continue

        # ── Roll day (Nth trading day of month) ─────────────────────────
        if d in roll_set:
            f1_prev = f1_cont[i - 1] if i > 0 else np.nan
            if i == 0 or np.isnan(f1_prev):
                f1_cont[i] = f1v
            elif not np.isnan(f2v) and not np.isnan(prev_f2):
                f1_cont[i] = f1_prev + (f2v - prev_f2)
            elif not np.isnan(f1v) and not np.isnan(prev_f1):
                f1_cont[i] = f1_prev + (f1v - prev_f1)
            else:
                f1_cont[i] = f1_prev

            phase_labels[i] = f"Roll_TD{roll_day}"
            is_roll[i] = True
            active_cont[i] = current_contract
            prev_f1 = f1v
            prev_f2 = f2v
            in_f2_phase = True
            continue

        # ── F2 tracking ─────────────────────────────────────────────────
        if in_f2_phase:
            f1_prev = f1_cont[i - 1] if i > 0 else np.nan
            if not np.isnan(f2v) and not np.isnan(prev_f2) and not np.isnan(f1_prev):
                f1_cont[i] = f1_prev + (f2v - prev_f2)
            elif not np.isnan(f1_prev):
                f1_cont[i] = f1_prev
            else:
                f1_cont[i] = f1v

            phase_labels[i] = "F2_Tracking"
            active_cont[i] = current_contract
            prev_f1 = f1v
            prev_f2 = f2v
            continue

        # ── Normal F1 tracking ──────────────────────────────────────────
        if i == 0 or np.isnan(f1_cont[i - 1]):
            f1_cont[i] = f1v
        elif not np.isnan(f1v) and not np.isnan(prev_f1):
            f1_cont[i] = f1_cont[i - 1] + (f1v - prev_f1)
        else:
            f1_cont[i] = f1_cont[i - 1]

        phase_labels[i] = "F1_Tracking"
        active_cont[i] = current_contract
        prev_f1 = f1v
        prev_f2 = f2v

    out = prices[["F1_raw", "F2_raw"]].copy()
    out["F1_continuous"] = f1_cont
    out["Phase"] = phase_labels
    out["is_roll_date"] = is_roll
    out["is_bridge_date"] = is_bridge
    out["active_contract"] = active_cont
    return out


def get_rolling_f1(
    metal_code: str,
    futures_file: str = DEFAULT_FUTURES_FILE,
    calendar_file: str = DEFAULT_CALENDAR_FILE,
    verbose: bool = True,
    config: dict | None = None,
    roll_day: int = DEFAULT_ROLL_DAY,
) -> pd.DataFrame:
    """
    End-to-end loader + builder for any configured product.

    Parameters
    ----------
    metal_code    : Key from config, e.g. "LP", "CL", "GC".
    futures_file  : Path to the futures curve workbook.
    calendar_file : Path to the expiry calendars workbook.
    verbose       : Print progress messages if True.
    config        : Config registry (defaults to METAL_CONFIG).
    roll_day      : Which trading day of the month to roll on (1–10, default 5).
    """
    if config is None:
        config = METAL_CONFIG
    if metal_code not in config:
        raise ValueError(f"Unknown code '{metal_code}'. Available: {list(config.keys())}")

    cfg = config[metal_code]
    if verbose:
        print(f"[{metal_code}] Loading prices from sheet '{cfg['price_sheet']}' ...")

    prices = load_metal_prices(
        filepath=futures_file,
        sheet_name=cfg["price_sheet"],
        f1_col=cfg["f1_col"],
        f2_col=cfg["f2_col"],
        data_start_row=cfg["data_start_row"],
    )

    if verbose:
        print(f"[{metal_code}] {len(prices)} trading days "
              f"({prices.index[0].date()} -> {prices.index[-1].date()})")

    cal = load_metal_calendar(filepath=calendar_file, sheet_name=cfg["calendar_sheet"])

    if verbose:
        print(f"[{metal_code}] {len(cal)} contracts in calendar")
        print(f"[{metal_code}] Building TD{roll_day} rolling F1 ...")

    result = build_rolling_f1(prices, cal, roll_day=roll_day)

    if verbose:
        rolls = result["is_roll_date"].sum()
        bridges = result["is_bridge_date"].sum()
        print(f"[{metal_code}] Done: {rolls} roll events, {bridges} bridge events")

    return result


# Keep the old name as an alias for backward compatibility
get_rolling_f1_5td = get_rolling_f1


# ─────────────────────────────────────────────────────────────────────────────
# Main: generate verification Excel files
# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import argparse
    from pathlib import Path
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    parser = argparse.ArgumentParser(description="Generate verification Excel files for Nth-TD rolling.")
    parser.add_argument("--roll-day", type=int, default=DEFAULT_ROLL_DAY,
                        choices=range(1, 11), metavar="N",
                        help="Roll on the Nth trading day of each month (1–10, default 5)")
    args = parser.parse_args()
    rd = args.roll_day

    OUTPUT_DIR = Path(_REPO_ROOT) / f"verification_td{rd}"
    OUTPUT_DIR.mkdir(exist_ok=True)

    PRODUCTS_TO_VERIFY = [
        ("LP", "LME Copper", METAL_CONFIG, DEFAULT_FUTURES_FILE, DEFAULT_CALENDAR_FILE),
        ("CL", "WTI Crude", ENERGY_CONFIG, ENERGY_FUTURES_FILE, ENERGY_CALENDAR_FILE),
        ("GC", "Gold COMEX", PRECIOUS_CONFIG, PRECIOUS_FUTURES_FILE, PRECIOUS_CALENDAR_FILE),
    ]

    HEADER_FILL = PatternFill("solid", fgColor="2B3A47")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
    ROLL_FILL = PatternFill("solid", fgColor="FFF3CD")
    BRIDGE_FILL = PatternFill("solid", fgColor="D4EDDA")
    F2_FILL = PatternFill("solid", fgColor="E8EAF6")

    for code, name, config, futures_file, calendar_file in PRODUCTS_TO_VERIFY:
        print(f"\n{'='*70}")
        print(f"{name} ({code}) — roll on TD{rd}")
        print(f"{'='*70}")

        df = get_rolling_f1(code, config=config, futures_file=futures_file,
                            calendar_file=calendar_file, verbose=True, roll_day=rd)

        print(f"\n  F1_raw     : {df['F1_raw'].iloc[0]:.2f} -> {df['F1_raw'].iloc[-1]:.2f}")
        print(f"  F1_cont    : {df['F1_continuous'].iloc[0]:.2f} -> {df['F1_continuous'].iloc[-1]:.2f}")
        print(f"  Diff (end) : {df['F1_continuous'].iloc[-1] - df['F1_raw'].iloc[-1]:.2f}")
        print(f"  Days <= 0  : {(df['F1_continuous'] <= 0).sum()}")
        print(f"  Phase dist : {df['Phase'].value_counts().to_dict()}")

        out_path = OUTPUT_DIR / f"{code}_{name.replace(' ', '_')}_TD{rd}_rolling.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = f"{code} Continuous"

        cols = ["Date", "F1_raw", "F2_raw", "F1_continuous", "Phase",
                "is_roll_date", "is_bridge_date", "active_contract",
                "dF1", "dF2", "dF1_cont"]
        ws.append(cols)
        for c in range(1, len(cols) + 1):
            cell = ws.cell(1, c)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")

        df_out = df.reset_index()
        df_out["dF1"] = df_out["F1_raw"].diff()
        df_out["dF2"] = df_out["F2_raw"].diff()
        df_out["dF1_cont"] = df_out["F1_continuous"].diff()

        for _, row in df_out.iterrows():
            vals = [
                row["Date"].strftime("%Y-%m-%d") if hasattr(row["Date"], "strftime") else str(row["Date"]),
                round(row["F1_raw"], 4) if not np.isnan(row["F1_raw"]) else None,
                round(row["F2_raw"], 4) if not np.isnan(row["F2_raw"]) else None,
                round(row["F1_continuous"], 4) if not np.isnan(row["F1_continuous"]) else None,
                row["Phase"],
                row["is_roll_date"],
                row["is_bridge_date"],
                row["active_contract"],
                round(row["dF1"], 4) if pd.notna(row["dF1"]) else None,
                round(row["dF2"], 4) if pd.notna(row["dF2"]) else None,
                round(row["dF1_cont"], 4) if pd.notna(row["dF1_cont"]) else None,
            ]
            ws.append(vals)
            r = ws.max_row
            if row["is_roll_date"]:
                for c in range(1, len(cols) + 1):
                    ws.cell(r, c).fill = ROLL_FILL
            elif row["is_bridge_date"]:
                for c in range(1, len(cols) + 1):
                    ws.cell(r, c).fill = BRIDGE_FILL
            elif row["Phase"] == "F2_Tracking":
                for c in range(1, len(cols) + 1):
                    ws.cell(r, c).fill = F2_FILL

        for c in range(1, len(cols) + 1):
            ws.column_dimensions[get_column_letter(c)].width = 16
        ws.freeze_panes = "A2"

        wb.save(out_path)
        print(f"  Saved -> {out_path}")

    print(f"\nDone. Verification files in {OUTPUT_DIR.resolve()}")
